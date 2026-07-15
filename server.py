from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from datetime import date, datetime, time as datetime_time, timedelta
from zoneinfo import ZoneInfo
from urllib.parse import parse_qs, urlencode, urlparse
from io import BytesIO
import base64
import json
import html
import hashlib
import hmac
import os
import random
import queue
import re
import secrets
import ssl
import time
import threading
import unicodedata
import urllib.error
import urllib.request
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed


ROOT = Path(__file__).parent
PUBLIC = ROOT / "public"
DATA = Path(os.getenv("COMPETIDOR_DATA_DIR", str(ROOT / "data"))).expanduser().resolve()
DATA.mkdir(parents=True, exist_ok=True)
CERTS = DATA / "certs"
CERTS.mkdir(exist_ok=True)
APP_DATA_FILE = "app.json"
CATALOG_DATA_FILE = "catalog.json"
SYNC_PROGRESS_FILE = "sync_progress.json"
SYNC_LOCK = threading.Lock()
DATA_LOCK = threading.RLock()
ACTIVE_SYNC_ACCOUNTS = set()
SYNC_PROGRESS_LOCK = threading.RLock()
SYNC_PROGRESS = {}
SYNC_PROGRESS_DATA_DIR = ""
SYNC_PROGRESS_LAST_PERSIST = 0.0
SYNC_WORKER_SEMAPHORE = threading.BoundedSemaphore(
    max(1, min(2, int(os.getenv("MELI_SYNC_CONCURRENT_ACCOUNTS", "1"))))
)
MELI_NOTIFICATION_QUEUE = queue.Queue(maxsize=5000)
CATEGORY_ATTRIBUTES_LOCK = threading.RLock()
CATEGORY_ATTRIBUTES_CACHE = {}
SELLER_PROFILE_LOCK = threading.RLock()
SELLER_PROFILE_CACHE = {}
OFFICIAL_STORE_CACHE_LOCK = threading.RLock()
OFFICIAL_STORE_CACHE = {}
PUBLIC_BUYBOX_CACHE_LOCK = threading.RLock()
PUBLIC_BUYBOX_CACHE = {}
CATALOG_OFFERS_CACHE_LOCK = threading.RLock()
CATALOG_OFFERS_CACHE = {}
STATISTICS_CACHE_LOCK = threading.RLock()
STATISTICS_CACHE = {}
STATISTICS_JOBS_LOCK = threading.RLock()
STATISTICS_JOBS = {}
REPORT_JOBS_LOCK = threading.RLock()
REPORT_JOBS = {}
SPREADSHEET_JOBS_LOCK = threading.RLock()
SPREADSHEET_JOBS = {}
SHIPMENT_MODE_CACHE_LOCK = threading.RLock()
SHIPMENT_MODE_CACHE = {}
CLONE_SOURCE_CACHE_LOCK = threading.RLock()
CLONE_SOURCE_CACHE = {}
ASYNC_OPERATION_JOBS_LOCK = threading.RLock()
ASYNC_OPERATION_JOBS = {}
ACTIVE_CLONE_OPERATIONS = set()
CLONE_WORKER_SEMAPHORE = threading.BoundedSemaphore(
    max(1, min(3, int(os.getenv("MELI_CLONE_CONCURRENT_JOBS", "2"))))
)

MELI_AUTH_URL = "https://auth.mercadolivre.com.br/authorization"
MELI_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
MELI_API_URL = "https://api.mercadolibre.com"
TELEGRAM_API_URL = "https://api.telegram.org"
APP_TZ = ZoneInfo(os.getenv("APP_TIMEZONE", "America/Sao_Paulo"))
SESSION_SECONDS = 60 * 60 * 12
SENSITIVE_MELI_PATH_TERMS = (
    "mercadopago",
    "mercado_pago",
    "/payments",
    "/payment",
    "/billing",
    "/bank",
    "/cards",
    "/wallet",
    "/mp/",
    "/money",
)
ALLOWED_MELI_PATHS = (
    re.compile(r"^/users/me$"),
    re.compile(r"^/users/[^/]+$"),
    re.compile(r"^/users/[^/]+/shipping_options/free(\?|$)"),
    re.compile(r"^/users/[^/]+/items/search(\?|$)"),
    re.compile(r"^/items[^/]*(\?|$)"),
    re.compile(r"^/items/[^/]+(\?|$)"),
    re.compile(r"^/items/[^/]+/description(\?|$)"),
    re.compile(r"^/items/[^/]+/price_to_win(\?|$)"),
    re.compile(r"^/categories/[^/]+/attributes(\?|$)"),
    re.compile(r"^/products/[^/]+(\?|$)"),
    re.compile(r"^/products/[^/]+/items(\?|$)"),
    re.compile(r"^/user-products/[^/]+(\?|$)"),
    re.compile(r"^/orders/search(\?|$)"),
    re.compile(r"^/shipments/[^/]+(\?|$)"),
    re.compile(r"^/shipments/[^/]+/sla(\?|$)"),
    re.compile(r"^/claims/search(\?|$)"),
    re.compile(r"^/v1/claims/search(\?|$)"),
    re.compile(r"^/post-purchase/v1/claims/search(\?|$)"),
    re.compile(r"^/post-purchase/v1/claims/[^/]+/detail(\?|$)"),
)


def meli_flag(value):
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() == "true"


def is_catalog_listing(item):
    """Only catalog_listing identifies an offer that actually competes in catalog."""
    return meli_flag((item or {}).get("catalog_listing"))


def read_json(name, fallback):
    with DATA_LOCK:
        path = DATA / name
        if not path.exists():
            write_json(name, fallback)
        return json.loads(path.read_text(encoding="utf-8"))


def write_json(name, payload):
    with DATA_LOCK:
        path = DATA / name
        temporary = DATA / f".{name}.{uuid.uuid4().hex}.tmp"
        compact = name == CATALOG_DATA_FILE
        temporary.write_text(
            json.dumps(
                payload,
                indent=None if compact else 2,
                separators=(",", ":") if compact else None,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        os.replace(temporary, path)


def ensure_sync_progress_loaded():
    """Load the small durable sync registry once for the active data directory."""
    global SYNC_PROGRESS_DATA_DIR
    data_dir = str(DATA)
    with SYNC_PROGRESS_LOCK:
        if SYNC_PROGRESS_DATA_DIR == data_dir:
            return
        path = DATA / SYNC_PROGRESS_FILE
        try:
            stored = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
        except (OSError, ValueError, TypeError):
            stored = {}
        SYNC_PROGRESS.clear()
        if isinstance(stored, dict):
            SYNC_PROGRESS.update(stored)
        SYNC_PROGRESS_DATA_DIR = data_dir


def persist_sync_progress(force=False):
    global SYNC_PROGRESS_LAST_PERSIST
    ensure_sync_progress_loaded()
    now = time.time()
    with SYNC_PROGRESS_LOCK:
        if not force and now - SYNC_PROGRESS_LAST_PERSIST < 1.0:
            return
        write_json(SYNC_PROGRESS_FILE, SYNC_PROGRESS)
        SYNC_PROGRESS_LAST_PERSIST = now


def set_sync_progress(account_id, values, force=False):
    account_id = str(account_id or "")
    ensure_sync_progress_loaded()
    with SYNC_PROGRESS_LOCK:
        current = SYNC_PROGRESS.get(account_id, {})
        SYNC_PROGRESS[account_id] = {**current, **values}
    persist_sync_progress(force=force)
    return SYNC_PROGRESS.get(account_id, {})


def sync_progress_snapshot():
    ensure_sync_progress_loaded()
    with SYNC_PROGRESS_LOCK:
        return json.loads(json.dumps(SYNC_PROGRESS, ensure_ascii=False))


def read_payload(include_catalog=True):
    with DATA_LOCK:
        payload = read_json(APP_DATA_FILE, empty_payload())
        embedded_catalog = payload.pop("catalog", None)
        catalog_path = DATA / CATALOG_DATA_FILE
        if embedded_catalog is not None:
            if embedded_catalog or not catalog_path.exists():
                write_json(CATALOG_DATA_FILE, embedded_catalog)
            migration_payload = {**payload, "catalog": embedded_catalog}
            payload["catalog_counts_snapshot"] = catalog_counts(embedded_catalog)
            payload["operations_snapshot"] = build_operations(migration_payload)
            write_json(APP_DATA_FILE, payload)
        payload["catalog"] = read_json(CATALOG_DATA_FILE, []) if include_catalog else []
        if include_catalog:
            for item in payload["catalog"]:
                if item.get("winner_source") in {
                    "catalog_lowest_active_offer",
                    "catalog_reference",
                    "products_items_winner_marker",
                    "public_purchase_options",
                    "public_product_page",
                }:
                    item.update(competition_snapshot(item))
        payload["_catalog_loaded"] = bool(include_catalog)
        payload.setdefault("_revision", 0)
        return payload


def record_key(collection, record):
    if collection == "users":
        return str(record.get("id") or (record.get("email") or "").lower())
    if collection == "accounts":
        return str(record.get("seller_id") or record.get("id") or record.get("nickname") or "")
    return str(record.get("id") or "")


def merge_clone_jobs(incoming, latest):
    """Keep previews created while an older background writer was running."""
    incoming_by_id = {str(row.get("id")): row for row in incoming or [] if row.get("id")}
    latest_by_id = {str(row.get("id")): row for row in latest or [] if row.get("id")}
    ordered_ids = [*latest_by_id, *[key for key in incoming_by_id if key not in latest_by_id]]
    merged = []
    status_rank = {
        "preview_ready": 1,
        "review_required": 2,
        "partial_error": 3,
        "error": 3,
        "copied": 4,
    }
    for job_id in ordered_ids:
        current = incoming_by_id.get(job_id)
        saved = latest_by_id.get(job_id)
        if current is None or saved is None:
            merged.append(current or saved)
            continue
        current_rank = status_rank.get(current.get("status"), 0)
        saved_rank = status_rank.get(saved.get("status"), 0)
        winner = current if current_rank > saved_rank else saved
        merged.append({**current, **saved, **winner})
    return merged[:500]


def merge_critical_records(collection, incoming, latest):
    merged = []
    incoming_by_key = {record_key(collection, row): row for row in incoming or [] if record_key(collection, row)}
    latest_by_key = {record_key(collection, row): row for row in latest or [] if record_key(collection, row)}
    for key in [*latest_by_key.keys(), *[key for key in incoming_by_key if key not in latest_by_key]]:
        current = incoming_by_key.get(key)
        saved = latest_by_key.get(key)
        if current is None:
            merged.append(saved)
            continue
        if saved is None:
            merged.append(current)
            continue
        if collection == "accounts":
            current_sync = str(current.get("last_sync") or current.get("sync_finished_at") or "")
            saved_sync = str(saved.get("last_sync") or saved.get("sync_finished_at") or "")
            row = {**current, **saved} if saved_sync >= current_sync else {**saved, **current}
            token_source = current if int(current.get("token_created_at") or 0) > int(saved.get("token_created_at") or 0) else saved
            for field in ("access_token", "refresh_token", "expires_in", "token_created_at", "status", "official"):
                if field in token_source:
                    row[field] = token_source[field]
            merged.append(row)
        else:
            merged.append({**current, **saved})
    return merged


CATALOG_COMPETITION_FIELDS = {
    "competition_status", "competition_consistent", "competition_checked_at", "competition_reason",
    "winner_item_id", "winner_seller_id", "winner_name", "winner_price", "winner_confirmed", "winner_source",
    "catalog_reference_seller_id", "catalog_reference_name", "catalog_reference_price", "price_to_win",
    "current_price", "visit_share", "competitors_sharing_first_place", "runner_up_item_id",
    "runner_up_seller_id", "runner_up_name", "runner_up_price", "runner_up_source",
    "internal_competition", "internal_winner_account_id", "share", "status", "action",
}
CATALOG_ITEM_FIELDS = {
    "title", "thumbnail", "sku", "brand", "gtin", "variation_count", "catalog_product_id",
    "catalog_listing", "listing_type_id", "shipping_logistic_type", "shipping_mode", "free_shipping",
    "package_weight", "package_height", "package_width", "package_length", "price", "stock",
    "meli_status", "permalink", "item_data_checked_at",
}


def merge_catalog_records(incoming, latest):
    def key(row):
        return f"{row.get('account_id') or row.get('account') or ''}:{row.get('id') or ''}"

    incoming_by_key = {key(row): row for row in incoming or [] if row.get("id")}
    latest_by_key = {key(row): row for row in latest or [] if row.get("id")}
    merged_rows = []
    for row_key in [*latest_by_key, *[value for value in incoming_by_key if value not in latest_by_key]]:
        current = incoming_by_key.get(row_key)
        saved = latest_by_key.get(row_key)
        if current is None:
            merged_rows.append(saved)
            continue
        if saved is None:
            merged_rows.append(current)
            continue
        merged = {**saved, **current}
        competition_source = current if str(current.get("competition_checked_at") or "") >= str(saved.get("competition_checked_at") or "") else saved
        item_source = current if str(current.get("item_data_checked_at") or "") >= str(saved.get("item_data_checked_at") or "") else saved
        for field in CATALOG_COMPETITION_FIELDS:
            if field in competition_source:
                merged[field] = competition_source[field]
        for field in CATALOG_ITEM_FIELDS:
            if field in item_source:
                merged[field] = item_source[field]
        merged_rows.append(merged)
    return merged_rows


def write_payload(payload, replace_collections=None):
    replace_collections = set(replace_collections or [])
    with DATA_LOCK:
        path = DATA / APP_DATA_FILE
        latest = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
        latest.pop("catalog", None)
        incoming_revision = int(payload.get("_revision") or 0)
        latest_revision = int(latest.get("_revision") or 0)
        if incoming_revision < latest_revision:
            for collection in ("users", "accounts"):
                if collection not in replace_collections:
                    payload[collection] = merge_critical_records(
                        collection,
                        payload.get(collection, []),
                        latest.get(collection, []),
                    )
            latest_notifications = latest.get("user_notifications") or {}
            payload["user_notifications"] = {**(payload.get("user_notifications") or {}), **latest_notifications}
            payload["clone_jobs"] = merge_clone_jobs(
                payload.get("clone_jobs", []),
                latest.get("clone_jobs", []),
            )
            if payload.get("_catalog_loaded", True):
                payload["catalog"] = merge_catalog_records(
                    payload.get("catalog", []),
                    read_json(CATALOG_DATA_FILE, []),
                )
        payload["_revision"] = latest_revision + 1
        catalog_loaded = bool(payload.get("_catalog_loaded", True))
        if catalog_loaded:
            write_json(CATALOG_DATA_FILE, payload.get("catalog", []))
            payload["catalog_counts_snapshot"] = catalog_counts(payload.get("catalog", []))
            payload["operations_snapshot"] = build_operations(payload)
        metadata = dict(payload)
        metadata.pop("catalog", None)
        metadata.pop("_catalog_loaded", None)
        write_json(APP_DATA_FILE, metadata)


def catalog_counts(catalog):
    return {
        "total": len(catalog or []),
        "winning": len([item for item in catalog or [] if item.get("status") == "winning"]),
        "losing": len([item for item in catalog or [] if item.get("status") == "losing"]),
        "sharing": len([item for item in catalog or [] if item.get("status") == "sharing"]),
        "paused": len(
            [
                item
                for item in catalog or []
                if item.get("status") == "paused" or item.get("meli_status") == "paused"
            ]
        ),
    }


def now_label():
    return datetime.now(APP_TZ).strftime("%Y-%m-%d %H:%M")


def parse_meli_datetime(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=APP_TZ)
    return parsed.astimezone(APP_TZ)


def dispatch_time_left(value, reference=None):
    deadline = parse_meli_datetime(value)
    if not deadline:
        return "-"
    current = reference or datetime.now(APP_TZ)
    seconds = int((deadline - current).total_seconds())
    if seconds <= 0:
        overdue = abs(seconds)
        hours, remainder = divmod(overdue, 3600)
        minutes = remainder // 60
        return f"Atrasado {hours}h {minutes:02d}min"
    days, remainder = divmod(seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes = remainder // 60
    return f"{days}d {hours}h {minutes:02d}min" if days else f"{hours}h {minutes:02d}min"


def brl_label(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "não informado pela API"
    formatted = f"{number:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {formatted}"


def https_port():
    return int(os.getenv("HTTPS_PORT", str(int(os.getenv("PORT", "8765")) + 1)))


def ensure_dev_certificate():
    cert_path = CERTS / "localhost.crt"
    key_path = CERTS / "localhost.key"
    if cert_path.exists() and key_path.exists():
        return cert_path, key_path

    from cryptography import x509
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa
    from cryptography.x509.oid import NameOID
    import datetime
    import ipaddress

    key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    subject = issuer = x509.Name(
        [
            x509.NameAttribute(NameOID.COUNTRY_NAME, "BR"),
            x509.NameAttribute(NameOID.ORGANIZATION_NAME, "CompeTIDOR local"),
            x509.NameAttribute(NameOID.COMMON_NAME, "127.0.0.1"),
        ]
    )
    cert = (
        x509.CertificateBuilder()
        .subject_name(subject)
        .issuer_name(issuer)
        .public_key(key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(datetime.datetime.utcnow() - datetime.timedelta(days=1))
        .not_valid_after(datetime.datetime.utcnow() + datetime.timedelta(days=365))
        .add_extension(
            x509.SubjectAlternativeName(
                [
                    x509.DNSName("localhost"),
                    x509.IPAddress(ipaddress.ip_address("127.0.0.1")),
                ]
            ),
            critical=False,
        )
        .sign(key, hashes.SHA256())
    )
    cert_path.write_bytes(cert.public_bytes(serialization.Encoding.PEM))
    key_path.write_bytes(
        key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.TraditionalOpenSSL,
            encryption_algorithm=serialization.NoEncryption(),
        )
    )
    return cert_path, key_path


def app_settings():
    settings = read_json("settings.json", {"meli": {}})
    return settings


def meli_config():
    settings = app_settings().get("meli", {})
    return {
        "client_id": os.getenv("MELI_CLIENT_ID") or settings.get("client_id", ""),
        "client_secret": os.getenv("MELI_CLIENT_SECRET") or settings.get("client_secret", ""),
        "redirect_uri": os.getenv("MELI_REDIRECT_URI") or settings.get("redirect_uri", ""),
    }


def app_configured():
    config = meli_config()
    return all([config["client_id"], config["client_secret"], config["redirect_uri"]])


def oauth_issues():
    issues = []
    config = meli_config()
    if not config["client_id"]:
        issues.append("MELI_CLIENT_ID ausente")
    if not config["client_secret"]:
        issues.append("MELI_CLIENT_SECRET ausente")
    if not config["redirect_uri"]:
        issues.append("MELI_REDIRECT_URI ausente")
    redirect_uri = config["redirect_uri"]
    if redirect_uri and not redirect_uri.startswith(("http://", "https://")):
        issues.append("MELI_REDIRECT_URI precisa comecar com http:// ou https://")
    return issues


def public_account(account):
    clean = dict(account)
    clean.pop("access_token", None)
    clean.pop("refresh_token", None)
    return clean


def hash_password(password, salt=None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120_000)
    return salt, digest.hex()


def verify_password(password, user):
    salt = user.get("password_salt", "")
    expected = user.get("password_hash", "")
    if not salt or not expected:
      return False
    _, candidate = hash_password(password, salt)
    return hmac.compare_digest(candidate, expected)


def public_user(user):
    clean = dict(user)
    clean.pop("password_hash", None)
    clean.pop("password_salt", None)
    return clean


def is_master(user):
    return (user or {}).get("role") == "master"


def can_manage_users(user):
    return (user or {}).get("role") in {"master", "admin"}


def visible_users_for(actor, users):
    visible = users if is_master(actor) else [user for user in users if user.get("role") != "master"]
    return [public_user(user) for user in visible]


def public_notifications(notifications):
    clean = json.loads(json.dumps(notifications or blank_notifications()))
    if clean.get("telegram", {}).get("bot_token"):
        clean["telegram"]["bot_token"] = "********"
    return clean


def blank_notifications():
    return {
        "telegram": {
            "enabled": False,
            "bot_token": "",
            "chat_id": "",
            "alert_types": ["stock", "catalog", "shipping", "scan"],
            "status": "Aguardando token criado no BotFather",
        }
    }


def user_notifications(payload, user, create=False):
    user_id = (user or {}).get("id")
    if not user_id:
        return blank_notifications()
    store = payload.setdefault("user_notifications", {}) if create else payload.get("user_notifications", {})
    if user_id in store:
        return store[user_id]
    if create:
        if is_master(user) and payload.get("notifications"):
            store[user_id] = json.loads(json.dumps(payload.get("notifications")))
        else:
            store[user_id] = blank_notifications()
        return store[user_id]
    if is_master(user) and payload.get("notifications"):
        return payload.get("notifications")
    return blank_notifications()


def notification_targets(payload):
    migrate_legacy_notifications(payload)
    targets = []
    store = payload.get("user_notifications") or {}
    for user_id, config in store.items():
        targets.append((user_id, config))
    if not targets and payload.get("notifications"):
        targets.append(("legacy", payload.get("notifications")))
    return targets


def migrate_legacy_notifications(payload):
    legacy = payload.get("notifications")
    if not legacy or payload.get("notifications_migrated_to_users"):
        return False
    users = ensure_users(payload)
    master = next((user for user in users if user.get("role") == "master"), None)
    if not master:
        return False
    payload.setdefault("user_notifications", {}).setdefault(master["id"], json.loads(json.dumps(legacy)))
    payload["notifications_migrated_to_users"] = True
    return True


def telegram_type_enabled(config, alert_type):
    telegram = (config or {}).get("telegram") or {}
    types = telegram.get("alert_types") or []
    return not types or alert_type in types


def public_payload(payload, actor=None, include_catalog=True):
    if include_catalog:
        reclassify_internal_competition(payload)
    if payload.get("catalog"):
        enrich_recent_sale_thumbnails(payload)
    if include_catalog:
        clean = json.loads(json.dumps(payload))
    else:
        clean = json.loads(json.dumps({key: value for key, value in payload.items() if key not in {"catalog", "item_logs"}}))
        clean["catalog"] = []
        clean["item_logs"] = []
    clean.setdefault("scan_items", [])
    clean.setdefault("recent_sales", [])
    for alert in clean.get("alerts", []):
        if alert.get("type") == "catalog":
            alert["message"] = normalize_catalog_alert_price(alert.get("message"))
    clean["clone_jobs"] = clean.get("clone_jobs", [])[:50]
    for job in clean["clone_jobs"]:
        if job.get("errors"):
            job["errors"] = job["errors"][:10]
    clean["item_logs"] = clean.get("item_logs", [])[:500]
    catalog = clean.get("catalog", [])
    clean["catalog_counts"] = catalog_counts(catalog) if include_catalog else clean.get("catalog_counts_snapshot") or catalog_counts([])
    live_progress = sync_progress_snapshot()
    public_accounts = []
    for account in clean.get("accounts", []):
        account_key = str(account.get("id") or "")
        seller_key = str(account.get("seller_id") or "")
        progress = live_progress.get(account_key) or live_progress.get(seller_key)
        public = public_account(account)
        if progress:
            public["sync_progress"] = progress
            public["sync_status"] = progress.get("message") or public.get("sync_status")
        elif "andamento" in str(public.get("sync_status") or "").lower():
            public["sync_progress"] = {
                "status": "running",
                "stage": "queued",
                "completed": 0,
                "total": int(public.get("sync_total_item_ids") or 0),
                "percent": 0,
                "message": "O servidor reiniciou; esta sincronização será retomada automaticamente.",
            }
            public["sync_status"] = public["sync_progress"]["message"]
        public_accounts.append(public)
    clean["accounts"] = public_accounts
    clean["notifications"] = public_notifications(user_notifications(payload, actor, create=False))
    clean["operations"] = (
        build_operations(clean)
        if include_catalog
        else clean.get("operations_snapshot") or build_operations(clean)
    )
    clean["users"] = visible_users_for(actor, ensure_users(clean))
    clean.pop("_catalog_loaded", None)
    clean.pop("_revision", None)
    clean.pop("operations_snapshot", None)
    clean.pop("catalog_counts_snapshot", None)
    return clean


def normalize_catalog_alert_price(message):
    text = str(message or "")
    pattern = re.compile(r"(Pre[cç]o para ganhar:\s*)(?!R\$\s*)(\d+(?:[.,]\d+)?)", re.I)

    def replace(match):
        raw = match.group(2).replace(".", "").replace(",", ".")
        return f"{match.group(1)}{brl_label(raw)}"

    return pattern.sub(replace, text)


def enrich_recent_sale_thumbnails(payload):
    catalog_by_id = {
        item.get("id"): item
        for item in payload.get("catalog", [])
        if item.get("id") and item.get("thumbnail")
    }
    changed = False
    for sale in payload.get("recent_sales", []):
        if sale.get("thumbnail"):
            continue
        catalog_item = catalog_by_id.get(sale.get("item_id")) or {}
        thumbnail = catalog_item.get("thumbnail") or ""
        if thumbnail:
            sale["thumbnail"] = thumbnail
            changed = True
    return changed


def percent_rate(value):
    try:
        return round(float(value or 0) * 100, 2)
    except (TypeError, ValueError):
        return 0


def brl(value):
    try:
        return f"R$ {float(value or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "R$ 0,00"


def default_tenant():
    return {
        "id": "tenant-production",
        "name": "Workspace CompeTIDOR",
        "plan": "Pro",
        "billing_status": "active",
        "account_limit": 10,
        "user_limit": 5,
    }


def default_users(tenant):
    return []


def ensure_users(payload):
    tenant = payload.get("tenant") or default_tenant()
    users = payload.get("users")
    if users is None:
        users = default_users(tenant)
    changed = False
    emails = set()
    for user in users:
        email = (user.get("email") or "").lower()
        if email in emails:
            user["email"] = f"{user.get('role') or 'usuario'}-{user.get('id', uuid.uuid4().hex[:6])}@competidor.umsoftware.com.br"
            changed = True
        emails.add((user.get("email") or "").lower())
    if users and not any(user.get("role") == "master" for user in users):
        first_admin = next((user for user in users if user.get("role") == "admin"), users[0] if users else None)
        if first_admin:
            first_admin["role"] = "master"
            first_admin["name"] = first_admin.get("name") or "Usuário Master"
            changed = True
    for user in users:
        if not user.get("password_hash"):
            password = secrets.token_urlsafe(18)
            salt, password_hash = hash_password(password)
            user["password_salt"] = salt
            user["password_hash"] = password_hash
            changed = True
    payload["users"] = users
    return users


def fallback_time(index):
    return time.strftime("%Y-%m-%d %H:%M", time.localtime(time.time() - (index * 2700)))


def current_month_period():
    now = datetime.now(APP_TZ)
    return f"{now.year:04d}-{now.month:02d}"


def current_month_window():
    now = datetime.now(APP_TZ)
    start = datetime(now.year, now.month, 1, tzinfo=APP_TZ)
    if now.month == 12:
        end = datetime(now.year + 1, 1, 1, tzinfo=APP_TZ)
    else:
        end = datetime(now.year, now.month + 1, 1, tzinfo=APP_TZ)
    return (
        current_month_period(),
        start.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
        end.strftime("%Y-%m-%dT%H:%M:%S.000-03:00"),
    )


def build_operations(payload):
    accounts = payload.get("accounts", [])
    catalog = payload.get("catalog", [])
    monthly = payload.get("monthly_revenue") or {}
    revenue_accounts = monthly.get("accounts") or {}
    period = monthly.get("period") or current_month_period()
    revenue = []
    revenue_source_accounts = [account for account in accounts if account.get("official")]
    for account in revenue_source_accounts:
        key = account.get("id") or account.get("nickname")
        record = revenue_accounts.get(key) or revenue_accounts.get(account.get("nickname")) or {}
        revenue.append(
            {
                "account": account.get("nickname"),
                "monthly_revenue": round(float(record.get("amount") or 0), 2),
                "currency": "BRL",
                "source": record.get("source") or "Pedidos oficiais Mercado Livre",
                "orders_count": int(record.get("orders_count") or 0),
                "period": period,
                "updated_at": record.get("updated_at") or "",
                "sync_status": record.get("sync_status") or account.get("sales_sync_status") or "Aguardando sincronização real",
            }
        )
    total_revenue = round(sum(item["monthly_revenue"] for item in revenue), 2)

    stock = []
    catalog_attention = []
    catalog_by_id = {item.get("id"): item for item in catalog if item.get("id")}
    seen_stock_items = set()
    for alert in payload.get("alerts", []):
        if alert.get("type") != "stock" or not alert.get("item_id"):
            continue
        item = catalog_by_id.get(alert.get("item_id")) or {}
        stock.append(
            {
                "id": alert.get("item_id"),
                "title": alert.get("product") or item.get("title") or alert.get("item_id"),
                "account": alert.get("account") or item.get("account"),
                "sku": alert.get("sku") or item.get("sku") or "-",
                "stock": 0,
                "price": item.get("price"),
                "thumbnail": item.get("thumbnail"),
                "occurred_at": alert.get("created_at") or "",
            }
        )
        seen_stock_items.add(alert.get("item_id"))
    for index, item in enumerate(catalog):
        row = {
            "id": item.get("id"),
            "title": item.get("title"),
            "account": item.get("account"),
            "sku": item.get("sku"),
            "stock": item.get("stock"),
            "price": item.get("price"),
            "thumbnail": item.get("thumbnail"),
            "occurred_at": item.get("updated_at") or fallback_time(index + 1),
        }
        if not item.get("internal_competition") and (
            item.get("status") == "losing"
            or item.get("competition_status") in {"competing", "sharing"}
            and item.get("winner_name") not in ("", None, item.get("account"))
        ):
            catalog_attention.append({**row, "winner_name": item.get("winner_name"), "winner_price": item.get("winner_price")})

    stock.sort(key=lambda item: item["occurred_at"], reverse=True)
    catalog_attention.sort(key=lambda item: item["occurred_at"], reverse=True)

    details_by_account = {}
    for detail in payload.get("claim_details", []):
        details_by_account.setdefault(detail.get("account"), []).append(detail)
    claims = payload.get("claims") or [
        {
            "account": account.get("nickname"),
            "open": 0,
            "mediations": 0,
            "updated_at": now_label(),
            "details": details_by_account.get(account.get("nickname"), []),
            "sync_status": account.get("claims_sync_status") or "Aguardando sincronização de reclamações",
        }
        for account in accounts
        if account.get("official")
    ]
    for claim in claims:
        claim["details"] = claim.get("details") or details_by_account.get(claim.get("account"), [])
        account = next((account for account in accounts if account.get("nickname") == claim.get("account")), {})
        claim["sync_status"] = claim.get("sync_status") or account.get("claims_sync_status") or ""
    shipments = payload.get("pending_shipments") or [
        {
            "account": account.get("nickname"),
            "order_id": "-",
            "buyer": "A sincronizar",
            "deadline": "Aguardando pedidos oficiais",
            "time_left": "-",
            "sync_status": account.get("sales_sync_status") or "Aguardando sincronização de pedidos",
        }
        for account in accounts
        if account.get("official")
    ]
    top_skus_today = sorted(
        payload.get("daily_sku_sales", []),
        key=lambda item: (int(item.get("units") or 0), float(item.get("revenue") or 0)),
        reverse=True,
    )[:20]

    return {
        "revenue": revenue,
        "total_monthly_revenue": total_revenue,
        "attention_stock": stock[:200],
        "attention_catalog": catalog_attention[:200],
        "claims": claims,
        "pending_shipments": shipments,
        "top_skus_today": top_skus_today,
    }


def request_json(url, method="GET", payload=None, headers=None, retries=None, timeout=None):
    body = None
    request_headers = headers or {}
    if payload is not None:
        body = json.dumps(payload).encode("utf-8")
        request_headers.setdefault("Content-Type", "application/json")
    method = method.upper()
    attempts = int(retries if retries is not None else (4 if method in {"GET", "PUT"} else 1))
    transient_statuses = {429, 500, 502, 503, 504}
    last_error = None
    request_timeout = max(2.0, float(timeout or os.getenv("MELI_REQUEST_TIMEOUT_SECONDS", "20")))
    for attempt in range(max(1, attempts)):
        req = urllib.request.Request(url, data=body, headers=request_headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=request_timeout) as response:
                text = response.read().decode("utf-8")
                return json.loads(text) if text else {}
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            last_error = RuntimeError(f"HTTP {exc.code}: {detail}")
            if exc.code not in transient_statuses or attempt + 1 >= attempts:
                raise last_error from exc
            retry_after = exc.headers.get("Retry-After") if exc.headers else None
            try:
                delay = float(retry_after) if retry_after else min(8.0, 0.6 * (2**attempt)) + random.uniform(0.1, 0.5)
            except (TypeError, ValueError):
                delay = min(8.0, 0.6 * (2**attempt)) + random.uniform(0.1, 0.5)
            time.sleep(delay)
        except urllib.error.URLError as exc:
            last_error = RuntimeError(f"Falha temporária de conexão com a API: {exc.reason}")
            if attempt + 1 >= attempts:
                raise last_error from exc
            time.sleep(min(8.0, 0.6 * (2**attempt)) + random.uniform(0.1, 0.5))
        except TimeoutError as exc:
            last_error = RuntimeError(f"A API do Mercado Livre não respondeu em até {request_timeout:g} segundos.")
            if attempt + 1 >= attempts:
                raise last_error from exc
            time.sleep(min(4.0, 0.4 * (2**attempt)) + random.uniform(0.1, 0.3))
    raise last_error or RuntimeError("Não foi possível concluir a chamada à API.")


def interactive_request_options():
    return {
        "retries": max(1, int(os.getenv("MELI_INTERACTIVE_RETRIES", "2"))),
        "timeout": max(2.0, float(os.getenv("MELI_INTERACTIVE_TIMEOUT_SECONDS", "8"))),
    }


def meli_background_work_busy():
    with SYNC_LOCK:
        sync_busy = bool(ACTIVE_SYNC_ACCOUNTS)
    with ASYNC_OPERATION_JOBS_LOCK:
        interactive_busy = bool(ACTIVE_CLONE_OPERATIONS)
    return sync_busy, interactive_busy


def wait_for_interactive_meli_priority():
    while True:
        with ASYNC_OPERATION_JOBS_LOCK:
            interactive_busy = bool(ACTIVE_CLONE_OPERATIONS)
        if not interactive_busy:
            return
        time.sleep(0.2)


def validate_meli_path(path):
    lowered = (path or "").lower()
    if any(term in lowered for term in SENSITIVE_MELI_PATH_TERMS):
        raise RuntimeError("Endpoint bloqueado por segurança: Mercado Pago, pagamentos e dados financeiros sensíveis não são permitidos.")
    if not any(pattern.match(path or "") for pattern in ALLOWED_MELI_PATHS):
        raise RuntimeError("Endpoint Mercado Livre não autorizado pela política de segurança do CompeTIDOR.")


def request_form(url, payload, headers=None):
    request_headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    request_headers.update(headers or {})
    req = urllib.request.Request(url, data=urlencode(payload).encode("utf-8"), headers=request_headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            text = response.read().decode("utf-8")
            return json.loads(text) if text else {}
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code}: {detail}") from exc


def empty_payload():
    return {
        "tenant": default_tenant(),
        "users": [],
        "accounts": [],
        "catalog": [],
        "alerts": [],
        "metrics": [],
        "competitors": [],
        "clone_jobs": [],
        "scan_items": [],
        "recent_sales": [],
        "claim_details": [],
        "pending_shipments": [],
        "monthly_revenue": {"period": current_month_period(), "accounts": {}},
        "user_notifications": {},
        "notifications_migrated_to_users": True,
    }

class MercadoLivreClient:
    def __init__(self, access_token=None):
        self.access_token = access_token

    def auth_url(self, state, redirect_uri=None, switch_account=False):
        config = meli_config()
        client_id = config["client_id"]
        redirect_uri = redirect_uri or config["redirect_uri"]
        params = {
            "response_type": "code",
            "client_id": client_id,
            "redirect_uri": redirect_uri,
            "state": state,
        }
        if switch_account:
            params["prompt"] = "login"
            params["switch_account"] = "true"
        return f"{MELI_AUTH_URL}?{urlencode(params)}"

    def exchange_code(self, code, redirect_uri=None):
        config = meli_config()
        payload = {
            "grant_type": "authorization_code",
            "client_id": config["client_id"],
            "client_secret": config["client_secret"],
            "code": code,
            "redirect_uri": redirect_uri or config["redirect_uri"],
        }
        return request_form(MELI_TOKEN_URL, payload)

    def refresh(self, refresh_token):
        payload = {
            "grant_type": "refresh_token",
            "client_id": meli_config()["client_id"],
            "client_secret": meli_config()["client_secret"],
            "refresh_token": refresh_token,
        }
        return request_form(MELI_TOKEN_URL, payload)

    def get(self, path, extra_headers=None, retries=None, timeout=None):
        validate_meli_path(path)
        headers = {"Authorization": f"Bearer {self.access_token}", "Accept": "application/json"}
        headers.update(extra_headers or {})
        return request_json(
            f"{MELI_API_URL}{path}",
            headers=headers,
            retries=retries,
            timeout=timeout,
        )

    def post(self, path, payload):
        validate_meli_path(path)
        return request_json(
            f"{MELI_API_URL}{path}",
            method="POST",
            payload=payload,
            headers={"Authorization": f"Bearer {self.access_token}", "Accept": "application/json"},
        )

    def put(self, path, payload, retries=None, timeout=None):
        validate_meli_path(path)
        return request_json(
            f"{MELI_API_URL}{path}",
            method="PUT",
            payload=payload,
            headers={"Authorization": f"Bearer {self.access_token}", "Accept": "application/json"},
            retries=retries,
            timeout=timeout,
        )

    def me(self):
        return self.get("/users/me")

    def user(self, seller_id):
        return self.get(f"/users/{seller_id}")

    def user_brands(self, seller_id, interactive=False):
        options = interactive_request_options() if interactive else {}
        return self.get(f"/users/{seller_id}/brands", **options)

    def seller_items(self, seller_id, limit=50, offset=0):
        return self.get(f"/users/{seller_id}/items/search?limit={limit}&offset={offset}")

    def seller_items_scan(self, seller_id, limit=100, scroll_id="", status=""):
        params = {"search_type": "scan", "limit": min(int(limit or 100), 100)}
        if scroll_id:
            params["scroll_id"] = scroll_id
        if status:
            params["status"] = status
        return self.get(
            f"/users/{seller_id}/items/search?{urlencode(params)}",
            retries=max(1, int(os.getenv("MELI_SYNC_LIST_RETRIES", "3"))),
            timeout=max(4.0, float(os.getenv("MELI_SYNC_LIST_TIMEOUT_SECONDS", "12"))),
        )

    def seller_all_items(self, seller_id, max_items=None, progress=None):
        max_items = None if max_items in (None, "all", 0, "0") else int(max_items)
        statuses = [
            status.strip()
            for status in os.getenv("MELI_SYNC_STATUSES", "active,paused,under_review").split(",")
            if status.strip()
        ]
        progress_lock = threading.Lock()
        discovered = {status: 0 for status in statuses}

        def scan_status(status):
            status_results = []
            status_seen = set()
            scroll_id = ""
            empty_pages = 0
            max_pages = max(1, int(os.getenv("MELI_SCAN_MAX_PAGES", "1000")))
            for _ in range(max_pages):
                wait_for_interactive_meli_priority()
                page = self.seller_items_scan(seller_id, limit=100, scroll_id=scroll_id, status=status)
                batch = page.get("results", []) or []
                added = 0
                for item_id in batch:
                    if item_id and item_id not in status_seen:
                        status_seen.add(item_id)
                        status_results.append(item_id)
                        added += 1
                with progress_lock:
                    discovered[status] = len(status_results)
                    if progress:
                        progress(sum(discovered.values()), status)
                scroll_id = page.get("scroll_id") or scroll_id
                if not batch or added == 0:
                    empty_pages += 1
                if not scroll_id or empty_pages >= 2:
                    break
                if max_items and len(status_results) >= max_items:
                    break
            return status_results

        results = []
        seen = set()
        list_workers = max(1, min(len(statuses), int(os.getenv("MELI_SYNC_LIST_WORKERS", "3"))))
        with ThreadPoolExecutor(max_workers=list_workers, thread_name_prefix="meli-list") as executor:
            futures = [executor.submit(scan_status, status) for status in statuses]
            for future in futures:
                for item_id in future.result():
                    if item_id not in seen:
                        seen.add(item_id)
                        results.append(item_id)
        if results:
            return results[:max_items] if max_items else results

        # Fallback para contas/permissões em que search_type=scan não esteja disponível.
        fallback_limit = max_items or int(os.getenv("MELI_OFFSET_FALLBACK_LIMIT", "1000"))
        results = []
        offset = 0
        page_size = 50
        while len(results) < fallback_limit:
            wait_for_interactive_meli_priority()
            page = self.seller_items(seller_id, page_size, offset)
            batch = page.get("results", [])
            results.extend(batch)
            total = (page.get("paging") or {}).get("total") or len(results)
            if not batch or len(results) >= total:
                break
            offset += page_size
        return results[:fallback_limit]

    def item(self, item_id):
        return self.get(f"/items/{item_id}")

    def item_for_clone(self, item_id):
        return self.get(f"/items/{item_id}?include_attributes=all", **interactive_request_options())

    def items_bulk(self, item_ids):
        ids = ",".join(item_ids)
        rows = self.get(f"/items?ids={ids}&include_attributes=all")
        items = []
        for row in rows if isinstance(rows, list) else []:
            body = row.get("body") if isinstance(row, dict) else None
            if body:
                items.append(body)
        return items

    def item_description(self, item_id, interactive=False):
        options = interactive_request_options() if interactive else {}
        return self.get(f"/items/{item_id}/description", **options)

    def product(self, catalog_product_id, interactive=False):
        options = interactive_request_options() if interactive else {}
        return self.get(f"/products/{catalog_product_id}", **options)

    def user_product(self, user_product_id, interactive=False):
        options = interactive_request_options() if interactive else {}
        return self.get(f"/user-products/{user_product_id}", **options)

    def item_shipping_cost(self, seller_id, item_id):
        params = {"item_id": item_id, "verbose": "true"}
        return self.get(f"/users/{seller_id}/shipping_options/free?{urlencode(params)}")

    def price_to_win(self, item_id):
        return self.get(f"/items/{item_id}/price_to_win?version=v2")

    def product_winners(self, catalog_product_id):
        return self.get(f"/products/{catalog_product_id}/items?site_id=MLB")

    def category_attributes(self, category_id):
        return self.get(f"/categories/{category_id}/attributes")

    def create_item(self, payload):
        return self.post("/items", payload)

    def create_item_description(self, item_id, plain_text):
        return self.post(f"/items/{item_id}/description", {"plain_text": plain_text})

    def update_item_description(self, item_id, plain_text, interactive=False):
        options = interactive_request_options() if interactive else {}
        return self.put(f"/items/{item_id}/description", {"plain_text": plain_text}, **options)

    def update_item(self, item_id, payload):
        return self.put(f"/items/{item_id}", payload)

    def seller_orders(self, seller_id, limit=50, offset=0, date_from=None, date_to=None):
        params = {
            "seller": seller_id,
            "sort": "date_desc",
            "limit": min(int(limit or 50), 50),
            "offset": max(int(offset or 0), 0),
        }
        if date_from:
            params["order.date_created.from"] = date_from
        if date_to:
            params["order.date_created.to"] = date_to
        return self.get(f"/orders/search?{urlencode(params)}")

    def shipment(self, shipment_id):
        return self.get(f"/shipments/{shipment_id}", extra_headers={"x-format-new": "true"})

    def shipment_sla(self, shipment_id):
        return self.get(f"/shipments/{shipment_id}/sla")

    def seller_claims(self, seller_id, limit=50, offset=0):
        params = {
            "players.user_id": seller_id,
            "players.role": "respondent",
            "status": "opened",
            "sort": "last_updated:desc",
            "limit": min(int(limit or 50), 100),
            "offset": max(int(offset or 0), 0),
        }
        return self.get(f"/post-purchase/v1/claims/search?{urlencode(params)}")

    def claim_detail(self, claim_id):
        return self.get(f"/post-purchase/v1/claims/{claim_id}/detail")


class Notifier:
    def __init__(self, config):
        self.config = config

    def telegram_get_me(self):
        cfg = self.config.get("telegram", {})
        if not cfg.get("bot_token"):
            return {"ok": False, "status": "Token do Telegram não configurado"}
        return request_json(f"{TELEGRAM_API_URL}/bot{cfg['bot_token']}/getMe")

    def telegram_updates(self):
        cfg = self.config.get("telegram", {})
        if not cfg.get("bot_token"):
            return {"ok": False, "status": "Token do Telegram não configurado"}
        return request_json(f"{TELEGRAM_API_URL}/bot{cfg['bot_token']}/getUpdates")

    def send_telegram(self, text):
        cfg = self.config.get("telegram", {})
        if not (cfg.get("enabled") and cfg.get("bot_token") and cfg.get("chat_id")):
            return {"ok": False, "status": "Telegram não configurado"}
        try:
            me = self.telegram_get_me()
            bot_id = str((me.get("result") or {}).get("id") or "")
            if bot_id and str(cfg.get("chat_id")) == bot_id:
                return {
                    "ok": False,
                    "error": "O Chat ID informado é o ID do próprio bot. Envie uma mensagem para o bot no Telegram e use o chat.id retornado em getUpdates.",
                }
        except Exception:
            pass
        url = f"{TELEGRAM_API_URL}/bot{cfg['bot_token']}/sendMessage"
        return request_json(url, method="POST", payload={"chat_id": cfg["chat_id"], "text": text})


def add_or_update_account(payload, account):
    accounts = payload["accounts"]
    for index, current in enumerate(accounts):
        if str(current.get("seller_id")) == str(account.get("seller_id")):
            accounts[index] = {**current, **account}
            return accounts[index]
    accounts.append(account)
    return account


def account_client(account):
    if not account.get("access_token"):
        raise RuntimeError("Conta oficial sem access token salvo. Refaça o login OAuth.")
    expires_at = int(account.get("token_created_at") or 0) + int(account.get("expires_in") or 0)
    if account.get("refresh_token") and expires_at and expires_at - 300 <= int(time.time()):
        token = MercadoLivreClient().refresh(account["refresh_token"])
        account["access_token"] = token.get("access_token", account.get("access_token"))
        account["refresh_token"] = token.get("refresh_token", account.get("refresh_token"))
        account["expires_in"] = token.get("expires_in", account.get("expires_in"))
        account["token_created_at"] = int(time.time())
    return MercadoLivreClient(account["access_token"])


def item_sku(item):
    if item.get("seller_custom_field"):
        return item["seller_custom_field"]
    for key in ("seller_sku", "seller_custom_field", "sku"):
        if item.get(key):
            return item[key]
    for attribute in item.get("attributes", []) or []:
        if attribute.get("id") in {"SELLER_SKU", "SKU"} and attribute.get("value_name"):
            return attribute["value_name"]
    for variation in item.get("variations", []) or []:
        if variation.get("seller_custom_field"):
            return variation["seller_custom_field"]
        for attribute in variation.get("attributes", []) or []:
            if attribute.get("id") in {"SELLER_SKU", "SKU"} and attribute.get("value_name"):
                return attribute["value_name"]
    return "-"


def item_available_quantity(item):
    quantity = item.get("available_quantity")
    if quantity not in (None, ""):
        try:
            return int(float(quantity or 0))
        except (TypeError, ValueError):
            pass
    total = 0
    for variation in item.get("variations", []) or []:
        try:
            total += int(float(variation.get("available_quantity") or 0))
        except (TypeError, ValueError):
            pass
    if total:
        return total
    try:
        return int(float(item.get("initial_quantity") or item.get("stock") or 0))
    except (TypeError, ValueError):
        return 0


def item_flex_logistic_type(item):
    shipping = item.get("shipping") or {}
    logistic_type = shipping.get("logistic_type") or ""
    mode = shipping.get("mode") or ""
    tags = shipping.get("tags") or item.get("shipping_tags") or item.get("tags") or item.get("sub_status") or []
    if isinstance(tags, str):
        tags = [tags]
    normalized_tags = {str(tag).strip().lower() for tag in tags if str(tag).strip()}
    active_tags = {"self_service_in", "self_service", "mercado_envios_flex", "flex"}
    if logistic_type == "self_service" or mode == "self_service" or normalized_tags.intersection(active_tags):
        return "self_service"
    # self_service_available only means that Flex can be enabled; self_service_out
    # explicitly means that the listing is currently outside the service.
    if normalized_tags.intersection({"self_service_out", "self_service_available"}):
        return logistic_type if logistic_type and logistic_type != "self_service" else mode
    return logistic_type or mode


def item_attribute_value(item, attr_ids):
    wanted = {str(attr_id).upper() for attr_id in attr_ids}
    for attribute in item.get("attributes", []) or []:
        if str(attribute.get("id") or "").upper() not in wanted:
            continue
        value = clean_attribute_value(attribute.get("value_name"))
        if value:
            return normalize_package_value(value, attr_ids)
        struct = attribute.get("value_struct") or {}
        if struct.get("number") and struct.get("unit"):
            return normalize_package_value(f"{struct.get('number')} {struct.get('unit')}", attr_ids)
    return ""


def package_dimensions_from_shipping(item):
    dimensions = first_present(item, ["shipping.dimensions"], "")
    if not dimensions:
        return {}
    match = re.match(r"\s*([\d.,]+)x([\d.,]+)x([\d.,]+)\s*,\s*([\d.,]+)", str(dimensions), flags=re.I)
    if not match:
        return {}
    width, height, length, weight = [parse_decimal_number(part) for part in match.groups()]
    return {
        "package_width": f"{width:g} cm",
        "package_height": f"{height:g} cm",
        "package_length": f"{length:g} cm",
        "package_weight": f"{weight / 1000:g} kg",
    }


def normalize_package_value(value, attr_ids):
    text = clean_attribute_value(value)
    match = re.search(r"([\d.,]+)\s*([a-zA-Z]+)", text)
    if not match:
        return text
    number = parse_decimal_number(match.group(1))
    unit = match.group(2).lower()
    attr_text = " ".join(str(attr).upper() for attr in attr_ids)
    if "WEIGHT" in attr_text:
        if unit in {"g", "gr", "gramas"}:
            return f"{number / 1000:g} kg"
        if unit in {"mg"}:
            return f"{number / 1000000:g} kg"
        if unit in {"kg", "kgs", "quilo", "quilos"}:
            return f"{number:g} kg"
        return f"{number:g} {unit}"
    if unit == "mm":
        return f"{number / 10:g} cm"
    if unit in {"m", "metro", "metros"}:
        return f"{number * 100:g} cm"
    return f"{number:g} {unit}"


def parse_decimal_number(value):
    text = str(value or "").strip()
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    return float(text)


def seller_package_api_value(field, value):
    text = clean_attribute_value(value)
    match = re.search(r"([\d.,]+)\s*([a-zA-Z]*)", text)
    if not match:
        raise RuntimeError(f"Valor inválido para {field}.")
    number = parse_decimal_number(match.group(1))
    unit = (match.group(2) or "").lower()
    if field == "package_weight":
        if unit in {"kg", "kgs", "quilo", "quilos"} or not unit:
            grams = round(number * 1000)
        elif unit in {"g", "gr", "grama", "gramas"}:
            grams = round(number)
        else:
            raise RuntimeError("Informe o peso em kg ou g. Exemplo: 1,16 kg.")
        return f"{max(1, grams)} g"
    if unit == "mm":
        number /= 10
    elif unit in {"m", "metro", "metros"}:
        number *= 100
    elif unit not in {"", "cm"}:
        raise RuntimeError("Informe as dimensões em cm. Exemplo: 31 cm.")
    return f"{number:g} cm"


def package_values_from_item(item):
    shipping_dimensions = package_dimensions_from_shipping(item)
    return {
        "package_weight": item_attribute_value(item, ["SELLER_PACKAGE_WEIGHT"]) or shipping_dimensions.get("package_weight") or item_attribute_value(item, ["PACKAGE_WEIGHT"]) or "",
        "package_height": item_attribute_value(item, ["SELLER_PACKAGE_HEIGHT"]) or shipping_dimensions.get("package_height") or item_attribute_value(item, ["PACKAGE_HEIGHT"]) or "",
        "package_width": item_attribute_value(item, ["SELLER_PACKAGE_WIDTH"]) or shipping_dimensions.get("package_width") or item_attribute_value(item, ["PACKAGE_WIDTH"]) or "",
        "package_length": item_attribute_value(item, ["SELLER_PACKAGE_LENGTH"]) or shipping_dimensions.get("package_length") or item_attribute_value(item, ["PACKAGE_LENGTH"]) or "",
    }


def package_value_number(field, value):
    normalized = normalize_package_value(value, ["SELLER_PACKAGE_WEIGHT" if field == "package_weight" else "SELLER_PACKAGE_LENGTH"])
    match = re.search(r"[\d.,]+", normalized or "")
    return parse_decimal_number(match.group(0)) if match else None


def verify_package_update(client, item_id, expected):
    expected_numbers = {
        field: package_value_number(field, value)
        for field, value in expected.items()
        if clean_attribute_value(value)
    }
    latest = {}
    mismatches = []
    for attempt in range(4):
        latest = client.item(item_id)
        actual = package_values_from_item(latest)
        mismatches = []
        for field, expected_number in expected_numbers.items():
            actual_number = package_value_number(field, actual.get(field))
            tolerance = 0.005 if field == "package_weight" else 0.05
            if actual_number is None or expected_number is None or abs(actual_number - expected_number) > tolerance:
                mismatches.append(field)
        if not mismatches:
            return latest
        if attempt < 3:
            time.sleep(0.5 * (attempt + 1))
    labels = {
        "package_weight": "peso",
        "package_height": "altura",
        "package_width": "largura",
        "package_length": "comprimento",
    }
    names = ", ".join(labels.get(field, field) for field in mismatches)
    raise RuntimeError(
        f"O Mercado Livre recebeu a solicitação, mas não confirmou a alteração de {names}. "
        "Em anúncios Full ou com dimensões gerenciadas pela logística, esses campos podem ser somente leitura."
    )


def item_thumbnail(item):
    thumbnail = item.get("secure_thumbnail") or item.get("thumbnail") or ""
    pictures = item.get("pictures") or []
    if pictures:
        thumbnail = pictures[0].get("secure_url") or pictures[0].get("url") or thumbnail
    return thumbnail.replace("http://", "https://") if thumbnail else ""


def product_thumbnail(product):
    pictures = product.get("pictures") or []
    if pictures:
        first = pictures[0]
        if isinstance(first, dict):
            thumbnail = first.get("secure_url") or first.get("url") or first.get("id") or ""
        else:
            thumbnail = str(first)
        return thumbnail.replace("http://", "https://") if thumbnail else ""
    thumbnail = product.get("secure_thumbnail") or product.get("thumbnail") or product.get("picture") or ""
    return thumbnail.replace("http://", "https://") if thumbnail else ""


def first_present(payload, keys, default=None):
    for key in keys:
        current = payload
        for part in key.split("."):
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                current = None
                break
        if current not in (None, ""):
            return current
    return default


def extract_meli_item_id(value):
    text = (value or "").strip().upper()
    match = re.search(r"\b(ML[A-Z]\d{5,})\b", text)
    if match:
        return match.group(1)
    compact = re.sub(r"[^A-Z0-9]", "", text)
    match = re.search(r"(ML[A-Z]\d{5,})", compact)
    if match:
        return match.group(1)
    return text if re.fullmatch(r"ML[A-Z]\d{5,}", text) else ""


def policy_error_message(exc, action):
    text = str(exc)
    if "PA_UNAUTHORIZED_RESULT_FROM_POLICIES" in text or "PolicyAgent" in text or "HTTP 403" in text:
        return (
            f"O Mercado Livre bloqueou {action} por política/permissão da aplicação. "
            "Confirme no painel de desenvolvedores se a aplicação tem permissões de leitura de anúncios/vendas "
            "e refaça o OAuth da conta conectada."
        )
    return text


def official_reader_client(payload):
    account = next((item for item in payload.get("accounts", []) if item.get("official") and item.get("access_token")), None)
    if not account:
        return None
    return account_client(account)


def meli_read(payload, path):
    validate_meli_path(path)
    client = official_reader_client(payload)
    if client:
        return client.get(path)
    return request_json(f"{MELI_API_URL}{path}")


def meli_public_read(path):
    validate_meli_path(path)
    return request_json(f"{MELI_API_URL}{path}")


def try_meli_sources(payload, paths):
    last_error = None
    for path in paths:
        for reader in (lambda p: meli_read(payload, p), meli_public_read):
            try:
                return reader(path)
            except Exception as exc:
                last_error = exc
    if last_error:
        raise last_error
    raise RuntimeError("Nenhuma rota Mercado Livre informada.")


def seller_name_for(payload, seller_id):
    if not seller_id:
        return "Vendedor não informado"
    try:
        seller = meli_read(payload, f"/users/{seller_id}")
        return seller.get("nickname") or f"Seller {seller_id}"
    except Exception:
        return f"Seller {seller_id}"


def resolve_scan_target(payload, target_id):
    try:
        item = meli_read(payload, f"/items/{target_id}")
        item["_scan_target_type"] = "item"
        return item
    except Exception as exc:
        text = str(exc)
        if "not_found" not in text and "HTTP 404" not in text:
            raise RuntimeError(policy_error_message(exc, "a leitura deste anúncio no Scan")) from exc

    try:
        product = meli_read(payload, f"/products/{target_id}")
        offers = meli_read(payload, f"/products/{target_id}/items?site_id=MLB")
    except Exception as exc:
        raise RuntimeError(
            "Não encontrei esse código como anúncio nem como produto de catálogo. "
            "Cole o link completo do anúncio/produto ou confira se o MLB está correto."
        ) from exc

    rows = offers.get("results") if isinstance(offers, dict) else offers
    if isinstance(rows, dict):
        rows = rows.get("results") or rows.get("items") or []
    rows = rows or []
    candidates = catalog_offer_candidates(rows)
    if not candidates:
        raise RuntimeError("Produto de catálogo encontrado, mas sem ofertas disponíveis para acompanhar.")
    live_candidates = live_catalog_offers(payload, candidates)
    validation_source = "live_item"
    if not live_candidates:
        live_candidates = candidates
        validation_source = "catalog_endpoint_unverified"
    best = live_candidates[0]
    item = {
        "id": best["item_id"] or target_id,
        "catalog_product_id": target_id,
        "title": product.get("name") or product.get("title") or best.get("title") or target_id,
        "seller_id": best["seller_id"],
        "price": best["price"],
        "permalink": best["permalink"] or product.get("permalink") or "",
        "thumbnail": product_thumbnail(product) or best.get("thumbnail") or "",
        "_scan_target_type": "catalog_product",
        "_scan_offer_count": len(live_candidates),
        "_scan_total_offer_count": len(candidates),
        "_scan_offers": live_candidates,
        "_scan_validation_source": validation_source,
    }
    return item


def offer_value(row, keys, default=""):
    return first_present(row, keys, default)


def catalog_offer_candidates(rows, sort_by_price=True):
    candidates = []
    for position, row in enumerate(rows or []):
        price = offer_value(row, ["price", "sale_price.amount", "current_price", "amount"], 0) or 0
        try:
            price = float(price)
        except (TypeError, ValueError):
            price = 0
        item_id = offer_value(row, ["item_id", "id", "item.id"], "")
        seller_id = offer_value(row, ["seller_id", "seller.id", "seller.id_seller"], "")
        permalink = offer_value(row, ["permalink", "item.permalink"], "")
        tags = row.get("tags") if isinstance(row, dict) else []
        tags = tags if isinstance(tags, list) else []
        winner_markers = [
            row.get("winner") if isinstance(row, dict) else None,
            row.get("winning") if isinstance(row, dict) else None,
            row.get("is_winner") if isinstance(row, dict) else None,
            first_present(row, ["metadata.winner", "catalog_position.winner"], None) if isinstance(row, dict) else None,
        ]
        is_winner = any(str(marker).lower() == "true" for marker in winner_markers) or any(
            str(tag).lower() in {"winner", "catalog_winner", "best_seller"} for tag in tags
        )
        if price:
            candidates.append(
                {
                    "price": price,
                    "item_id": item_id,
                    "seller_id": str(seller_id) if seller_id else "",
                    "permalink": permalink,
                    "position": position,
                    "is_winner": is_winner,
                    "raw": row,
                }
            )
    if sort_by_price:
        return sorted(candidates, key=lambda item: item["price"])
    return sorted(candidates, key=lambda item: (0 if item.get("is_winner") else 1, item.get("position", 999999)))


def live_catalog_offers(payload, candidates, sort_by_price=True):
    live = []
    max_offers = max(1, int(os.getenv("SCAN_VALIDATE_OFFERS_LIMIT", "150")))
    for candidate in candidates[:max_offers]:
        item_id = candidate.get("item_id")
        if not item_id:
            continue
        try:
            item = try_meli_sources(payload, [f"/items/{item_id}"])
        except Exception:
            continue
        status = item.get("status") or ""
        if status and status not in {"active", "under_review"}:
            continue
        try:
            price = float(item.get("price") or 0)
        except (TypeError, ValueError):
            price = 0
        if not price:
            continue
        live.append(
            {
                **candidate,
                "price": price,
                "seller_id": str(item.get("seller_id") or candidate.get("seller_id") or ""),
                "permalink": item.get("permalink") or candidate.get("permalink") or "",
                "thumbnail": item_thumbnail(item),
                "status": status or "active",
                "title": item.get("title") or "",
            }
        )
    if sort_by_price:
        return sorted(live, key=lambda item: item["price"])
    return sorted(live, key=lambda item: (0 if item.get("is_winner") else 1, item.get("position", 999999)))


def seller_name(payload, seller_id, client=None):
    seller_id = str(seller_id or "")
    if not seller_id:
        return ""
    for account in payload.get("accounts", []):
        if str(account.get("seller_id")) == seller_id:
            return account.get("nickname") or f"Seller {seller_id}"
    if client:
        try:
            seller = client.user(seller_id)
            return seller.get("nickname") or f"Seller {seller_id}"
        except Exception:
            pass
    return f"Seller {seller_id}"


def scan_meli_item(payload, scan_id):
    scans = payload.setdefault("scan_items", [])
    scan = next((item for item in scans if item.get("id") == scan_id), None)
    if not scan:
        raise RuntimeError("Produto de scan não encontrado.")
    item_id = scan.get("item_id") or extract_meli_item_id(scan.get("url"))
    if not item_id:
        raise RuntimeError("Não foi possível identificar o código MLB do anúncio.")
    item_data = resolve_scan_target(payload, item_id)
    seller_id = item_data.get("seller_id")
    seller_name = seller_name_for(payload, seller_id)
    price = float(item_data.get("price") or 0)
    minimum = float(scan.get("minimum_price") or 0)
    below_minimum_offers = []
    for offer in item_data.get("_scan_offers", []) or []:
        offer_price = float(offer.get("price") or 0)
        if minimum and offer_price <= minimum:
            offer_seller_id = offer.get("seller_id") or ""
            below_minimum_offers.append(
                {
                    "item_id": offer.get("item_id") or "",
                    "seller_id": offer_seller_id,
                    "seller_name": seller_name_for(payload, offer_seller_id),
                    "price": offer_price,
                    "permalink": offer.get("permalink") or "",
                }
            )
    below_minimum_offers = sorted(below_minimum_offers, key=lambda item: item["price"])
    history = scan.setdefault("history", [])
    last_price = history[0].get("price") if history else None
    changed = last_price != price
    entry = {
        "id": f"scan-log-{uuid.uuid4().hex[:10]}",
        "item_id": item_id,
        "title": item_data.get("title") or scan.get("name") or item_id,
        "seller_id": seller_id or "",
        "seller_name": seller_name,
        "price": price,
        "permalink": item_data.get("permalink") or scan.get("url") or "",
        "thumbnail": item_thumbnail(item_data),
        "target_type": item_data.get("_scan_target_type", "item"),
        "offer_count": item_data.get("_scan_offer_count", 1),
        "validation_source": item_data.get("_scan_validation_source", "live_item"),
        "created_at": now_label(),
        "changed": changed,
    }
    if changed or not history:
        history.insert(0, entry)
    scan["item_id"] = item_id
    scan["last_title"] = entry["title"]
    scan["last_seller_name"] = seller_name
    scan["last_price"] = price
    scan["last_thumbnail"] = entry["thumbnail"]
    scan["last_scan_at"] = entry["created_at"]
    scan["last_permalink"] = entry["permalink"]
    scan["target_type"] = entry["target_type"]
    scan["offer_count"] = entry["offer_count"]
    scan["validation_source"] = entry["validation_source"]
    scan["below_minimum_offers"] = below_minimum_offers
    scan["catalog_reference_offers"] = []
    alert_signature = "|".join(f"{offer['item_id']}:{offer['price']}" for offer in below_minimum_offers) or f"{item_id}:{price}"
    if minimum and (below_minimum_offers or price <= minimum) and scan.get("last_alert_signature") != alert_signature:
        if below_minimum_offers:
            seller_lines = "\n".join(
                f"- {offer['seller_name']} | {offer['item_id']} | {brl(offer['price'])}"
                for offer in below_minimum_offers[:6]
            )
            message = "\n".join(
                [
                    "CompeTIDOR | Alerta Scan",
                    "------------------------",
                    f"Produto: {scan.get('name') or entry['title']}",
                    f"Mínimo configurado: {brl(minimum)}",
                    f"Ofertas abaixo: {len(below_minimum_offers)}",
                    "",
                    seller_lines,
                ]
            )
        else:
            message = "\n".join(
                [
                    "CompeTIDOR | Alerta Scan",
                    "------------------------",
                    f"Produto: {scan.get('name') or entry['title']}",
                    f"Preço atual: {brl(price)}",
                    f"Mínimo configurado: {brl(minimum)}",
                    f"Vendedor: {seller_name}",
                    f"Anúncio: {item_id}",
                ]
            )
        try:
            send_telegram_message_to_users(payload, "scan", message)
            scan["last_alert_price"] = price
            scan["last_alert_signature"] = alert_signature
            scan["last_alert_at"] = now_label()
        except Exception as exc:
            scan["last_alert_error"] = str(exc)
        payload.setdefault("alerts", []).insert(
            0,
            {
                "id": f"alert-scan-{uuid.uuid4().hex[:8]}",
                "type": "scan",
                "severity": "critical",
                "title": f"Preço abaixo do mínimo no Scan: {scan.get('name') or item_id}",
                "message": message,
                "account": "Scan",
                "channel": ["dashboard", "telegram"],
                "created_at": now_label(),
                "read": False,
            },
        )
    scan["history"] = history[:120]
    return scan, entry


def scan_competitor_profile(payload, seller_id, limit=50):
    seller_id = str(seller_id or "").strip()
    if not seller_id:
        raise RuntimeError("Informe o ID do vendedor concorrente.")
    requested_limit = max(1, min(int(limit or 100), 500))
    errors = []
    try:
        profile = try_meli_sources(payload, [f"/users/{seller_id}"])
    except Exception as exc:
        errors.append(str(exc))
        profile = {"nickname": f"Seller {seller_id}", "permalink": ""}
    ids = []
    total = 0
    source = "users_items_search"
    routes = [
        ("users_items_search", lambda size, offset: f"/users/{seller_id}/items/search?limit={size}&offset={offset}"),
        ("sites_search_public", lambda size, offset: f"/sites/MLB/search?seller_id={seller_id}&limit={size}&offset={offset}"),
    ]
    for route_name, route_builder in routes:
        ids = []
        total = 0
        source = route_name
        offset = 0
        try:
            while len(ids) < requested_limit:
                page_size = min(50, requested_limit - len(ids))
                page = try_meli_sources(payload, [route_builder(page_size, offset)])
                batch = page.get("results", []) or []
                for row in batch:
                    ids.append(row if isinstance(row, str) else row.get("id"))
                ids = [item_id for item_id in ids if item_id]
                total = (page.get("paging") or {}).get("total") or len(ids)
                if not batch or len(ids) >= total:
                    break
                offset += page_size
            if ids:
                break
        except Exception as exc:
            errors.append(str(exc))
            ids = []
            continue
    items = []
    for item_id in ids[:requested_limit]:
        try:
            item = try_meli_sources(payload, [f"/items/{item_id}"])
            price = float(item.get("price") or 0)
            sold = int(item.get("sold_quantity") or 0)
            items.append(
                {
                    "id": item.get("id"),
                    "title": item.get("title") or item.get("id"),
                    "price": price,
                    "sold_quantity": sold,
                    "available_quantity": item_available_quantity(item),
                    "status": item.get("status") or "-",
                    "listing_type_id": item.get("listing_type_id") or "",
                    "thumbnail": item_thumbnail(item),
                    "permalink": item.get("permalink") or "",
                    "estimated_revenue": round(price * sold, 2),
                }
            )
        except Exception:
            continue
    prices = [item["price"] for item in items if item.get("price")]
    competitor = {
        "id": f"competitor-{seller_id}",
        "seller_id": seller_id,
        "name": profile.get("nickname") or f"Seller {seller_id}",
        "permalink": profile.get("permalink") or "",
        "reputation": first_present(profile, ["seller_reputation.level_id"], "Não informado"),
        "transactions_total": first_present(profile, ["seller_reputation.transactions.total"], None),
        "transactions_completed": first_present(profile, ["seller_reputation.transactions.completed"], None),
        "items_total": total or len(ids),
        "items_loaded": len(items),
        "analysis_limit": requested_limit,
        "source": source,
        "sync_status": "ok" if items else "blocked",
        "sync_error": policy_error_message(Exception(errors[-1]), "a análise deste concorrente") if errors and not items else "",
        "price_min": min(prices) if prices else 0,
        "price_max": max(prices) if prices else 0,
        "price_avg": round(sum(prices) / len(prices), 2) if prices else 0,
        "estimated_revenue": round(sum(item.get("estimated_revenue") or 0 for item in items), 2),
        "items": items,
        "updated_at": now_label(),
        "note": "Dados públicos oficiais. Faturamento é estimado por preço atual x sold_quantity público, quando disponível."
        if items
        else "O Mercado Livre bloqueou a listagem pública de anúncios deste vendedor por política da API. Tente acompanhar concorrência por produto/catálogo via Scan.",
    }
    competitors = payload.setdefault("competitors", [])
    existing = next((index for index, item in enumerate(competitors) if str(item.get("seller_id")) == seller_id), None)
    if existing is None:
        competitors.insert(0, competitor)
    else:
        competitors[existing] = {**competitors[existing], **competitor}
    return competitor


def process_meli_notification(event):
    topic = str(event.get("topic") or "")
    if topic not in {"items", "orders_v2", "catalog_item_competition_status"}:
        return
    seller_id = str(event.get("user_id") or "")
    payload = read_payload(include_catalog=False)
    account = next(
        (
            row
            for row in payload.get("accounts", [])
            if str(row.get("seller_id") or "") == seller_id and row.get("official")
        ),
        None,
    )
    if not account:
        return
    client = account_client(account)
    if topic == "orders_v2":
        sync_recent_sales(payload, account, client)
        account["last_webhook_at"] = now_label()
        write_payload(payload)
        return

    resource = str(event.get("resource") or "")
    match = re.fullmatch(r"/items/(MLB\d+)(?:/price_to_win)?", resource, flags=re.I)
    if not match:
        return
    item_id = match.group(1).upper()
    official = client.item(item_id)
    existing = next(
        (
            row
            for row in payload.get("catalog", [])
            if row.get("id") == item_id and row.get("account_id") == account.get("id")
        ),
        None,
    )
    before = dict(existing) if existing else None
    is_catalog = is_catalog_listing(official)
    competition = normalize_competition(client, account, official) if is_catalog else {}
    updated = synced_catalog_item(account, official, competition)
    updated["first_seen_at"] = (before or {}).get("first_seen_at") or now_label()
    updated["updated_at"] = now_label()
    updated["last_webhook_at"] = now_label()
    if existing is None:
        payload.setdefault("catalog", []).append(updated)
    else:
        existing.update(updated)

    if before and int(before.get("stock") or 0) > 0 and int(updated.get("stock") or 0) == 0:
        alert = stock_alert(f"stock-{account.get('id')}-{item_id}-{uuid.uuid4().hex[:8]}", account, updated)
        payload.setdefault("alerts", []).insert(0, alert)
        notify_alert(payload, alert)
    if before and before.get("status") != "losing" and updated.get("status") == "losing":
        alert = {
            "id": f"catalog-{account.get('id')}-{item_id}-{uuid.uuid4().hex[:8]}",
            "type": "catalog",
            "severity": "danger",
            "title": "Produto começou a perder catálogo",
            "message": (
                f"{updated.get('title')} deixou de vencer o catálogo. "
                f"Preço para ganhar: {brl_label(updated.get('price_to_win'))}."
            ),
            "account": updated.get("account"),
            "item_id": item_id,
            "sku": updated.get("sku"),
            "product": updated.get("title"),
            "created_at": now_label(),
            "read": False,
        }
        payload.setdefault("alerts", []).insert(0, alert)
        notify_alert(payload, alert)
    account["last_webhook_at"] = now_label()
    account["webhook_status"] = f"Última notificação processada: {topic} {item_id}"
    write_payload(payload)


def meli_notification_loop():
    while True:
        event = MELI_NOTIFICATION_QUEUE.get()
        try:
            process_meli_notification(event)
        except Exception:
            pass
        finally:
            MELI_NOTIFICATION_QUEUE.task_done()


def auto_scan_loop():
    interval = max(60, int(os.getenv("SCAN_INTERVAL_SECONDS", "300")))
    time.sleep(8)
    while True:
        try:
            sync_busy, interactive_busy = meli_background_work_busy()
            if sync_busy or interactive_busy:
                time.sleep(min(30, interval))
                continue
            payload = read_payload()
            scans = [item for item in payload.get("scan_items", []) if item.get("status", "active") == "active"]
            changed = False
            for scan in scans:
                try:
                    scan_meli_item(payload, scan.get("id"))
                    scan["auto_scan_status"] = "Scan automático atualizado"
                    scan["auto_scan_error"] = ""
                    scan["last_auto_scan_at"] = now_label()
                    scan["auto_scan_interval_seconds"] = interval
                    changed = True
                except Exception as exc:
                    scan["auto_scan_status"] = "Scan automático com erro"
                    scan["auto_scan_error"] = str(exc)
                    scan["last_auto_scan_at"] = now_label()
                    scan["auto_scan_interval_seconds"] = interval
                    changed = True
            if changed:
                payload["scan_items"] = payload.get("scan_items", [])
                write_payload(payload)
        except Exception:
            pass
        time.sleep(interval)


def account_sync_is_active(account):
    keys = {
        str((account or {}).get("id") or ""),
        str((account or {}).get("seller_id") or ""),
        str((account or {}).get("nickname") or ""),
    }
    with SYNC_LOCK:
        return bool(keys.intersection(ACTIVE_SYNC_ACCOUNTS))


def auto_official_sync_loop():
    interval = max(120, int(os.getenv("AUTO_SYNC_INTERVAL_SECONDS", "300")))
    startup_delay = max(20, int(os.getenv("AUTO_SYNC_STARTUP_DELAY_SECONDS", "45")))
    operations_every = max(1, int(os.getenv("AUTO_OPERATIONS_SYNC_EVERY_N_RUNS", "6")))
    run_count = 0
    time.sleep(startup_delay)
    while True:
        try:
            sync_busy, interactive_busy = meli_background_work_busy()
            if sync_busy or interactive_busy:
                time.sleep(min(30, interval))
                continue
            payload = read_payload()
            accounts = [
                account
                for account in payload.get("accounts", [])
                if account.get("official")
                and account.get("access_token")
                and account.get("status") == "connected"
                and not account_sync_is_active(account)
            ]
            changed = False
            run_count += 1
            for account in accounts:
                try:
                    result = refresh_official_account_items(payload, account, include_competition=False)
                    account["auto_sync_status"] = account.get("auto_refresh_status") or f"Atualização automática OK: {result.get('items', 0)} anúncios"
                    if account.get("operations_refresh_requested") or run_count % operations_every == 0:
                        client = account_client(account)
                        sync_recent_sales(payload, account, client)
                        sync_claims(payload, account, client)
                        account["operations_refresh_requested"] = False
                    account["auto_sync_error"] = ""
                    account["last_auto_sync_at"] = now_label()
                    changed = True
                except Exception as exc:
                    account["auto_sync_status"] = "Sincronização automática com erro"
                    account["auto_sync_error"] = str(exc)
                    account["last_auto_sync_at"] = now_label()
                    changed = True
            if changed:
                write_payload(payload)
        except Exception:
            pass
        time.sleep(interval)


def auto_catalog_competition_loop():
    interval = max(30, int(os.getenv("AUTO_COMPETITION_INTERVAL_SECONDS", "60")))
    startup_delay = max(30, int(os.getenv("AUTO_COMPETITION_STARTUP_DELAY_SECONDS", "75")))
    batch_size = max(1, int(os.getenv("AUTO_COMPETITION_BATCH_SIZE", "250")))
    workers = max(1, min(12, int(os.getenv("AUTO_COMPETITION_WORKERS", "8"))))
    time.sleep(startup_delay)
    while True:
        try:
            sync_busy, interactive_busy = meli_background_work_busy()
            if sync_busy or interactive_busy:
                time.sleep(min(15, interval))
                continue
            payload = read_payload()
            accounts = [
                account for account in payload.get("accounts", [])
                if account.get("official")
                and account.get("access_token")
                and account.get("status") == "connected"
                and not account_sync_is_active(account)
            ]
            jobs = []
            for account in accounts:
                rows = [
                    item for item in payload.get("catalog", [])
                    if item.get("official_source")
                    and item.get("account_id") == account.get("id")
                    and is_catalog_listing(item)
                    and item.get("meli_status") != "closed"
                ]
                cursor_key = f"competition_refresh_cursor_{account.get('id')}"
                selected, account[cursor_key] = rotating_batch(rows, account.get(cursor_key) or 0, batch_size)
                client = account_client(account)
                for item in selected:
                    jobs.append((account, client, item, dict(item)))

            completed_by_account = {}
            with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="meli-buybox") as executor:
                future_jobs = {
                    executor.submit(normalize_competition, client, account, item): (account, item, before)
                    for account, client, item, before in jobs
                }
                for future in as_completed(future_jobs):
                    account, item, before = future_jobs[future]
                    try:
                        competition = future.result()
                        item.update(competition)
                        item["status"], item["action"] = classified_catalog_status(
                            account, item, int(item.get("stock") or 0), True, competition
                        )
                        item["share"] = competition_share(competition.get("competition_status"), competition.get("visit_share"))
                        item["updated_at"] = now_label()
                        item.pop("competition_refresh_error", None)
                        completed_by_account[account.get("id")] = completed_by_account.get(account.get("id"), 0) + 1
                        if before.get("status") != "losing" and item.get("status") == "losing" and not item.get("internal_competition"):
                            alert_id = f"catalog-{account.get('id')}-{item.get('id')}-{now_label()[:10]}"
                            if not any(alert.get("id") == alert_id for alert in payload.get("alerts", [])):
                                alert = {
                                    "id": alert_id,
                                    "type": "catalog",
                                    "severity": "danger",
                                    "title": "Produto começou a perder catálogo",
                                    "message": f"{item.get('title')} perdeu a buybox. Vencedor: {item.get('winner_name') or '-'} por {brl(item.get('winner_price')) if item.get('winner_price') else '-' }.",
                                    "account": item.get("account"),
                                    "item_id": item.get("id"),
                                    "sku": item.get("sku"),
                                    "product": item.get("title"),
                                    "created_at": now_label(),
                                }
                                payload.setdefault("alerts", []).insert(0, alert)
                                notify_alert(payload, alert)
                    except Exception as exc:
                        item["competition_refresh_error"] = str(exc)
                        item["competition_checked_at"] = now_label()
            reclassify_internal_competition(payload)
            for account in accounts:
                checked = completed_by_account.get(account.get("id"), 0)
                account["catalog_refresh_status"] = f"Varredura rotativa: {checked} anúncios de catálogo verificados"
                account["last_catalog_refresh_at"] = now_label()
            if accounts:
                write_payload(payload)
        except Exception:
            pass
        time.sleep(interval)


def parse_brl_price(text):
    match = re.search(r"R\$\s*([\d\.]+)(?:,(\d{2}))?", text or "")
    if not match:
        return None
    whole = match.group(1).replace(".", "")
    cents = match.group(2) or "00"
    try:
        return float(f"{whole}.{cents}")
    except ValueError:
        return None


def visible_page_lines(html_text):
    text = re.sub(r"(?is)<(script|style|noscript).*?</\1>", "\n", html_text or "")
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</(p|div|section|article|li|h[1-6]|span|a)>", "\n", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = html.unescape(text)
    return [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if re.sub(r"\s+", " ", line).strip()]


def canonical_product_url(catalog_product_id, url=""):
    if catalog_product_id:
        return f"https://www.mercadolivre.com.br/p/{catalog_product_id}"
    if not url:
        return ""
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}{parsed.path}" if parsed.scheme and parsed.netloc else url.split("?", 1)[0].split("#", 1)[0]


def product_public_urls(client, catalog_product_id, item=None):
    urls = [canonical_product_url(catalog_product_id)]
    try:
        product = client.product(catalog_product_id)
        permalink = product.get("permalink") or product.get("url")
        if permalink:
            urls.append(canonical_product_url(catalog_product_id, permalink))
        product_title = product.get("name") or product.get("title") or ""
    except Exception:
        product_title = ""
    item_permalink = (item or {}).get("permalink") or ""
    if item_permalink:
        urls.append(canonical_product_url(catalog_product_id, item_permalink))
    deduped = []
    for url in urls:
        if url and url not in deduped:
            deduped.append(url)
    return deduped, product_title


def fetch_public_buybox(client, catalog_product_id, item=None):
    urls, product_title = product_public_urls(client, catalog_product_id, item)
    last_error = None
    for url in urls:
        try:
            return fetch_public_buybox_url(url, product_title)
        except Exception as exc:
            last_error = exc
    raise last_error or RuntimeError("Não foi possível consultar a página pública do Mercado Livre.")


def fetch_public_buybox_url(url, product_title=""):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0 Safari/537.36",
            "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
        },
    )
    with urllib.request.urlopen(req, timeout=float(os.getenv("MELI_PUBLIC_PAGE_TIMEOUT_SECONDS", "4"))) as response:
        html_text = response.read().decode("utf-8", errors="replace")
        final_url = response.geturl() or url
    lines = visible_page_lines(html_text)
    seller_name = ""
    price = None
    price_patterns = [
        r'itemprop="price"\s+content="(\d+(?:\.\d+)?)"',
        r'content="(\d+(?:\.\d+)?)"\s+itemprop="price"',
        r'ui-pdp-price__second-line[\s\S]{0,1200}?andes-money-amount__fraction[^>]*>([\d\.]+)<',
    ]
    for pattern_index, pattern in enumerate(price_patterns):
        match = re.search(pattern, html_text, flags=re.I)
        if match:
            try:
                parsed_price = float(match.group(1)) if pattern_index < 2 else parse_brl_price(f"R$ {match.group(1)}")
                if parsed_price:
                    price = parsed_price
                    break
            except (TypeError, ValueError):
                pass
    anchors = []
    if product_title:
        anchors.append(product_title)
    anchors.extend(["Novo |", "Adicionar aos favoritos", "Compartilhar"])
    search_text = "\n".join(lines)
    start = 0
    for anchor in anchors:
        pos = search_text.lower().find(anchor.lower())
        if pos >= 0:
            start = pos
            break
    end = len(search_text)
    for anchor in ["Opções de compra", "Ir para a compra", "O que você precisa saber", "Características"]:
        pos = search_text.find(anchor, start)
        if pos > start:
            end = min(end, pos)
    main_area = search_text[start:end]
    sold_match = re.search(r"Vendido por\s+([^\n]+)", main_area, flags=re.I)
    if sold_match:
        seller_name = re.split(r"\s{2,}|\+\d+\s+vendas|NF-e|verificada", sold_match.group(1).strip())[0].strip()
    if not price:
        price_candidates = []
        for line in main_area.splitlines():
            if re.search(r"\b\d+x\b|sem juros|meios de pagamento", line, flags=re.I):
                continue
            for match in re.finditer(r"R\$\s*[\d\.]+(?:,\d{2})?", line):
                parsed = parse_brl_price(match.group(0))
                if parsed:
                    price_candidates.append(parsed)
        if price_candidates:
            price = min(price_candidates)
    if not seller_name or not price:
        raise RuntimeError("A página pública não confirmou simultaneamente vendedor e preço da buybox.")
    return {
        "seller_name": seller_name,
        "price": price,
        "url": final_url,
        "source": "public_product_page",
    }


def public_purchase_options_winner(lines):
    best_index = next((i for i, line in enumerate(lines) if "Melhor preço" in line), -1)
    if best_index < 0:
        options_index = next((i for i, line in enumerate(lines) if "Outras opções de compra" in line or "Opções de compra" in line), -1)
        if options_index >= 0:
            best_index = options_index + 1
    if best_index < 0:
        return {}

    window_start = max(0, best_index - 8)
    window_end = min(len(lines), best_index + 18)
    window = lines[window_start:window_end]
    price = None
    seller_name = ""
    price_candidates = []
    for line in window:
        if re.search(r"R\$\s*[\d\.]+(?:,\d{2})?", line):
            parsed = parse_brl_price(line)
            if parsed:
                price_candidates.append(parsed)
    if price_candidates:
        price = min(price_candidates)
    for index, line in enumerate(window):
        match = re.search(r"Loja oficial\s+(.+)", line, flags=re.I)
        if match:
            seller_name = re.split(r"\s{2,}|\s\+\d|NF-e|verificada", match.group(1).strip())[0].strip()
            break
        if "Vendido por" in line:
            seller_name = line.split("Vendido por", 1)[1].strip(" :·-")
            break
        if re.search(r"Loja\s+oficial", line, flags=re.I) and index + 1 < len(window):
            seller_name = window[index + 1].strip()
            break
    return {
        "seller_name": seller_name or "",
        "price": price,
        "source": "public_purchase_options",
    }


def seller_nickname(client, seller_id):
    seller_id = str(seller_id or "").strip()
    if not seller_id:
        return ""
    ttl = max(300, int(os.getenv("SELLER_PROFILE_CACHE_SECONDS", "21600")))
    now = time.monotonic()
    with SELLER_PROFILE_LOCK:
        cached = SELLER_PROFILE_CACHE.get(seller_id)
        if cached and now - cached["time"] < ttl:
            return cached["nickname"]
    profile = client.user(seller_id)
    nickname = str((profile or {}).get("nickname") or f"Vendedor {seller_id}")
    with SELLER_PROFILE_LOCK:
        SELLER_PROFILE_CACHE[seller_id] = {"time": now, "nickname": nickname}
    return nickname


def winner_embedded_name(value):
    if not isinstance(value, dict):
        return ""
    seller = value.get("seller") if isinstance(value.get("seller"), dict) else {}
    official_store = value.get("official_store") if isinstance(value.get("official_store"), dict) else {}
    for candidate in (
        value.get("seller_name"),
        value.get("seller_nickname"),
        value.get("nickname"),
        value.get("official_store_name"),
        seller.get("nickname"),
        seller.get("name"),
        official_store.get("name"),
    ):
        name = clean_attribute_value(candidate)
        if name:
            return name
    return ""


def cached_public_buybox(client, catalog_product_id, item=None):
    cache_key = str(catalog_product_id or "")
    ttl = max(60, int(os.getenv("MELI_PUBLIC_BUYBOX_CACHE_SECONDS", "300")))
    now = time.monotonic()
    with PUBLIC_BUYBOX_CACHE_LOCK:
        cached = PUBLIC_BUYBOX_CACHE.get(cache_key)
    if cached and now - cached.get("time", 0) < ttl:
        return cached.get("value") or {}
    value = fetch_public_buybox(client, catalog_product_id, item)
    with PUBLIC_BUYBOX_CACHE_LOCK:
        PUBLIC_BUYBOX_CACHE[cache_key] = {"time": now, "value": value}
    return value


def resolve_winner_identity(client, winner, winner_item_id, catalog_product_id, source_item=None, expected_price=None):
    winner = winner if isinstance(winner, dict) else {}
    seller_id = str(winner.get("seller_id") or first_present(winner, ["seller.id"], "") or "")
    seller_name = winner_embedded_name(winner)
    winner_item = {}
    if winner_item_id:
        try:
            winner_item = client.item(winner_item_id) or {}
            seller_id = seller_id or str(winner_item.get("seller_id") or first_present(winner_item, ["seller.id"], "") or "")
            seller_name = seller_name or winner_embedded_name(winner_item)
        except Exception:
            winner_item = {}
    if catalog_product_id and (not seller_id or not seller_name):
        try:
            offers = client.product_winners(catalog_product_id)
            rows = offers.get("results") if isinstance(offers, dict) else offers
            for offer in rows or []:
                offer_item_id = str(offer.get("item_id") or offer.get("id") or "")
                if winner_item_id and offer_item_id != str(winner_item_id):
                    continue
                seller_id = seller_id or str(offer.get("seller_id") or first_present(offer, ["seller.id"], "") or "")
                seller_name = seller_name or winner_embedded_name(offer)
                if seller_id or seller_name:
                    break
        except Exception:
            pass
    if seller_id and not seller_name:
        try:
            seller_name = seller_nickname(client, seller_id)
        except Exception:
            seller_name = ""
    # The public page is only used to resolve the display name when its buy-box
    # price agrees with the price already returned by the official API.
    if not seller_name and catalog_product_id and expected_price not in (None, ""):
        try:
            public = cached_public_buybox(client, catalog_product_id, source_item)
            if abs(float(public.get("price") or 0) - float(expected_price or 0)) <= 0.02:
                seller_name = clean_attribute_value(public.get("seller_name"))
        except Exception:
            pass
    return seller_id, seller_name


def catalog_next_visible_offer(client, catalog_product_id, winner_item_id):
    catalog_product_id = str(catalog_product_id or "")
    if not catalog_product_id:
        return {}
    ttl = max(120, int(os.getenv("MELI_CATALOG_OFFERS_CACHE_SECONDS", "900")))
    now = time.monotonic()
    with CATALOG_OFFERS_CACHE_LOCK:
        cached = CATALOG_OFFERS_CACHE.get(catalog_product_id)
    if cached and now - cached.get("time", 0) < ttl:
        candidates = cached.get("candidates") or []
    else:
        response = client.product_winners(catalog_product_id)
        rows = response.get("results") if isinstance(response, dict) else response
        candidates = catalog_offer_candidates(rows or [], sort_by_price=True)
        with CATALOG_OFFERS_CACHE_LOCK:
            CATALOG_OFFERS_CACHE[catalog_product_id] = {"time": now, "candidates": candidates}
    candidate = next((row for row in candidates if str(row.get("item_id") or "") != str(winner_item_id or "")), None)
    if not candidate:
        return {}
    seller_id = str(candidate.get("seller_id") or "")
    seller = ""
    if seller_id:
        try:
            seller = seller_nickname(client, seller_id)
        except Exception:
            seller = f"Vendedor {seller_id}"
    return {
        "runner_up_item_id": candidate.get("item_id") or "",
        "runner_up_seller_id": seller_id,
        "runner_up_name": seller or (f"Anúncio {candidate.get('item_id')}" if candidate.get("item_id") else ""),
        "runner_up_price": candidate.get("price"),
        "runner_up_source": "catalog_visible_offer",
    }


def normalize_competition(client, account, item):
    item_id = item.get("id")
    catalog_product_id = item.get("catalog_product_id")
    data = {
        "competition_status": "not_checked",
        "competition_consistent": None,
        "competition_checked_at": now_label(),
        "winner_item_id": "",
        "winner_seller_id": "",
        "winner_name": "Indisponível",
        "winner_price": None,
        "winner_confirmed": False,
        "winner_source": "",
        "catalog_reference_seller_id": "",
        "catalog_reference_name": "",
        "catalog_reference_price": None,
        "price_to_win": None,
        "current_price": None,
        "visit_share": None,
        "competition_reason": "",
        "competitors_sharing_first_place": None,
        "runner_up_item_id": "",
        "runner_up_seller_id": "",
        "runner_up_name": "",
        "runner_up_price": None,
        "runner_up_source": "",
    }

    try:
        price = client.price_to_win(item_id)
        winner = price.get("winner") or {}
        data.update(
            {
                "competition_status": price.get("status") or "unknown",
                "competition_consistent": price.get("consistent"),
                "price_to_win": price.get("price_to_win"),
                "current_price": price.get("current_price"),
                "visit_share": price.get("visit_share"),
                "competitors_sharing_first_place": price.get("competitors_sharing_first_place"),
                "competition_reason": ", ".join(price.get("reason") or []),
            }
        )
        if isinstance(winner, dict) and winner.get("item_id"):
            winner_item_id = str(winner.get("item_id") or "")
            winner_seller_id, winner_name = resolve_winner_identity(
                client,
                winner,
                winner_item_id,
                catalog_product_id,
                item,
                winner.get("price"),
            )
            if str(winner_item_id) == str(item_id):
                winner_seller_id = str(account.get("seller_id") or winner_seller_id)
                winner_name = account.get("nickname") or winner_name or "Sua conta"
            winner_price = winner.get("price")
            data["winner_item_id"] = winner_item_id
            data["winner_seller_id"] = winner_seller_id
            data["winner_name"] = winner_name or (f"Vendedor {winner_seller_id}" if winner_seller_id else f"Anúncio {winner_item_id}")
            data["winner_price"] = winner_price
            data["winner_confirmed"] = winner_price not in (None, "")
            data["winner_source"] = "price_to_win_winner"
        elif price.get("status") in {"winning", "winner"}:
            data["winner_item_id"] = str(item_id or "")
            data["winner_seller_id"] = account.get("seller_id", "")
            data["winner_name"] = account.get("nickname", "Sua conta")
            data["winner_price"] = price.get("current_price")
            data["winner_confirmed"] = price.get("current_price") not in (None, "")
            data["winner_source"] = "price_to_win"
    except Exception as exc:
        data["competition_reason"] = str(exc)

    # Compatibility fallback for accounts where the product resource still exposes
    # buy_box_winner. It remains an official API source and is never inferred from
    # the lowest offer or from unrelated text on the public product page.
    if catalog_product_id and not data.get("winner_confirmed"):
        try:
            product = client.product(catalog_product_id)
            winner = (product or {}).get("buy_box_winner") or {}
            winner_item_id = str(winner.get("item_id") or "")
            winner_price = winner.get("price")
            if winner_item_id and winner_price not in (None, ""):
                winner_seller_id, winner_name = resolve_winner_identity(
                    client,
                    winner,
                    winner_item_id,
                    catalog_product_id,
                    item,
                    winner_price,
                )
                data.update(
                    {
                        "winner_item_id": winner_item_id,
                        "winner_seller_id": winner_seller_id,
                        "winner_name": winner_name or (f"Vendedor {winner_seller_id}" if winner_seller_id else f"Anúncio {winner_item_id}"),
                        "winner_price": winner_price,
                        "winner_confirmed": True,
                        "winner_source": "product_buy_box_winner",
                    }
                )
        except Exception:
            pass

    if not data.get("winner_confirmed"):
        data["winner_item_id"] = ""
        data["winner_seller_id"] = ""
        data["winner_name"] = "Aguardando API oficial"
        data["winner_price"] = None
    if data["competition_status"] == "not_listed" and not data.get("winner_seller_id"):
        data["winner_name"] = "Sem vencedor disponível"
    if data.get("winner_confirmed") and str(data.get("winner_item_id") or "") == str(item_id or ""):
        try:
            data.update(catalog_next_visible_offer(client, catalog_product_id, data.get("winner_item_id")))
        except Exception:
            pass
    return data


def classified_catalog_status(account, item, stock, is_catalog, competition):
    if item.get("status") == "paused":
        return "paused", "Anúncio pausado no Mercado Livre. Ative o anúncio e revise estoque, preço e políticas antes de disputar catálogo."
    if stock == 0:
        return "paused", "Anúncio importado sem estoque disponível. Pode estar pausado ou em risco de pausa por estoque zerado."
    if not is_catalog:
        return "winning", "Anúncio importado da API oficial. Sem vínculo de catálogo identificado neste retorno."

    own_seller_id = str(account.get("seller_id") or "")
    winner_seller_id = str(competition.get("winner_seller_id") or "")
    winner_name = str(competition.get("winner_name") or "")
    own_name = str(account.get("nickname") or "")
    status_text = str(competition.get("competition_status") or "").lower()
    try:
        own_price = float(item.get("price") or competition.get("current_price") or 0)
    except (TypeError, ValueError):
        own_price = 0
    try:
        winner_price = float(competition.get("winner_price") or 0)
    except (TypeError, ValueError):
        winner_price = 0
    try:
        price_to_win = float(competition.get("price_to_win") or 0)
    except (TypeError, ValueError):
        price_to_win = 0

    # price_to_win v2 is authoritative. Do not infer the position only from
    # prices: reputation, delivery and installments also decide the buy box.
    if status_text in {"winning", "winner"}:
        return "winning", "Sua conta está vencendo a buybox deste catálogo."
    if status_text in {"sharing_first_place", "sharing", "shared"}:
        return "sharing", "Sua conta está compartilhando a primeira posição deste catálogo."
    if status_text in {"competing", "losing", "not_winning", "listed", "not_listed"}:
        return "losing", "Sua conta não está vencendo este catálogo. Revise preço, reputação, entrega e condições comerciais."
    if winner_seller_id and own_seller_id and winner_seller_id == own_seller_id:
        return "winning", "Sua conta está vencendo a buybox deste catálogo."
    if winner_name and own_name and winner_name.strip().lower() == own_name.strip().lower():
        return "winning", "Sua conta está vencendo a buybox deste catálogo."
    if competition.get("competitors_sharing_first_place"):
        return "sharing", "Sua conta está compartilhando a primeira posição do catálogo."
    if price_to_win and own_price and own_price > price_to_win:
        return "losing", "Sua conta não está vencendo este catálogo. Revise preço, reputação e condições comerciais."
    if competition.get("winner_confirmed") and winner_price and own_price:
        if own_price > winner_price:
            return "losing", "Sua conta está acima do preço vencedor informado para a buybox."
        if abs(own_price - winner_price) < 0.01:
            return "sharing", "Sua conta está no mesmo preço da oferta vencedora; pode estar compartilhando a disputa."
        return "winning", "Sua conta está com preço abaixo da oferta vencedora informada; confirme reputação e frete."
    return "sharing", "Anúncio de catálogo importado da API oficial. Aguardando confirmação completa da buybox."


def competition_share(status, visit_share=None):
    status = str(status or "").lower()
    if status in {"winning", "winner"}:
        return 100
    if status in {"sharing_first_place", "sharing", "shared"}:
        return 50
    if status in {"competing", "losing", "not_winning", "listed", "not_listed"}:
        return 0
    return {"maximum": 100, "medium": 50, "minimum": 0}.get(str(visit_share or "").lower(), 0)


def synced_catalog_item(account, item, competition=None):
    stock = item_available_quantity(item)
    catalog_product_id = item.get("catalog_product_id") or "-"
    listing_type_id = item.get("listing_type_id") or first_present(item, ["listing_type.id", "listing_type"], "")
    shipping_logistic_type = item_flex_logistic_type(item)
    shipping_mode = first_present(item, ["shipping.mode"], "")
    is_catalog = is_catalog_listing(item)
    competition = competition or {}
    status, action = classified_catalog_status(account, item, stock, is_catalog, competition)
    package_values = package_values_from_item(item)
    identifier_values = source_clone_identifiers(item, {})
    return {
        "id": item.get("id"),
        "title": item.get("title") or item.get("id"),
        "account": account.get("nickname"),
        "account_id": account.get("id"),
        "official_source": True,
        "item_data_checked_at": now_label(),
        "thumbnail": item_thumbnail(item),
        "sku": item_sku(item),
        "brand": source_attribute_value(item, ["BRAND"]),
        "gtin": ", ".join(identifier_values),
        "variation_count": len(item.get("variations") or []),
        "catalog_product_id": catalog_product_id,
        "catalog_listing": is_catalog,
        "listing_type_id": listing_type_id,
        "shipping_logistic_type": shipping_logistic_type,
        "shipping_mode": shipping_mode,
        "free_shipping": bool(first_present(item, ["shipping.free_shipping"], False)),
        **package_values,
        "status": status,
        "share": competition_share(competition.get("competition_status"), competition.get("visit_share")) if is_catalog else 100,
        "price": item.get("price") or 0,
        "stock": stock,
        "competitor": "A sincronizar",
        "action": action,
        "meli_status": item.get("status", "-"),
        "permalink": item.get("permalink", ""),
        **competition,
    }


def shipping_quote_values(response):
    data = response if isinstance(response, dict) else {}
    coverage = first_present(data, ["coverage.all_country"], {})
    if not isinstance(coverage, dict):
        coverage = {}
    raw_cost = coverage.get("list_cost")
    if raw_cost in (None, ""):
        raw_cost = first_present(data, ["list_cost", "shipping_option.list_cost", "cost"], None)
    try:
        cost = round(float(raw_cost), 2) if raw_cost not in (None, "") else None
    except (TypeError, ValueError):
        cost = None
    currency = (
        coverage.get("currency_id")
        or data.get("currency_id")
        or first_present(data, ["shipping_option.currency_id"], "BRL")
        or "BRL"
    )
    billable_weight = first_present(
        data,
        ["billable_weight", "shipping_option.billable_weight", "coverage.all_country.billable_weight"],
        None,
    )
    return {"cost": cost, "currency": currency, "billable_weight": billable_weight}


def refresh_account_shipping_costs(payload, account, item_ids=None, limit=None):
    maximum = max(1, int(limit or os.getenv("AUTO_SHIPPING_COST_BATCH_SIZE", "20")))
    requested = {str(item_id) for item_id in (item_ids or []) if item_id}
    rows = [
        item
        for item in payload.get("catalog", [])
        if item.get("official_source")
        and item.get("account_id") == account.get("id")
        and item.get("id")
        and (not requested or item.get("id") in requested)
    ]
    if not rows:
        return []
    if requested:
        selected = rows[:maximum]
    else:
        cursor_key = f"shipping_cost_cursor_{account.get('id')}"
        selected, account[cursor_key] = rotating_batch(rows, account.get(cursor_key) or 0, maximum)
    client = account_client(account)
    results = []
    for item in selected:
        try:
            quote = client.item_shipping_cost(account.get("seller_id"), item.get("id"))
            values = shipping_quote_values(quote)
            item["shipping_cost"] = values["cost"]
            item["shipping_cost_currency"] = values["currency"]
            item["shipping_billable_weight"] = values["billable_weight"]
            item["shipping_cost_source"] = "Cotação oficial Mercado Livre"
            item["shipping_cost_status"] = "ok" if values["cost"] is not None else "not_available"
            item["shipping_cost_error"] = "" if values["cost"] is not None else "A API não informou uma cotação para este anúncio."
        except Exception as exc:
            item["shipping_cost_status"] = "error"
            item["shipping_cost_error"] = policy_error_message(exc, "a cotação de frete do anúncio")
        item["shipping_cost_updated_at"] = now_label()
        results.append(
            {
                "id": item.get("id"),
                "account_id": item.get("account_id"),
                "shipping_cost": item.get("shipping_cost"),
                "shipping_cost_currency": item.get("shipping_cost_currency") or "BRL",
                "shipping_billable_weight": item.get("shipping_billable_weight"),
                "shipping_cost_status": item.get("shipping_cost_status"),
                "shipping_cost_error": item.get("shipping_cost_error") or "",
                "shipping_cost_updated_at": item.get("shipping_cost_updated_at"),
            }
        )
    return results


def preserve_shipping_cost_snapshot(target, previous):
    for key in (
        "shipping_cost",
        "shipping_cost_currency",
        "shipping_billable_weight",
        "shipping_cost_source",
        "shipping_cost_status",
        "shipping_cost_error",
        "shipping_cost_updated_at",
    ):
        if key in (previous or {}):
            target[key] = previous[key]
    return target


def preserve_identifier_snapshot(target, previous):
    if not target.get("gtin") and previous.get("gtin"):
        target["gtin"] = previous.get("gtin")
    for key in ("gtin_status", "gtin_updated_at", "gtin_error"):
        if key in previous and key not in target:
            target[key] = previous[key]
    return target


def refresh_identifiers_for_items(payload, item_ids):
    maximum = max(1, int(os.getenv("MELI_GTIN_ON_DEMAND_LIMIT", "15")))
    requested = list(dict.fromkeys(str(item_id) for item_id in (item_ids or []) if item_id))[:maximum]
    catalog_by_id = {item.get("id"): item for item in payload.get("catalog", []) if item.get("id") in requested}
    accounts_by_id = {
        str(account.get("id")): account
        for account in payload.get("accounts", [])
        if account.get("official") and account.get("access_token")
    }
    results = []
    for item_id in requested:
        target = catalog_by_id.get(item_id)
        account = accounts_by_id.get(str((target or {}).get("account_id") or ""))
        if not target or not account:
            continue
        try:
            client = account_client(account)
            source = enrich_clone_source_item(client, clone_source_item(client, item_id))
            identifiers = source_clone_identifiers(source, clone_source_catalog_product(client, source))
            target["gtin"] = ", ".join(identifiers)
            target["gtin_status"] = "ok" if identifiers else "not_available"
            target["gtin_error"] = ""
        except Exception as exc:
            target["gtin_status"] = "error"
            target["gtin_error"] = str(exc)
        target["gtin_updated_at"] = now_label()
        results.append(
            {
                key: target.get(key)
                for key in ("id", "gtin", "gtin_status", "gtin_error", "gtin_updated_at")
            }
        )
    return results


def refresh_shipping_costs_for_items(payload, item_ids):
    maximum = max(1, int(os.getenv("MELI_SHIPPING_COST_ON_DEMAND_LIMIT", "25")))
    requested = list(dict.fromkeys(str(item_id) for item_id in (item_ids or []) if item_id))[:maximum]
    by_account = {}
    catalog_by_id = {item.get("id"): item for item in payload.get("catalog", []) if item.get("id") in requested}
    for item_id in requested:
        item = catalog_by_id.get(item_id)
        if item and item.get("account_id"):
            by_account.setdefault(item.get("account_id"), []).append(item_id)
    results = []
    for account_id, ids in by_account.items():
        account = next(
            (
                row
                for row in payload.get("accounts", [])
                if row.get("id") == account_id and row.get("official") and row.get("access_token")
            ),
            None,
        )
        if account:
            results.extend(refresh_account_shipping_costs(payload, account, ids, len(ids)))
    return results


def normalized_account_name(value):
    return normalized_attribute_label(value or "")


def reclassify_internal_competition(payload):
    accounts = [account for account in payload.get("accounts") or [] if account.get("official")]
    by_seller = {str(account.get("seller_id")): account for account in accounts if account.get("seller_id") not in (None, "")}
    by_name = {normalized_account_name(account.get("nickname")): account for account in accounts if account.get("nickname")}
    by_item = {
        str(item.get("id")): next(
            (
                account
                for account in accounts
                if account.get("id") == item.get("account_id") or account.get("nickname") == item.get("account")
            ),
            None,
        )
        for item in payload.get("catalog") or []
        if item.get("id")
    }
    changed = False
    for item in payload.get("catalog") or []:
        if not is_catalog_listing(item):
            continue
        winner_account = by_seller.get(str(item.get("winner_seller_id") or ""))
        if not winner_account:
            winner_account = by_item.get(str(item.get("winner_item_id") or ""))
        if not winner_account:
            winner_account = by_name.get(normalized_account_name(item.get("winner_name")))
        is_internal_loss = bool(winner_account) and (item.get("status") == "losing" or item.get("internal_competition"))
        if is_internal_loss:
            if not item.get("internal_competition"):
                item["external_status"] = item.get("status") or "losing"
            target_name = winner_account.get("nickname") or item.get("winner_name")
            updates = {
                "internal_competition": True,
                "internal_winner_account_id": winner_account.get("id"),
                "winner_name": target_name,
                "status": "sharing",
                "share": 50,
                "action": f"A oferta vencedora pertence à conta conectada {target_name}. Esta disputa não é tratada como perda externa.",
            }
            for key, value in updates.items():
                if item.get(key) != value:
                    item[key] = value
                    changed = True
        elif item.get("internal_competition"):
            item.pop("internal_competition", None)
            item.pop("internal_winner_account_id", None)
            restored = item.pop("external_status", None)
            if restored:
                item["status"] = restored
            changed = True
    return changed


def append_item_log(payload, item, user, action, changes=None, sale_id=None):
    logs = payload.setdefault("item_logs", [])
    logs.insert(
        0,
        {
            "id": f"log-{uuid.uuid4().hex[:10]}",
            "item_id": item.get("id"),
            "account": item.get("account"),
            "title": item.get("title"),
            "user": (user or {}).get("name") or (user or {}).get("email") or "Sistema",
            "action": action,
            "changes": changes or {},
            "sale_id": sale_id or "",
            "created_at": now_label(),
        },
    )
    payload["item_logs"] = logs[:1000]


def upsert_metric(payload, account, user_profile):
    reputation = user_profile.get("seller_reputation", {}) or {}
    metrics = reputation.get("metrics", {}) or {}
    claims = percent_rate((metrics.get("claims") or {}).get("rate"))
    cancellations = percent_rate((metrics.get("cancellations") or {}).get("rate"))
    late = percent_rate((metrics.get("delayed_handling_time") or {}).get("rate"))
    metric = {
        "account": account.get("nickname"),
        "claims": claims,
        "mediations": percent_rate((metrics.get("sales") or {}).get("completed")),
        "cancellations": cancellations,
        "late_shipments": late,
        "agency_score": max(0, round(100 - late, 2)),
        "flex_score": max(0, round(100 - late, 2)),
        "period": "Período informado pela API oficial",
    }
    existing = payload.get("metrics", [])
    for index, current in enumerate(existing):
        if current.get("account") == account.get("nickname"):
            existing[index] = metric
            break
    else:
        existing.append(metric)
    payload["metrics"] = existing


def stock_alert(alert_id, account, item):
    account_name = account.get("nickname") or item.get("account")
    sku = item.get("sku") or "-"
    item_id = item.get("id") or "-"
    return {
        "id": alert_id,
        "type": "stock",
        "severity": "critical",
        "title": "Produto sem estoque",
        "message": f"{item.get('title')} está com estoque zerado na conta {account_name}. SKU: {sku}. MLB: {item_id}.",
        "account": account_name,
        "item_id": item_id,
        "sku": sku,
        "product": item.get("title") or item_id,
        "channel": ["dashboard", "telegram"],
        "created_at": now_label(),
        "read": False,
    }


def product_context_for_alert(payload, alert):
    if alert.get("type") == "scan":
        return {}
    item_id = alert.get("item_id") or ""
    if not item_id and str(alert.get("id", "")).startswith("stock-"):
        item_id = str(alert.get("id", "")).split("-")[-1]
    match = next(
        (
            item
            for item in payload.get("catalog", [])
            if item.get("id") == item_id
            or (
                alert.get("account")
                and item.get("account") == alert.get("account")
                and item.get("title")
                and item.get("title") in (alert.get("message") or "")
            )
        ),
        None,
    )
    if not match:
        return {
            "account": alert.get("account") or "",
            "item_id": alert.get("item_id") or "",
            "sku": alert.get("sku") or "",
            "product": alert.get("product") or "",
        }
    alert.setdefault("item_id", match.get("id") or "")
    alert.setdefault("sku", match.get("sku") or "")
    alert.setdefault("product", match.get("title") or "")
    return {
        "account": match.get("account") or alert.get("account") or "",
        "item_id": match.get("id") or alert.get("item_id") or "",
        "sku": match.get("sku") or alert.get("sku") or "",
        "product": match.get("title") or alert.get("product") or "",
    }


def telegram_alert_text(payload, alert):
    header = f"CompeTIDOR | {alert.get('title') or 'Alerta'}"
    lines = [
        header,
        "-" * min(42, max(18, len(header))),
    ]
    context = product_context_for_alert(payload, alert)
    if alert.get("type") != "scan" and any(context.values()):
        if context.get("product"):
            lines.append(f"Produto: {context['product']}")
        if context.get("account"):
            lines.append(f"Loja: {context['account']}")
        if context.get("item_id"):
            lines.append(f"MLB: {context['item_id']}")
        if context.get("sku"):
            lines.append(f"SKU: {context['sku']}")
        lines.append("")
    elif alert.get("account"):
        lines.append(f"Conta: {alert.get('account')}")
        lines.append("")
    lines.extend(
        [
            "Detalhe:",
            str(alert.get("message") or "-"),
            "",
            f"Horário: {alert.get('created_at') or now_label()}",
        ]
    )
    return "\n".join(lines)


def notify_alert(payload, alert):
    results = {}
    for user_id, config in notification_targets(payload):
        telegram = config.get("telegram") or {}
        if not telegram.get("enabled") or not telegram.get("bot_token") or not telegram.get("chat_id"):
            continue
        if not telegram_type_enabled(config, alert.get("type")):
            continue
        try:
            results[user_id] = Notifier(config).send_telegram(telegram_alert_text(payload, alert))
        except Exception as exc:
            results[user_id] = {"ok": False, "error": str(exc)}
    if results:
        alert["telegram_results"] = results
    return results or None


def send_telegram_message_to_users(payload, alert_type, text):
    results = {}
    for user_id, config in notification_targets(payload):
        telegram = config.get("telegram") or {}
        if not telegram.get("enabled") or not telegram.get("bot_token") or not telegram.get("chat_id"):
            continue
        if not telegram_type_enabled(config, alert_type):
            continue
        try:
            results[user_id] = Notifier(config).send_telegram(text)
        except Exception as exc:
            results[user_id] = {"ok": False, "error": str(exc)}
    return results


def add_stock_alerts(payload, account, items):
    created = 0
    existing_ids = {alert.get("id") for alert in payload.get("alerts", [])}
    for item in items:
        if item.get("stock") != 0:
            continue
        alert_id = f"stock-{account.get('id')}-{item.get('id')}"
        if alert_id in existing_ids:
            continue
        alert = stock_alert(alert_id, account, item)
        payload.setdefault("alerts", []).insert(0, alert)
        notify_alert(payload, alert)
        existing_ids.add(alert_id)
        created += 1
    return created


def ensure_stock_alerts(payload, notify=False):
    created = 0
    existing_ids = {alert.get("id") for alert in payload.get("alerts", [])}
    accounts = payload.get("accounts", [])
    for item in payload.get("catalog", []):
        if item.get("stock") != 0:
            continue
        account = next(
            (
                account
                for account in accounts
                if account.get("id") == item.get("account_id") or account.get("nickname") == item.get("account")
            ),
            {"id": item.get("account_id") or item.get("account"), "nickname": item.get("account")},
        )
        alert_id = f"stock-{account.get('id')}-{item.get('id')}"
        if alert_id in existing_ids:
            continue
        alert = stock_alert(alert_id, account, item)
        payload.setdefault("alerts", []).insert(0, alert)
        existing_ids.add(alert_id)
        created += 1
        if notify:
            notify_alert(payload, alert)
    return created


def enrich_product_alerts(payload):
    changed = False
    for alert in payload.get("alerts", []):
        if alert.get("type") == "scan":
            continue
        before = (alert.get("item_id"), alert.get("sku"), alert.get("product"))
        context = product_context_for_alert(payload, alert)
        if context.get("item_id"):
            alert["item_id"] = context["item_id"]
        if context.get("sku"):
            alert["sku"] = context["sku"]
        if context.get("product"):
            alert["product"] = context["product"]
        changed = changed or before != (alert.get("item_id"), alert.get("sku"), alert.get("product"))
    return changed


def sync_recent_sales(payload, account, client):
    period, date_from, date_to = current_month_window()
    try:
        orders = []
        offset = 0
        page_size = 50
        max_orders = int(os.getenv("MELI_MONTHLY_ORDERS_LIMIT", "500"))
        while len(orders) < max_orders:
            data = client.seller_orders(account.get("seller_id"), limit=page_size, offset=offset, date_from=date_from, date_to=date_to)
            batch = data.get("results", []) or []
            orders.extend(batch)
            total = (data.get("paging") or {}).get("total") or len(orders)
            if not batch or len(orders) >= total:
                break
            offset += page_size
    except Exception as exc:
        account["sales_sync_status"] = policy_error_message(exc, "a leitura das vendas reais do mês")
        mark_monthly_revenue_error(payload, account, period, account["sales_sync_status"])
        return []
    rows = []
    revenue_total = 0.0
    revenue_orders = 0
    ignored_statuses = {"cancelled", "canceled", "invalid"}
    catalog_by_id = {item.get("id"): item for item in payload.get("catalog", []) if item.get("id")}
    for order in orders:
        status = str(order.get("status") or "").lower()
        order_total = float(order.get("total_amount") or order.get("paid_amount") or 0)
        if order_total > 0 and status not in ignored_statuses:
            revenue_total += order_total
            revenue_orders += 1
        order_items = order.get("order_items") or []
        for order_item in order_items or [{}]:
            item = order_item.get("item") or {}
            quantity = order_item.get("quantity") or 1
            unit_price = order_item.get("unit_price") or order_item.get("full_unit_price") or 0
            line_total = round(float(unit_price or 0) * int(quantity or 1), 2)
            sale = {
                "id": f"{order.get('id')}-{item.get('id') or len(rows)}",
                "order_id": order.get("id"),
                "account": account.get("nickname"),
                "product": item.get("title") or "Produto sem título",
                "item_id": item.get("id") or "",
                "thumbnail": item.get("thumbnail") or item.get("secure_thumbnail") or (catalog_by_id.get(item.get("id")) or {}).get("thumbnail") or "",
                "sku": item.get("seller_sku") or item.get("seller_custom_field") or "-",
                "quantity": quantity,
                "unit_price": unit_price,
                "total": line_total,
                "order_total": order_total,
                "channel": order.get("context", {}).get("channel") or order.get("tags", ["Mercado Livre"])[0],
                "status": order.get("status") or "-",
                "date": order.get("date_created") or order.get("last_updated") or now_label(),
            }
            rows.append(sale)
    existing = [
        item
        for item in payload.get("recent_sales", [])
        if item.get("account") != account.get("nickname")
    ]
    payload["recent_sales"] = sorted([*rows, *existing], key=lambda item: item.get("date") or "", reverse=True)[:80]
    today = datetime.now(APP_TZ).date()
    by_sku = {}
    for sale in rows:
        sold_at = parse_meli_datetime(sale.get("date"))
        if not sold_at or sold_at.date() != today:
            continue
        sku = str(sale.get("sku") or "-")
        key = (sku, sale.get("item_id") or "")
        summary = by_sku.setdefault(
            key,
            {
                "account": account.get("nickname"),
                "sku": sku,
                "item_id": sale.get("item_id") or "",
                "product": sale.get("product") or "Produto vendido",
                "thumbnail": sale.get("thumbnail") or "",
                "units": 0,
                "revenue": 0.0,
                "date": today.isoformat(),
            },
        )
        summary["units"] += int(sale.get("quantity") or 0)
        summary["revenue"] = round(float(summary["revenue"]) + float(sale.get("total") or 0), 2)
    daily = [row for row in payload.get("daily_sku_sales", []) if row.get("account") != account.get("nickname")]
    daily.extend(by_sku.values())
    payload["daily_sku_sales"] = daily
    account["sales_sync_status"] = f"{revenue_orders} pedidos reais sincronizados no mês"
    upsert_monthly_revenue(payload, account, revenue_total, revenue_orders, period, account["sales_sync_status"])
    sync_pending_shipments_from_orders(payload, account, orders)
    return rows


def sync_claims(payload, account, client):
    try:
        data = client.seller_claims(account.get("seller_id"), 50, 0)
        rows = data.get("data") or data.get("results") or data.get("claims") or []
    except Exception as exc:
        account["claims_sync_status"] = policy_error_message(exc, "a leitura das reclamações")
        if "404" in str(exc) or "resource not found" in str(exc).lower():
            account["claims_sync_status"] = (
                "Reclamações não retornadas pela API oficial nesta credencial. "
                "Verifique no painel do Mercado Livre se a aplicação tem permissão de pós-venda/reclamações."
            )
        return []
    details = []
    open_count = 0
    mediation_count = 0
    detail_limit = max(0, int(os.getenv("MELI_CLAIM_DETAIL_LIMIT", "20")))
    for index, claim in enumerate(rows):
        status = str(claim.get("status") or claim.get("stage") or "").lower()
        claim_type = str(claim.get("type") or "").lower()
        claim_stage = str(claim.get("stage") or "").lower()
        if "medi" in status or "medi" in claim_type or claim_stage == "dispute":
            mediation_count += 1
        elif status not in {"closed", "resolved", "cancelled", "canceled"}:
            open_count += 1
        detail = {}
        if index < detail_limit and claim.get("id"):
            try:
                detail = client.claim_detail(claim.get("id")) or {}
            except Exception:
                detail = {}
        details.append(
            {
                "id": claim.get("id") or claim.get("claim_id") or "-",
                "account": account.get("nickname"),
                "status": claim.get("status") or claim.get("stage") or "-",
                "subject": detail.get("title") or claim.get("type") or claim.get("reason") or "Reclamação Mercado Livre",
                "description": detail.get("problem") or detail.get("description") or claim.get("description") or claim.get("detail") or claim.get("reason") or "",
                "created_at": claim.get("date_created") or claim.get("created_at") or now_label(),
                "due_date": detail.get("due_date") or "",
                "order_id": claim.get("resource_id") if claim.get("resource") == "order" else "",
            }
        )
    payload["claim_details"] = [item for item in payload.get("claim_details", []) if item.get("account") != account.get("nickname")]
    payload["claim_details"].extend(details[:50])
    claims = [item for item in payload.get("claims", []) if item.get("account") != account.get("nickname")]
    claims.append({"account": account.get("nickname"), "open": open_count, "mediations": mediation_count, "updated_at": now_label(), "details": details[:20]})
    payload["claims"] = claims
    account["claims_sync_status"] = f"{len(details)} reclamações sincronizadas"
    return details


def sync_pending_shipments_from_orders(payload, account, orders):
    pending = []
    pending_statuses = {"paid", "confirmed", "payment_required", "partially_paid"}
    final_statuses = {"cancelled", "canceled", "invalid"}
    final_shipping_statuses = {"shipped", "delivered", "cancelled", "canceled", "not_delivered"}
    sla_lookup_limit = max(0, int(os.getenv("MELI_PENDING_SLA_LIMIT", "100")))
    sla_lookups = 0
    for order in orders:
        status = str(order.get("status") or "").lower()
        tags = [str(tag).lower() for tag in order.get("tags", []) or []]
        if status in final_statuses or status not in pending_statuses:
            continue
        if any(tag in {"shipped", "delivered", "cancelled", "canceled"} for tag in tags):
            continue
        shipping = order.get("shipping") or {}
        shipment_id = shipment_id_from_order(order)
        official_shipment = {}
        if shipment_id not in (None, "", 0, "0"):
            try:
                official_shipment = account_client(account).shipment(shipment_id) or {}
            except Exception:
                official_shipment = {}
        if official_shipment:
            shipping_status = str(official_shipment.get("status") or "").lower()
        else:
            shipping_status = str(shipping.get("status") or first_present(shipping, ["status_history.status"], "") or "").lower()
        shipping_substatus = str(official_shipment.get("substatus") or first_present(shipping, ["substatus", "status_history.substatus"], "") or "").lower()
        if shipping_status in final_shipping_statuses or shipping_substatus in final_shipping_statuses:
            continue
        deadline = ""
        sla_status = ""
        if shipment_id not in (None, "", 0, "0") and sla_lookups < sla_lookup_limit:
            try:
                sla = account_client(account).shipment_sla(shipment_id) or {}
                deadline = sla.get("expected_date") or ""
                sla_status = sla.get("status") or ""
            except Exception:
                deadline = ""
            sla_lookups += 1
        if not deadline:
            deadline = (
                first_present(official_shipment, ["lead_time.estimated_handling_limit.date", "shipping_option.estimated_handling_limit.date"])
                or first_present(shipping, ["estimated_handling_limit.date", "estimated_handling_limit", "shipping_option.estimated_handling_limit"])
                or "Aguardando SLA oficial"
            )
        pending.append(
            {
                "account": account.get("nickname"),
                "order_id": order.get("id") or "-",
                "buyer": first_present(order, ["buyer.nickname", "buyer.first_name"], "Comprador Mercado Livre"),
                "deadline": deadline,
                "time_left": dispatch_time_left(deadline),
                "shipment_id": shipment_id or "-",
                "sla_status": sla_status,
            }
        )
    payload["pending_shipments"] = [item for item in payload.get("pending_shipments", []) if item.get("account") != account.get("nickname")]
    payload["pending_shipments"].extend(pending[:200])
    return pending


COMPETITION_FIELDS = (
    "competition_status",
    "competition_consistent",
    "competition_checked_at",
    "winner_item_id",
    "winner_seller_id",
    "winner_name",
    "winner_price",
    "winner_confirmed",
    "winner_source",
    "catalog_reference_seller_id",
    "catalog_reference_name",
    "catalog_reference_price",
    "price_to_win",
    "current_price",
    "visit_share",
    "competition_reason",
    "competitors_sharing_first_place",
    "public_buybox_error",
)


def competition_snapshot(item):
    snapshot = {field: item.get(field) for field in COMPETITION_FIELDS if field in item}
    if snapshot.get("winner_source") in {
        "catalog_lowest_active_offer",
        "catalog_reference",
        "products_items_winner_marker",
        "public_purchase_options",
        "public_product_page",
    }:
        snapshot.update(
            {
                "winner_seller_id": "",
                "winner_name": "Não exposto pela API oficial",
                "winner_price": None,
                "winner_confirmed": False,
                "winner_source": "",
                "competition_reason": (
                    "A API oficial informa o status da disputa e o preço para ganhar, "
                    "mas este registro antigo não possuía vencedor confirmado."
                ),
            }
        )
    return snapshot


def upsert_monthly_revenue(payload, account, amount, orders_count, period, status):
    monthly = payload.setdefault("monthly_revenue", {"period": period, "accounts": {}})
    if monthly.get("period") != period:
        monthly["period"] = period
        monthly["accounts"] = {}
    monthly.setdefault("accounts", {})[account.get("id") or account.get("nickname")] = {
        "account": account.get("nickname"),
        "amount": round(float(amount or 0), 2),
        "orders_count": int(orders_count or 0),
        "source": "Pedidos oficiais Mercado Livre",
        "sync_status": status,
        "updated_at": now_label(),
    }


def mark_monthly_revenue_error(payload, account, period, status):
    monthly = payload.setdefault("monthly_revenue", {"period": period, "accounts": {}})
    if monthly.get("period") != period:
        monthly["period"] = period
        monthly["accounts"] = {}
    accounts = monthly.setdefault("accounts", {})
    record = accounts.setdefault(
        account.get("id") or account.get("nickname"),
        {
            "account": account.get("nickname"),
            "amount": 0,
            "orders_count": 0,
            "source": "Pedidos oficiais Mercado Livre",
        },
    )
    record["sync_status"] = status
    record["updated_at"] = now_label()


def sync_official_account(payload, account_id, limit=None, progress=None):
    def update_progress(stage, completed=0, total=0, message=""):
        if progress:
            progress(stage, completed, total, message)

    account = next(
        (
            item
            for item in payload.get("accounts", [])
            if item.get("id") == account_id or str(item.get("seller_id")) == str(account_id) or item.get("nickname") == account_id
        ),
        None,
    )
    if not account:
        raise RuntimeError("Conta não encontrada.")
    if not account.get("official"):
        raise RuntimeError("Apenas contas com OAuth oficial podem ser sincronizadas.")

    update_progress("preparing", 0, 0, "Validando a conta no Mercado Livre.")
    client = account_client(account)
    user_profile = client.user(account["seller_id"])
    update_progress("listing", 0, 0, "Listando todos os anúncios da conta no Mercado Livre.")
    item_ids = client.seller_all_items(
        account["seller_id"],
        max_items=None if limit in (None, "all") else int(limit),
        progress=lambda found, status: update_progress(
            "listing",
            found,
            0,
            f"Listando anúncios no Mercado Livre: {found} identificados (status {status}).",
        ),
    )
    update_progress("items", 0, len(item_ids), f"{len(item_ids)} anúncios encontrados; importando os dados oficiais.")
    imported = []
    existing_by_id = {
        item.get("id"): item
        for item in payload.get("catalog", [])
        if item.get("official_source") and item.get("account_id") == account.get("id") and item.get("id")
    }
    competition_inline_limit = max(0, int(os.getenv("MELI_SYNC_COMPETITION_INLINE_LIMIT", "0")))
    batch_size = max(1, min(20, int(os.getenv("MELI_ITEM_BULK_SIZE", "20"))))
    batches = [(start, item_ids[start : start + batch_size]) for start in range(0, len(item_ids), batch_size)]

    def fetch_item_batch(batch):
        batch_start, batch_ids = batch
        while True:
            with ASYNC_OPERATION_JOBS_LOCK:
                clone_busy = bool(ACTIVE_CLONE_OPERATIONS)
            if not clone_busy:
                break
            time.sleep(0.25)
        try:
            batch_items = client.items_bulk(batch_ids)
        except Exception:
            batch_items = []

            def fetch_single(item_id):
                try:
                    return client.item(item_id)
                except Exception:
                    return None

            fallback_workers = max(1, min(4, int(os.getenv("MELI_SYNC_FALLBACK_WORKERS", "4"))))
            with ThreadPoolExecutor(max_workers=fallback_workers, thread_name_prefix="meli-item-retry") as fallback:
                batch_items = [item for item in fallback.map(fetch_single, batch_ids) if item]
        return batch_start, batch_items, len(batch_ids)

    batch_workers = max(1, min(8, int(os.getenv("MELI_SYNC_BATCH_WORKERS", "4"))))
    fetched_count = 0
    with ThreadPoolExecutor(max_workers=batch_workers, thread_name_prefix="meli-items") as executor:
        futures = [executor.submit(fetch_item_batch, batch) for batch in batches]
        for future in as_completed(futures):
            batch_start, batch_items, planned_count = future.result()
            for batch_offset, item in enumerate(batch_items):
                item_id = item.get("id")
                index = batch_start + batch_offset
                if not item_id:
                    continue
                is_catalog = is_catalog_listing(item)
                if is_catalog and index < competition_inline_limit:
                    competition = normalize_competition(client, account, item)
                else:
                    competition = competition_snapshot(existing_by_id.get(item_id, {}))
                row = synced_catalog_item(account, item, competition)
                previous = existing_by_id.get(item_id, {})
                preserve_shipping_cost_snapshot(row, previous)
                preserve_identifier_snapshot(row, previous)
                row["first_seen_at"] = previous.get("first_seen_at") or now_label()
                imported.append(row)
            fetched_count += planned_count
            update_progress(
                "items",
                min(fetched_count, len(item_ids)),
                len(item_ids),
                f"Importando e organizando anúncios: {min(fetched_count, len(item_ids))} de {len(item_ids)}.",
            )

    update_progress("saving", len(imported), len(item_ids), "Salvando os anúncios importados.")
    payload["catalog"] = [
        item
        for item in payload.get("catalog", [])
        if not (item.get("official_source") and item.get("account_id") == account.get("id"))
    ]
    payload["catalog"].extend(imported)
    account["last_sync"] = now_label()
    account["sync_status"] = f"{len(imported)} anúncios importados da API oficial"
    account["sync_total_item_ids"] = len(item_ids)
    upsert_metric(payload, account, user_profile)
    # Estoque zerado antigo não gera alerta na sincronização completa; alertas vêm de transição no refresh automático.
    return {"account": public_account(account), "items": len(imported), "catalog": imported}


def rotating_batch(rows, cursor, batch_size):
    if not rows or batch_size <= 0:
        return [], 0
    count = min(int(batch_size), len(rows))
    cursor = int(cursor or 0) % len(rows)
    end = cursor + count
    if end <= len(rows):
        selected = rows[cursor:end]
    else:
        selected = [*rows[cursor:], *rows[: end - len(rows)]]
    next_cursor = 0 if count >= len(rows) else end % len(rows)
    return selected, next_cursor


def refresh_official_account_items(payload, account, batch_size=None, include_competition=True):
    batch_size = max(1, int(batch_size or os.getenv("AUTO_REFRESH_BATCH_SIZE", "1000")))
    rows = [
        item
        for item in payload.get("catalog", [])
        if item.get("official_source") and item.get("account_id") == account.get("id") and item.get("id")
    ]
    if not rows:
        return {"items": 0, "alerts": 0}
    cursor_key = f"auto_refresh_cursor_{account.get('id')}"
    cursor = int(account.get(cursor_key) or 0)
    selected, account[cursor_key] = rotating_batch(rows, cursor, batch_size)

    catalog_rows = [item for item in rows if is_catalog_listing(item) and item.get("meli_status") != "closed"]
    competition_limit = max(0, int(os.getenv("AUTO_COMPETITION_BATCH_SIZE", "200"))) if include_competition else 0
    competition_cursor_key = f"competition_refresh_cursor_{account.get('id')}"
    competition_cursor = int(account.get(competition_cursor_key) or 0)
    competition_selected, account[competition_cursor_key] = rotating_batch(
        catalog_rows,
        competition_cursor,
        competition_limit,
    )

    client = account_client(account)
    refreshed = 0
    prior_by_id = {item.get("id"): dict(item) for item in [*selected, *competition_selected]}
    official_by_id = {}
    bulk_size = max(1, min(20, int(os.getenv("MELI_ITEM_BULK_SIZE", "20"))))
    selected_ids = [item.get("id") for item in selected if item.get("id")]
    for start in range(0, len(selected_ids), bulk_size):
        batch_ids = selected_ids[start : start + bulk_size]
        try:
            for official in client.items_bulk(batch_ids):
                if official.get("id"):
                    official_by_id[official["id"]] = official
        except Exception:
            for item_id in batch_ids:
                try:
                    official_by_id[item_id] = client.item(item_id)
                except Exception:
                    pass
    for current in selected:
        try:
            official = official_by_id.get(current.get("id"))
            if not official:
                raise RuntimeError("Anúncio não retornado pelo lote oficial nesta rodada.")
            competition = competition_snapshot(current)
            updated = synced_catalog_item(account, official, competition)
            preserve_identifier_snapshot(updated, current)
            preserve_shipping_cost_snapshot(updated, current)
            current.update(updated)
            current["updated_at"] = now_label()
            refreshed += 1
        except Exception as exc:
            current["auto_refresh_error"] = str(exc)
            current["auto_refresh_at"] = now_label()

    competition_checked = 0
    request_delay = max(0.0, float(os.getenv("AUTO_COMPETITION_REQUEST_DELAY_SECONDS", "0.05")))
    for current in competition_selected:
        try:
            competition = normalize_competition(client, account, current)
            current.update(competition)
            current["status"], current["action"] = classified_catalog_status(
                account,
                current,
                int(current.get("stock") or 0),
                True,
                competition,
            )
            current["share"] = competition_share(
                competition.get("competition_status"),
                competition.get("visit_share"),
            )
            current["updated_at"] = now_label()
            current.pop("competition_refresh_error", None)
            competition_checked += 1
        except Exception as exc:
            current["competition_refresh_error"] = str(exc)
            current["competition_checked_at"] = now_label()
        if request_delay:
            time.sleep(request_delay)

    reclassify_internal_competition(payload)
    alerts_created = 0
    existing_alert_ids = {alert.get("id") for alert in payload.get("alerts", [])}
    for item in selected:
        before = prior_by_id.get(item.get("id")) or {}
        if int(before.get("stock") or 0) > 0 and int(item.get("stock") or 0) == 0:
            alert_id = f"stock-{account.get('id')}-{item.get('id')}-{now_label()[:10]}"
            if alert_id not in existing_alert_ids:
                alert = stock_alert(alert_id, account, item)
                payload.setdefault("alerts", []).insert(0, alert)
                notify_alert(payload, alert)
                alerts_created += 1
                existing_alert_ids.add(alert_id)
    for item in competition_selected:
        before = prior_by_id.get(item.get("id")) or {}
        if (
            before
            and before.get("status") != item.get("status")
            and item.get("status") == "losing"
            and not item.get("internal_competition")
        ):
            alert_id = f"catalog-{account.get('id')}-{item.get('id')}-{now_label()[:10]}"
            existing_ids = {alert.get("id") for alert in payload.get("alerts", [])}
            if alert_id not in existing_ids:
                alert = {
                    "id": alert_id,
                    "type": "catalog",
                    "severity": "danger",
                    "title": "Produto começou a perder catálogo",
                    "message": f"{item.get('title')} perdeu a buybox. Vencedor: {item.get('winner_name') or '-'} por {item.get('winner_price') or '-'}.",
                    "account": item.get("account"),
                    "item_id": item.get("id"),
                    "sku": item.get("sku"),
                    "product": item.get("title"),
                    "created_at": now_label(),
                }
                payload.setdefault("alerts", []).insert(0, alert)
                notify_alert(payload, alert)
                alerts_created += 1

    shipping_quotes = refresh_account_shipping_costs(payload, account)
    account["auto_refresh_status"] = (
        f"Atualização automática OK: {refreshed}/{len(rows)} anúncios revisados; "
        f"{competition_checked} disputas de catálogo recalculadas; "
        f"{len(shipping_quotes)} cotações de frete atualizadas"
    )
    account["last_auto_refresh_at"] = now_label()
    return {"items": refreshed, "total": len(rows), "alerts": alerts_created, "shipping_quotes": len(shipping_quotes)}


def enqueue_official_sync(account_id, limit=None, reason="manual"):
    requested_account_id = str(account_id or "")
    payload = read_payload()
    account = next(
        (
            item
            for item in payload.get("accounts", [])
            if item.get("id") == requested_account_id
            or str(item.get("seller_id")) == requested_account_id
            or item.get("nickname") == requested_account_id
        ),
        None,
    )
    if not account:
        raise RuntimeError("Conta não encontrada.")
    account_id = str(account.get("id") or account.get("seller_id") or requested_account_id)
    with SYNC_LOCK:
        if account_id in ACTIVE_SYNC_ACCOUNTS:
            return {"queued": False, "status": "running", "message": "Sincronização desta conta já está em andamento."}
        ACTIVE_SYNC_ACCOUNTS.add(account_id)
    account["sync_status"] = "Sincronização aguardando a fila do servidor"
    account["sync_requested_at"] = now_label()
    account["sync_reason"] = reason
    queued_total = int(account.get("sync_total_item_ids") or 0)
    set_sync_progress(
        account_id,
        {
            "status": "running",
            "stage": "queued",
            "completed": 0,
            "total": queued_total,
            "percent": 0,
            "message": "Sincronização adicionada à fila; aguardando disponibilidade do servidor.",
            "started_at": now_label(),
            "started_epoch": time.time(),
            "updated_epoch": time.time(),
            "limit": limit,
            "reason": reason,
        },
        force=True,
    )
    write_payload(payload)

    def worker():
        def progress(stage, completed=0, total=0, message=""):
            completed = max(0, int(completed or 0))
            total = max(0, int(total or 0))
            ratio = min(1.0, completed / total) if total else 0
            if stage == "completed":
                percent = 100.0
            elif stage == "operations":
                percent = 99.0
            elif stage == "saving":
                percent = 97.0
            elif stage == "items":
                percent = round(5 + ratio * 91, 1)
            elif stage == "listing":
                percent = 4.0
            elif stage == "preparing":
                percent = 2.0
            else:
                percent = 0.0
            previous = sync_progress_snapshot().get(account_id, {})
            started_epoch = float(previous.get("started_epoch") or time.time())
            elapsed = max(0.0, time.time() - started_epoch)
            eta_seconds = None
            if 5 <= percent < 97 and elapsed >= 5:
                eta_seconds = max(0, round((elapsed / (percent / 100.0)) - elapsed))
            set_sync_progress(
                account_id,
                {
                    **previous,
                    "status": "completed" if stage == "completed" else "running",
                    "stage": stage,
                    "completed": completed,
                    "total": total,
                    "percent": min(100, percent),
                    "message": message,
                    "eta_seconds": eta_seconds,
                    "updated_epoch": time.time(),
                },
                force=stage in {"completed", "error", "saving", "operations"},
            )

        try:
            with SYNC_WORKER_SEMAPHORE:
                set_sync_progress(
                    account_id,
                    {
                        "status": "running",
                        "stage": "preparing",
                        "message": "Iniciando a sincronização desta conta.",
                        "started_at": now_label(),
                        "started_epoch": time.time(),
                        "updated_epoch": time.time(),
                    },
                    force=True,
                )
                payload_inner = read_payload()
                result = sync_official_account(payload_inner, account_id, limit, progress=progress)
                refreshed = next(
                    (
                        item
                        for item in payload_inner.get("accounts", [])
                        if item.get("id") == account_id or str(item.get("seller_id")) == account_id
                    ),
                    None,
                )
                if refreshed:
                    refreshed["sync_status"] = f"Sincronização concluída: {result.get('items', 0)} anúncios importados"
                    refreshed["sync_finished_at"] = now_label()
                    refreshed["sync_error"] = ""
                    refreshed["operations_refresh_requested"] = True
                write_payload(payload_inner)
                progress(
                    "completed",
                    result.get("items", 0),
                    result.get("items", 0),
                    f"Sincronização concluída: {result.get('items', 0)} anúncios importados. Vendas e reclamações serão atualizadas na próxima rodada automática.",
                )
        except Exception as exc:
            previous = sync_progress_snapshot().get(account_id, {})
            set_sync_progress(
                account_id,
                {
                    **previous,
                    "status": "error",
                    "stage": "error",
                    "message": str(exc),
                    "updated_epoch": time.time(),
                },
                force=True,
            )
            payload_inner = read_payload()
            failed = next(
                (
                    item
                    for item in payload_inner.get("accounts", [])
                    if item.get("id") == account_id or str(item.get("seller_id")) == account_id
                ),
                None,
            )
            if failed:
                failed["sync_status"] = "Sincronização com erro"
                failed["sync_error"] = str(exc)
                failed["sync_finished_at"] = now_label()
            write_payload(payload_inner)
        finally:
            with SYNC_LOCK:
                ACTIVE_SYNC_ACCOUNTS.discard(account_id)

    threading.Thread(target=worker, daemon=True).start()
    return {"queued": True, "status": "queued", "message": "Sincronização adicionada à fila persistente do servidor."}


def resume_pending_official_syncs():
    """Resume queued/running jobs after a deploy or unexpected process restart."""
    time.sleep(max(2, int(os.getenv("MELI_SYNC_RESUME_DELAY_SECONDS", "5"))))
    payload = read_payload(include_catalog=False)
    progress_rows = sync_progress_snapshot()
    pending = []
    for account in payload.get("accounts", []):
        account_id = str(account.get("id") or account.get("seller_id") or "")
        progress = progress_rows.get(account_id) or progress_rows.get(str(account.get("seller_id") or "")) or {}
        stored_status = str(account.get("sync_status") or "").lower()
        if progress.get("status") in {"running", "queued"} or "andamento" in stored_status or "aguardando a fila" in stored_status:
            pending.append((account_id, progress.get("limit"), "retomada após reinício"))
    for account_id, limit, reason in pending:
        try:
            enqueue_official_sync(account_id, limit, reason)
        except Exception as exc:
            set_sync_progress(
                account_id,
                {
                    "status": "error",
                    "stage": "error",
                    "message": f"Não foi possível retomar automaticamente: {exc}",
                    "updated_epoch": time.time(),
                },
                force=True,
            )


def unlink_official_account(payload, account_id):
    account = next(
        (
            item
            for item in payload.get("accounts", [])
            if item.get("id") == account_id or str(item.get("seller_id")) == str(account_id)
        ),
        None,
    )
    if not account:
        raise RuntimeError("Conta não encontrada.")
    if not account.get("official"):
        raise RuntimeError("Apenas contas oficiais podem ser desvinculadas.")

    payload["accounts"] = [item for item in payload.get("accounts", []) if item.get("id") != account.get("id")]
    payload["catalog"] = [
        item
        for item in payload.get("catalog", [])
        if not (item.get("account_id") == account.get("id") or item.get("account") == account.get("nickname") and item.get("official_source"))
    ]
    payload["alerts"] = [
        item
        for item in payload.get("alerts", [])
        if not (item.get("account") == account.get("nickname") and str(item.get("id", "")).startswith(("stock-", "catalog-")))
    ]
    payload["metrics"] = [item for item in payload.get("metrics", []) if item.get("account") != account.get("nickname")]
    return public_account(account)


def cleanup_async_operation_jobs():
    cutoff = time.time() - max(300, int(os.getenv("ASYNC_OPERATION_TTL_SECONDS", "1800")))
    with ASYNC_OPERATION_JOBS_LOCK:
        expired = [
            job_id
            for job_id, job in ASYNC_OPERATION_JOBS.items()
            if float(job.get("updated_epoch") or job.get("created_epoch") or 0) < cutoff
        ]
        for job_id in expired:
            ASYNC_OPERATION_JOBS.pop(job_id, None)


def async_operation_result(job_id):
    cleanup_async_operation_jobs()
    with ASYNC_OPERATION_JOBS_LOCK:
        job = ASYNC_OPERATION_JOBS.get(str(job_id or ""))
        return json.loads(json.dumps(job, ensure_ascii=False)) if job else None


def start_async_operation(kind, work, message="Processamento adicionado à fila."):
    cleanup_async_operation_jobs()
    job_id = f"op-{uuid.uuid4().hex[:12]}"
    now = time.time()
    job = {
        "id": job_id,
        "kind": kind,
        "status": "queued",
        "message": message,
        "created_epoch": now,
        "updated_epoch": now,
    }
    with ASYNC_OPERATION_JOBS_LOCK:
        ASYNC_OPERATION_JOBS[job_id] = job

    def worker():
        with ASYNC_OPERATION_JOBS_LOCK:
            ACTIVE_CLONE_OPERATIONS.add(job_id)
        try:
            with CLONE_WORKER_SEMAPHORE:
                with ASYNC_OPERATION_JOBS_LOCK:
                    current = ASYNC_OPERATION_JOBS.get(job_id)
                    if current:
                        current.update({"status": "running", "message": "Processando dados oficiais no Mercado Livre.", "updated_epoch": time.time()})
                try:
                    result = work()
                    with ASYNC_OPERATION_JOBS_LOCK:
                        current = ASYNC_OPERATION_JOBS.get(job_id)
                        if current:
                            current.update(
                                {
                                    "status": "completed",
                                    "message": "Processamento concluído.",
                                    "result": result,
                                    "updated_epoch": time.time(),
                                }
                            )
                except Exception as exc:
                    with ASYNC_OPERATION_JOBS_LOCK:
                        current = ASYNC_OPERATION_JOBS.get(job_id)
                        if current:
                            current.update(
                                {
                                    "status": "error",
                                    "message": str(exc),
                                    "updated_epoch": time.time(),
                                }
                            )
        finally:
            with ASYNC_OPERATION_JOBS_LOCK:
                ACTIVE_CLONE_OPERATIONS.discard(job_id)

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id, "status": "queued", "message": message, "poll_url": f"/api/async/jobs/{job_id}"}


def item_description_operation(request, actor=None):
    metadata = read_payload(include_catalog=False)
    account_id = str(request.get("account_id") or "")
    account_name = str(request.get("account") or "")
    account = next(
        (
            item
            for item in metadata.get("accounts", [])
            if str(item.get("id") or "") == account_id or item.get("nickname") == account_name
        ),
        None,
    )
    if not account or not account.get("official") or not account.get("access_token"):
        raise RuntimeError("Conta oficial não encontrada para acessar a descrição.")
    item_id = str(request.get("item_id") or "").strip()
    if not item_id:
        raise RuntimeError("Informe o código do anúncio.")
    client = account_client(account)
    if request.get("action") != "update":
        description = client.item_description(item_id, interactive=True) or {}
        return {"ok": True, "description": description.get("plain_text") or description.get("text") or ""}

    text = str(request.get("description") or "").strip()
    if not text:
        raise RuntimeError("A descrição não pode ficar vazia.")
    official = client.update_item_description(item_id, text, interactive=True)
    payload = read_payload()
    item = next(
        (
            row
            for row in payload.get("catalog", [])
            if row.get("id") == item_id and row.get("account_id") == account.get("id")
        ),
        None,
    )
    if item:
        item["description_updated_at"] = now_label()
        append_item_log(
            payload,
            item,
            actor or {},
            "Descrição alterada",
            {"description": {"from": "Descrição anterior", "to": "Descrição atualizada"}},
        )
        write_payload(payload)
    return {"ok": True, "description": text, "official": official}


def official_account_by_name(payload, name):
    return next(
        (
            account
            for account in payload.get("accounts", [])
            if account.get("nickname") == name or account.get("id") == name or str(account.get("seller_id")) == str(name)
        ),
        None,
    )


def clone_picture_payload(item):
    pictures = item.get("pictures") or []
    rows = []
    for picture in pictures:
        url = picture.get("secure_url") or picture.get("url")
        if url:
            rows.append({"source": url})
    return rows


GTIN_IDENTIFIER_ATTRS = {
    "GTIN",
    "EAN",
    "UPC",
    "ISBN",
    "JAN",
    "UNIVERSAL_PRODUCT_CODE",
}
PRODUCT_IDENTIFIER_ATTRS = {*GTIN_IDENTIFIER_ATTRS, "MPN", "PART_NUMBER"}

CLONE_ATTRIBUTE_ALIASES = {
    "PACKAGE_HEIGHT": "SELLER_PACKAGE_HEIGHT",
    "PACKAGE_LENGTH": "SELLER_PACKAGE_LENGTH",
    "PACKAGE_WIDTH": "SELLER_PACKAGE_WIDTH",
    "PACKAGE_WEIGHT": "SELLER_PACKAGE_WEIGHT",
    "PACKAGE_TYPE": "SELLER_PACKAGE_TYPE",
}
CLONE_INTERNAL_READ_ONLY_ATTRIBUTES = {
    "PACKAGE_DATA_SOURCE",
    "SYI_PYMES_ID",
}
CLONE_ALWAYS_ALLOWED_ATTRIBUTES = {
    "SELLER_SKU",
    "SKU",
    "SELLER_PACKAGE_HEIGHT",
    "SELLER_PACKAGE_LENGTH",
    "SELLER_PACKAGE_WIDTH",
    "SELLER_PACKAGE_WEIGHT",
    "SELLER_PACKAGE_TYPE",
}


def clean_attribute_value(text):
    if text in (None, "", [], {}):
        return ""
    value = str(text).strip()
    return "" if value.lower() in {"null", "none", "nan", "undefined"} else value


def clone_attributes_payload(item, source_sku):
    attributes = []
    for attribute in item.get("attributes") or []:
        attr_id = attribute.get("id")
        if not attr_id:
            continue
        row = clone_attribute_row(attribute)
        if row:
            attributes.append(row)
    if source_sku:
        existing = next((attr for attr in attributes if attr.get("id") in {"SELLER_SKU", "SKU"}), None)
        if existing:
            existing.pop("value_id", None)
            existing["value_name"] = source_sku
        else:
            attributes.append({"id": "SELLER_SKU", "value_name": source_sku})
    return attributes


def clone_attribute_row(attribute):
    attr_id = attribute.get("id")
    if not attr_id:
        return {}
    row = {"id": attr_id}
    if attribute.get("value_id") not in (None, ""):
        row["value_id"] = attribute.get("value_id")
    elif clean_attribute_value(attribute.get("value_name")):
        row["value_name"] = clean_attribute_value(attribute.get("value_name"))
    elif isinstance(attribute.get("value_struct"), dict) and attribute.get("value_struct"):
        struct = attribute.get("value_struct")
        if struct.get("number") is not None:
            unit = clean_attribute_value(struct.get("unit"))
            row["value_name"] = f"{struct.get('number')}{f' {unit}' if unit else ''}"
        else:
            row["value_struct"] = struct
    elif attribute.get("values"):
        values = []
        for value in attribute.get("values") or []:
            clean = {}
            if value.get("id") not in (None, ""):
                clean["id"] = value.get("id")
            if clean_attribute_value(value.get("name")):
                clean["name"] = clean_attribute_value(value.get("name"))
            if isinstance(value.get("struct"), dict) and value.get("struct"):
                clean["struct"] = value.get("struct")
            if clean:
                values.append(clean)
        if values:
            row["values"] = values
    return row if len(row) > 1 else {}


def clone_source_item(client, item_id):
    method = getattr(client, "item_for_clone", None)
    if callable(method):
        try:
            return method(item_id)
        except Exception:
            pass
    return client.item(item_id)


def call_interactive_client_method(method, *args):
    try:
        return method(*args, interactive=True)
    except TypeError as exc:
        if "interactive" not in str(exc):
            raise
        return method(*args)


def clone_source_user_product(client, source_item):
    user_product_id = (source_item or {}).get("user_product_id")
    method = getattr(client, "user_product", None)
    if not user_product_id or not callable(method):
        return {}
    try:
        user_product = call_interactive_client_method(method, user_product_id)
        return user_product if isinstance(user_product, dict) else {}
    except Exception:
        return {}


def enrich_clone_source_item(client, source_item):
    """Merge the current User Product sheet into the legacy item representation."""
    source_item = json.loads(json.dumps(source_item or {}, ensure_ascii=False))
    user_product = clone_source_user_product(client, source_item)
    if not user_product:
        return source_item

    attributes = {}
    # The User Product sheet is the current source of truth and wins over legacy item values.
    for container in (source_item, user_product):
        for attribute in container.get("attributes") or []:
            attr_id = str(attribute.get("id") or "").upper()
            if attr_id and clone_attribute_row(attribute):
                attributes[attr_id] = attribute
    if attributes:
        source_item["attributes"] = list(attributes.values())

    user_variations = user_product.get("variations") or []
    if user_variations:
        source_variations = source_item.get("variations") or []
        if not source_variations:
            source_item["variations"] = json.loads(json.dumps(user_variations, ensure_ascii=False))
        else:
            user_by_id = {str(row.get("id")): row for row in user_variations if row.get("id") not in (None, "")}
            for index, variation in enumerate(source_variations):
                user_variation = user_by_id.get(str(variation.get("id")))
                if not user_variation and index < len(user_variations):
                    user_variation = user_variations[index]
                if not user_variation:
                    continue
                for section_name in ("attributes", "attribute_combinations"):
                    merged = {}
                    for container in (variation, user_variation):
                        for attribute in container.get(section_name) or []:
                            attr_id = str(attribute.get("id") or "").upper()
                            if attr_id and clone_attribute_row(attribute):
                                merged[attr_id] = attribute
                    if merged:
                        variation[section_name] = list(merged.values())

    for source_key, target_key in (
        ("name", "title"),
        ("family_name", "family_name"),
        ("domain_id", "domain_id"),
        ("catalog_product_id", "catalog_product_id"),
        ("pictures", "pictures"),
    ):
        if not source_item.get(target_key) and user_product.get(source_key):
            source_item[target_key] = user_product.get(source_key)
    source_item["_user_product_hydrated"] = True
    return source_item


def clone_source_catalog_product(client, source_item):
    product_id = source_item.get("catalog_product_id")
    if not product_id:
        return {}
    try:
        product = call_interactive_client_method(client.product, product_id)
        return product if isinstance(product, dict) else {}
    except Exception:
        return {}


def merge_clone_source_local_snapshot(source_item, payload, account, item_id):
    """Complete seller-owned fields that can be absent from the public item representation."""
    source_item = json.loads(json.dumps(source_item or {}, ensure_ascii=False))
    if not source_item.get("seller_id") and account.get("seller_id"):
        source_item["seller_id"] = account.get("seller_id")
    local_item = next(
        (
            row
            for row in payload.get("catalog", []) or []
            if str(row.get("id") or "") == str(item_id or "")
            and (
                row.get("account_id") == account.get("id")
                or row.get("account") == account.get("nickname")
            )
        ),
        None,
    )
    if not local_item:
        return source_item
    package_map = {
        "package_height": "SELLER_PACKAGE_HEIGHT",
        "package_width": "SELLER_PACKAGE_WIDTH",
        "package_length": "SELLER_PACKAGE_LENGTH",
        "package_weight": "SELLER_PACKAGE_WEIGHT",
    }
    attributes = source_item.setdefault("attributes", [])
    existing_ids = {canonical_clone_attribute_id(row.get("id")) for row in attributes if row.get("id")}
    for field, attr_id in package_map.items():
        value = clean_attribute_value(local_item.get(field))
        if not value or attr_id in existing_ids:
            continue
        try:
            value = seller_package_api_value(field, value)
        except Exception:
            continue
        attributes.append({"id": attr_id, "value_name": value})
        existing_ids.add(attr_id)
    local_gtin = clean_attribute_value(local_item.get("gtin"))
    if local_gtin and not source_clone_identifiers(source_item):
        attributes.append({"id": "GTIN", "value_name": local_gtin})
    return source_item


def clone_source_bundle(payload, account_identifier, item_id, include_description=False, force=False):
    account = official_account_by_name(payload, account_identifier)
    if not account or not account.get("official") or not account.get("access_token"):
        raise RuntimeError("Conta origem oficial não encontrada para ler o anúncio.")
    item_id = str(item_id or "").strip()
    cache_key = f"{account.get('id') or account.get('seller_id')}:{item_id}"
    ttl = max(30, int(os.getenv("MELI_CLONE_SOURCE_CACHE_SECONDS", "300")))
    now = time.monotonic()
    with CLONE_SOURCE_CACHE_LOCK:
        cached = CLONE_SOURCE_CACHE.get(cache_key)
        cache_valid = cached and not force and now - cached.get("time", 0) < ttl
        if cache_valid and (not include_description or cached.get("description_loaded")):
            bundle = json.loads(json.dumps(cached["bundle"], ensure_ascii=False))
            bundle["source_item"] = merge_clone_source_local_snapshot(bundle.get("source_item"), payload, account, item_id)
            return bundle

    if cache_valid and include_description:
        client = account_client(account)
        try:
            description = call_interactive_client_method(client.item_description, item_id) or {}
            description_text = description.get("plain_text") or description.get("text") or ""
        except Exception:
            description_text = ""
        with CLONE_SOURCE_CACHE_LOCK:
            cached["bundle"]["description"] = description_text
            cached["description_loaded"] = True
            cached["time"] = now
            bundle = json.loads(json.dumps(cached["bundle"], ensure_ascii=False))
            bundle["source_item"] = merge_clone_source_local_snapshot(bundle.get("source_item"), payload, account, item_id)
            return bundle

    client = account_client(account)
    raw_item = clone_source_item(client, item_id)
    with ThreadPoolExecutor(max_workers=3, thread_name_prefix="clone-source") as executor:
        enriched_future = executor.submit(enrich_clone_source_item, client, raw_item)
        catalog_future = executor.submit(clone_source_catalog_product, client, raw_item)
        description_future = executor.submit(call_interactive_client_method, client.item_description, item_id) if include_description else None
        source_item = enriched_future.result()
        catalog_product = catalog_future.result()
        description_text = ""
        if description_future:
            try:
                description = description_future.result() or {}
                description_text = description.get("plain_text") or description.get("text") or ""
            except Exception:
                description_text = ""
    source_item = merge_clone_source_local_snapshot(source_item, payload, account, item_id)
    if catalog_product:
        source_item["_clone_catalog_product"] = catalog_product
    bundle = {
        "source_item": source_item,
        "catalog_product": catalog_product,
        "description": description_text,
    }
    with CLONE_SOURCE_CACHE_LOCK:
        CLONE_SOURCE_CACHE[cache_key] = {
            "time": now,
            "description_loaded": bool(include_description),
            "bundle": json.loads(json.dumps(bundle, ensure_ascii=False)),
        }
        if len(CLONE_SOURCE_CACHE) > 500:
            oldest = sorted(CLONE_SOURCE_CACHE, key=lambda key: CLONE_SOURCE_CACHE[key].get("time", 0))[:100]
            for key in oldest:
                CLONE_SOURCE_CACHE.pop(key, None)
    return bundle


def source_clone_attribute(source_item, catalog_product, attr_id):
    wanted = str(attr_id or "").upper()
    wanted_ids = {wanted}
    wanted_ids.update(key for key, value in CLONE_ATTRIBUTE_ALIASES.items() if value == wanted)
    if wanted == "GTIN":
        wanted_ids.update({"EAN", "UPC", "JAN", "ISBN", "UNIVERSAL_PRODUCT_CODE"})
    embedded_catalog = source_item.get("_clone_catalog_product") or {}
    for container in (source_item, catalog_product or {}, embedded_catalog):
        for attribute in container.get("attributes") or []:
            if str(attribute.get("id") or "").upper() in wanted_ids:
                row = clone_attribute_row(attribute)
                if row:
                    row["id"] = wanted
                    return row

        if wanted == "GTIN":
            for key in ("gtin", "ean", "upc", "isbn", "jan"):
                value = clean_attribute_value(container.get(key))
                if value:
                    return {"id": "GTIN", "value_name": value}

    variations = source_item.get("variations") or []
    variation_rows = []
    for variation in variations:
        row = {}
        for section in (variation.get("attributes") or [], variation.get("attribute_combinations") or []):
            attribute = next((item for item in section if str(item.get("id") or "").upper() in wanted_ids), None)
            if attribute:
                row = clone_attribute_row(attribute)
                if row:
                    row["id"] = wanted
                break
        if not row:
            return {}
        variation_rows.append(row)
    if variation_rows:
        signatures = {json.dumps(row, sort_keys=True, ensure_ascii=False) for row in variation_rows}
        if len(signatures) == 1:
            return variation_rows[0]
    return {}


def clone_attribute_display_value(attribute):
    if not attribute:
        return ""
    if clean_attribute_value(attribute.get("value_name")):
        return clean_attribute_value(attribute.get("value_name"))
    values = []
    for value in attribute.get("values") or []:
        name = clean_attribute_value(value.get("name"))
        if name and name not in values:
            values.append(name)
    return ", ".join(values)


def source_clone_identifiers(source_item, catalog_product=None):
    wanted = GTIN_IDENTIFIER_ATTRS
    values = []
    containers = [source_item, catalog_product or {}, source_item.get("_clone_catalog_product") or {}]
    for container in containers:
        sections = [container.get("attributes") or []]
        for variation in container.get("variations") or []:
            sections.append(variation.get("attributes") or [])
            sections.append(variation.get("attribute_combinations") or [])
        for section in sections:
            for attribute in section:
                if str(attribute.get("id") or "").upper() not in wanted:
                    continue
                row = clone_attribute_row(attribute)
                value = clone_attribute_display_value(row)
                if value and value not in values:
                    values.append(value)
        for key in ("gtin", "ean", "upc", "isbn", "jan"):
            value = clean_attribute_value(container.get(key))
            if value and value not in values:
                values.append(value)
    return values


def normalize_clone_identifier_codes(value):
    raw = clean_attribute_value(value)
    if not raw:
        return []
    codes = []
    for part in re.split(r"[,;\n]+", raw):
        code = re.sub(r"[\s.\-]+", "", part or "")
        if not code:
            continue
        if not code.isdigit() or len(code) not in {8, 10, 12, 13, 14}:
            raise RuntimeError(
                f'O código "{part.strip()}" não é um GTIN/EAN/UPC válido. Use apenas 8, 10, 12, 13 ou 14 dígitos.'
            )
        if code not in codes:
            codes.append(code)
    return codes


def item_gtin_update_fragment(client, item_id, raw_codes):
    codes = normalize_clone_identifier_codes(raw_codes)
    if not codes:
        raise RuntimeError("Informe um EAN, UPC ou GTIN válido para atualizar o anúncio.")
    source_item = enrich_clone_source_item(client, clone_source_item(client, item_id))
    variations = source_item.get("variations") or []
    if variations:
        if len(codes) != len(variations):
            raise RuntimeError(
                f"Este anúncio possui {len(variations)} variações. Informe exatamente {len(variations)} códigos, um por variação, separados por vírgula."
            )
        rows = []
        for variation, code in zip(variations, codes):
            variation_id = variation.get("id")
            if variation_id in (None, ""):
                raise RuntimeError("O Mercado Livre não retornou o ID de uma das variações; sincronize o anúncio e tente novamente.")
            attributes = [
                row
                for attribute in variation.get("attributes") or []
                if str(attribute.get("id") or "").upper() not in GTIN_IDENTIFIER_ATTRS
                if (row := clone_attribute_row(attribute))
            ]
            attributes.append({"id": "GTIN", "value_name": code})
            rows.append({"id": variation_id, "attributes": attributes})
        return {"variations": rows}, codes
    if len(codes) != 1:
        raise RuntimeError("Este anúncio não possui variações. Informe somente um código EAN, UPC ou GTIN.")
    return {"attributes": [{"id": "GTIN", "value_name": codes[0]}]}, codes


def verify_gtin_update(client, item_id, expected_codes, attempts=4):
    last_item = {}
    for attempt in range(max(1, attempts)):
        if attempt:
            time.sleep(min(2.0, 0.35 * attempt))
        last_item = enrich_clone_source_item(client, clone_source_item(client, item_id))
        current_codes = normalize_clone_identifier_codes(",".join(source_clone_identifiers(last_item, {})))
        if current_codes == expected_codes:
            return last_item
    raise RuntimeError(
        "O Mercado Livre recebeu a atualização, mas não confirmou o EAN/UPC/GTIN na leitura de verificação. Nenhuma confirmação local foi gravada."
    )


def remove_clone_product_identifiers(attributes):
    return [
        attribute
        for attribute in attributes or []
        if str(attribute.get("id") or "").upper() not in GTIN_IDENTIFIER_ATTRS
    ]


def clone_variations_payload(item, sku_override="", price_override="", stock_override=""):
    source_variations = item.get("variations") or []
    if not source_variations:
        return []
    rows = []
    single_variation = len(source_variations) == 1
    for variation in source_variations:
        row = {}
        combinations = [
            clean
            for attribute in variation.get("attribute_combinations") or []
            if (clean := clone_attribute_row(attribute))
        ]
        attributes = [
            clean
            for attribute in variation.get("attributes") or []
            if (clean := clone_attribute_row(attribute))
        ]
        if single_variation and sku_override:
            attributes = [
                attribute for attribute in attributes if str(attribute.get("id") or "").upper() not in {"SELLER_SKU", "SKU"}
            ]
            attributes.append({"id": "SELLER_SKU", "value_name": sku_override})
        if combinations:
            row["attribute_combinations"] = combinations
        if attributes:
            row["attributes"] = attributes
        price = price_override or variation.get("price") or item.get("price")
        if price not in (None, ""):
            row["price"] = float(price)
        quantity = stock_override if single_variation and stock_override not in (None, "") else variation.get("available_quantity")
        if quantity in (None, ""):
            quantity = variation.get("available_quantity_by_channel") or 0
        try:
            row["available_quantity"] = max(0, int(float(quantity or 0)))
        except (TypeError, ValueError):
            row["available_quantity"] = 0
        picture_ids = [picture_id for picture_id in variation.get("picture_ids") or [] if picture_id]
        if picture_ids:
            row["picture_ids"] = picture_ids
        sale_terms = clone_sale_terms_payload(variation)
        if sale_terms:
            row["sale_terms"] = sale_terms
        rows.append(row)
    return rows


def apply_clone_gtin_override(create_payload, raw_value):
    codes = normalize_clone_identifier_codes(raw_value)
    if not codes:
        return create_payload
    remove_clone_attributes(create_payload, ["EMPTY_GTIN_REASON"])
    create_payload["attributes"] = remove_clone_product_identifiers(create_payload.get("attributes") or [])
    variations = create_payload.get("variations") or []
    if variations:
        if len(codes) != len(variations):
            raise RuntimeError(
                f"Este anúncio possui {len(variations)} variações. Informe {len(variations)} códigos GTIN/EAN/UPC separados por vírgula, um para cada variação."
            )
        for variation, code in zip(variations, codes):
            variation["attributes"] = remove_clone_product_identifiers(variation.get("attributes") or [])
            variation.setdefault("attributes", []).append({"id": "GTIN", "value_name": code})
    else:
        create_payload.setdefault("attributes", []).append({"id": "GTIN", "value_name": ",".join(codes)})
    return create_payload


def clone_required_attribute_satisfied(create_payload, attr_id):
    attr_id = canonical_clone_attribute_id(attr_id)
    if not attr_id:
        return True
    if clone_payload_has_attribute(create_payload, attr_id):
        return True
    if attr_id == "GTIN" and clone_payload_has_attribute(create_payload, "EMPTY_GTIN_REASON"):
        return True
    if attr_id == "EMPTY_GTIN_REASON" and clone_payload_has_attribute(create_payload, "GTIN"):
        return True
    return False


def hydrate_required_clone_attributes(create_payload, source_item, category_attributes, catalog_product=None):
    copied = []
    for definition in category_attributes or []:
        attr_id = definition.get("id")
        if (
            not attr_id
            or not clone_attribute_is_required(definition)
            or not clone_attribute_user_editable(definition, attr_id)
            or clone_required_attribute_satisfied(create_payload, attr_id)
        ):
            continue
        row = source_clone_attribute(source_item, catalog_product or {}, attr_id)
        if not row:
            continue
        create_payload.setdefault("attributes", []).append(row)
        copied.append(str(attr_id).upper())
    return copied


def hydrate_clone_package_attributes(create_payload, source_item):
    copied = []
    field_map = {
        "package_height": "SELLER_PACKAGE_HEIGHT",
        "package_length": "SELLER_PACKAGE_LENGTH",
        "package_width": "SELLER_PACKAGE_WIDTH",
        "package_weight": "SELLER_PACKAGE_WEIGHT",
    }
    values = package_values_from_item(source_item)
    for field, attr_id in field_map.items():
        if clone_payload_has_attribute(create_payload, attr_id) or not values.get(field):
            continue
        try:
            value = seller_package_api_value(field, values[field])
        except Exception:
            continue
        create_payload.setdefault("attributes", []).append({"id": attr_id, "value_name": value})
        copied.append(attr_id)
    if not clone_payload_has_attribute(create_payload, "SELLER_PACKAGE_TYPE"):
        package_type = source_clone_attribute(source_item, {}, "SELLER_PACKAGE_TYPE")
        if package_type:
            create_payload.setdefault("attributes", []).append(package_type)
            copied.append("SELLER_PACKAGE_TYPE")
    return copied


def restore_clone_attribute_from_source(create_payload, source_item, category_attributes, attr_id):
    canonical_id = canonical_clone_attribute_id(attr_id)
    if not canonical_id or clone_required_attribute_satisfied(create_payload, canonical_id):
        return bool(canonical_id)
    if canonical_id.startswith("SELLER_PACKAGE_"):
        hydrate_clone_package_attributes(create_payload, source_item)
        if clone_required_attribute_satisfied(create_payload, canonical_id):
            return True
    row = source_clone_attribute(source_item, {}, canonical_id)
    if not row:
        return False
    definition = category_attribute_definition(category_attributes, canonical_id)
    if definition and not clone_attribute_user_editable(definition, canonical_id):
        return False
    create_payload.setdefault("attributes", []).append(row)
    return True


def clone_sale_terms_payload(item):
    non_modifiable = {
        term.strip().upper()
        for term in os.getenv("MELI_CLONE_EXCLUDED_SALE_TERMS", "SUBSCRIBABLE").split(",")
        if term.strip()
    }
    terms = []
    for term in item.get("sale_terms") or []:
        term_id = term.get("id")
        if not term_id or str(term_id).upper() in non_modifiable:
            continue
        row = {"id": term_id}
        if term.get("value_id"):
            row["value_id"] = term.get("value_id")
        elif term.get("value_name"):
            row["value_name"] = term.get("value_name")
        if len(row) > 1:
            terms.append(row)
    return terms


def remove_clone_sale_terms(create_payload, term_ids):
    wanted = {str(term_id or "").upper() for term_id in term_ids if term_id}
    if not wanted:
        return []
    removed = []
    for section in [create_payload, *(create_payload.get("variations") or [])]:
        kept = []
        for term in section.get("sale_terms") or []:
            term_id = str(term.get("id") or "").upper()
            if term_id in wanted:
                removed.append(term_id)
            else:
                kept.append(term)
        if kept:
            section["sale_terms"] = kept
        else:
            section.pop("sale_terms", None)
    return list(dict.fromkeys(removed))


def clone_shipping_payload(item):
    shipping = item.get("shipping") or {}
    allowed_keys = ("local_pick_up", "free_shipping", "store_pick_up")
    payload = {key: shipping.get(key) for key in allowed_keys if key in shipping and shipping.get(key) is not None}
    return payload


def clone_extra_required_fields(item):
    fields = {}
    body_fields = (
        "family_name",
        "domain_id",
    )
    for field in body_fields:
        if item.get(field) not in (None, "", [], {}):
            fields[field] = item.get(field)
    attribute_to_body = {
        "FAMILY_NAME": "family_name",
        "MODEL": "family_name",
    }
    for attribute in item.get("attributes") or []:
        attr_id = str(attribute.get("id") or "").upper()
        field = attribute_to_body.get(attr_id)
        if field and not fields.get(field):
            value = attribute.get("value_name")
            if not value and attribute.get("values"):
                value = (attribute.get("values") or [{}])[0].get("name")
            if value:
                fields[field] = value
    if fields.get("family_name"):
        fields["family_name"] = normalize_family_name(fields["family_name"])
    return fields


def build_clone_item_payload(source_item, edits):
    source_sku = source_item.get("seller_custom_field") or item_sku(source_item) or source_item.get("id") or "SEM-SKU"
    new_sku = clean_attribute_value(edits.get("sku") or edits.get("sku_suffix")) or source_sku
    requested_listing_type = edits.get("listing_type_id") or ""
    listing_type_id = requested_listing_type if requested_listing_type in {"gold_special", "gold_pro"} else source_item.get("listing_type_id") or "gold_special"
    source_quantity = item_available_quantity(source_item)
    if source_quantity in (None, ""):
        source_quantity = 1
    payload = {
        "title": (edits.get("title") or source_item.get("title") or "").strip()[:60],
        "category_id": source_item.get("category_id"),
        "price": float(edits.get("price") or source_item.get("price") or 0),
        "currency_id": source_item.get("currency_id") or "BRL",
        "available_quantity": max(
            0,
            int(float(edits.get("stock") if edits.get("stock") not in (None, "") else source_quantity)),
        ),
        "buying_mode": source_item.get("buying_mode") or "buy_it_now",
        "listing_type_id": listing_type_id,
        "condition": source_item.get("condition") or "new",
        "pictures": clone_picture_payload(source_item),
        "attributes": clone_attributes_payload(source_item, new_sku),
        "sale_terms": clone_sale_terms_payload(source_item),
    }
    variations = clone_variations_payload(
        source_item,
        new_sku,
        edits.get("price"),
        edits.get("stock"),
    )
    if variations:
        payload["variations"] = variations
        payload.pop("available_quantity", None)
    payload.update(clone_extra_required_fields(source_item))
    if source_item.get("catalog_product_id"):
        payload["catalog_product_id"] = source_item.get("catalog_product_id")
    if is_catalog_listing(source_item):
        payload["catalog_listing"] = True
    shipping = clone_shipping_payload(source_item)
    if shipping:
        payload["shipping"] = shipping
    if source_item.get("seller_custom_field"):
        payload["seller_custom_field"] = new_sku
    if clean_attribute_value(edits.get("gtin")):
        apply_clone_gtin_override(payload, edits.get("gtin"))
    return {key: value for key, value in payload.items() if value not in (None, "", [], {})}


def apply_target_account_clone_rules(create_payload, source_item, source_account, target_account):
    source_seller_id = str(first_present(source_item, ["seller_id", "seller.id"], "") or source_account.get("seller_id") or "")
    target_seller_id = str(target_account.get("seller_id") or "")
    same_seller = bool(source_seller_id and target_seller_id and source_seller_id == target_seller_id)
    if same_seller and source_item.get("official_store_id") not in (None, "", [], {}):
        create_payload["official_store_id"] = source_item.get("official_store_id")
    else:
        create_payload.pop("official_store_id", None)
    return create_payload


OFFICIAL_STORE_CLONE_KEYS = {
    "official_store_id",
    "official_store_ids",
    "official_store",
    "official_store_name",
}


def clone_row_is_official_store_reference(value):
    if not isinstance(value, dict):
        return False
    reference_id = str(value.get("id") or value.get("code") or "").strip().lower()
    return reference_id in OFFICIAL_STORE_CLONE_KEYS


def strip_official_store_clone_fields(value):
    removed = False
    if isinstance(value, dict):
        for key in list(value):
            if str(key).strip().lower() in OFFICIAL_STORE_CLONE_KEYS:
                value.pop(key, None)
                removed = True
                continue
            removed = strip_official_store_clone_fields(value.get(key)) or removed
    elif isinstance(value, list):
        for row in list(value):
            if clone_row_is_official_store_reference(row):
                value.remove(row)
                removed = True
                continue
            removed = strip_official_store_clone_fields(row) or removed
    return removed


def normalized_store_name(value):
    return re.sub(
        r"[^a-z0-9]+",
        "",
        unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower(),
    )


def target_official_store_id(target_client, target_account, source_item):
    seller_id = str(target_account.get("seller_id") or "")
    if not seller_id:
        return None
    ttl = max(300, int(os.getenv("MELI_OFFICIAL_STORE_CACHE_SECONDS", "21600")))
    now = time.time()
    with OFFICIAL_STORE_CACHE_LOCK:
        cached = OFFICIAL_STORE_CACHE.get(seller_id)
    if cached and now - cached.get("time", 0) < ttl:
        brands = cached.get("brands") or []
    else:
        try:
            response = target_client.user_brands(seller_id, interactive=True)
            brands = response.get("brands") or [] if isinstance(response, dict) else []
        except Exception:
            brands = []
        with OFFICIAL_STORE_CACHE_LOCK:
            OFFICIAL_STORE_CACHE[seller_id] = {"time": now, "brands": brands}
    available = [
        brand for brand in brands
        if brand.get("official_store_id") not in (None, "")
        and str(brand.get("status") or "active").lower() not in {"inactive", "offline", "disabled"}
    ]
    if not available:
        return None
    source_names = {
        normalized_store_name(source_item.get("official_store_name")),
        normalized_store_name(first_present(source_item, ["official_store.name", "official_store.fantasy_name"], "")),
    }
    for attribute in source_item.get("attributes") or []:
        if str(attribute.get("id") or "").upper() == "BRAND":
            source_names.add(normalized_store_name(clone_attribute_display_value(attribute)))
    source_names.discard("")
    for brand in available:
        target_names = {
            normalized_store_name(brand.get("name")),
            normalized_store_name(brand.get("fantasy_name")),
            normalized_store_name(brand.get("normalized_name")),
            normalized_store_name(first_present(brand, ["brand_registry.brand_name"], "")),
        }
        if source_names.intersection(target_names):
            return brand.get("official_store_id")
    if len(available) == 1:
        return available[0].get("official_store_id")
    return None


def prepare_cross_account_official_store_payload(payload, destination_store_id=None):
    removed = strip_official_store_clone_fields(payload)
    payload["official_store_id"] = destination_store_id
    return removed


def required_fields_from_error(exc):
    text = str(exc)
    match = re.search(r"properties \[([^\]]+)\]", text)
    if not match:
        return []
    return [field.strip().strip("'\"") for field in match.group(1).split(",") if field.strip()]


def meli_error_detail(exc):
    text = str(exc)
    raw_json = ""
    if ": " in text:
        raw_json = text.split(": ", 1)[1]
    else:
        raw_json = text
    try:
        return json.loads(raw_json)
    except Exception:
        return {}


def meli_error_causes(exc):
    detail = meli_error_detail(exc)
    causes = detail.get("cause") if isinstance(detail, dict) else []
    return causes if isinstance(causes, list) else []


def meli_error_text(exc):
    detail = meli_error_detail(exc)
    parts = [str(exc)]
    if isinstance(detail, dict):
        parts.extend(str(detail.get(key) or "") for key in ("message", "error"))
        for cause in detail.get("cause") or []:
            parts.append(str(cause.get("message") or ""))
            parts.append(str(cause.get("code") or ""))
    return " ".join(part for part in parts if part)


def meli_error_status(exc):
    detail = meli_error_detail(exc)
    if isinstance(detail, dict):
        try:
            return int(detail.get("status"))
        except (TypeError, ValueError):
            pass
    match = re.search(r"HTTP\s+(\d{3})", str(exc), flags=re.I)
    return int(match.group(1)) if match else 0


def meli_rate_limited_error(exc):
    text = meli_error_text(exc).lower()
    return meli_error_status(exc) == 429 or "local_rate_limited" in text or "rate limit" in text


def clone_rate_limit_delay(attempt):
    base = max(0.1, float(os.getenv("MELI_CLONE_RATE_LIMIT_BASE_SECONDS", "1.5")))
    maximum = max(base, float(os.getenv("MELI_CLONE_RATE_LIMIT_MAX_SECONDS", "8")))
    return min(maximum, base * (2**attempt)) + random.uniform(0.05, 0.25)


def invalid_fields_from_error(exc):
    text = str(exc)
    fields = []
    match = re.search(r"fields \[([^\]]+)\] are invalid", text, flags=re.IGNORECASE)
    if match:
        fields.extend(field.strip().strip("'\"") for field in match.group(1).split(",") if field.strip())
    detail = meli_error_detail(exc)
    if isinstance(detail, dict):
        for cause in detail.get("cause") or []:
            code = str(cause.get("code") or cause.get("cause_id") or "")
            if "invalid_fields" not in code:
                continue
            for reference in cause.get("references") or []:
                field = str(reference or "").replace("body.", "").strip()
                if field and field != "body":
                    fields.append(field)
        error_text = " ".join(str(detail.get(key) or "") for key in ("message", "error"))
        match = re.search(r"fields \[([^\]]+)\] are invalid", error_text, flags=re.IGNORECASE)
        if match:
            fields.extend(field.strip().strip("'\"") for field in match.group(1).split(",") if field.strip())
    return list(dict.fromkeys(fields))


def attribute_ids_from_error_text(text):
    ids = []
    for pattern in (
        r"Attribute \[([A-Z0-9_]+)\]",
        r"attribute \[([A-Z0-9_]+)\]",
        r"attribute ([A-Z0-9_]+)",
        r"atributo \[([A-Z0-9_]+)\]",
    ):
        ids.extend(match.group(1).upper() for match in re.finditer(pattern, text or "", flags=re.I))
    return list(dict.fromkeys(ids))


def human_attribute_names_from_error_text(text):
    names = []
    for pattern in (
        r'em "([^"]+)"',
        r"em '([^']+)'",
        r'Attribute ([A-Z0-9_]+) with value null was omitted',
    ):
        for match in re.finditer(pattern, text or "", flags=re.I):
            value = match.group(1).strip()
            if value:
                names.append(value)
    return list(dict.fromkeys(names))


ATTRIBUTE_NAME_TO_ID = {
    "quantidade de pares": "UNITS_PER_PACKAGE",
    "comprimento do cabo": "CABLE_LENGTH",
}


def attribute_id_from_human_name(name):
    normalized = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii").lower().strip()
    return ATTRIBUTE_NAME_TO_ID.get(normalized) or re.sub(r"[^A-Z0-9]+", "_", (name or "").upper()).strip("_")


def remove_clone_attributes(create_payload, attr_ids):
    wanted = {str(attr_id).upper() for attr_id in attr_ids if attr_id}
    if not wanted:
        return []
    removed = []
    kept = []
    for attribute in create_payload.get("attributes") or []:
        attr_id = str(attribute.get("id") or "").upper()
        if attr_id in wanted:
            removed.append(attr_id)
        else:
            kept.append(attribute)
    create_payload["attributes"] = kept
    return list(dict.fromkeys(removed))


def add_or_update_clone_attribute(create_payload, attr_id, value):
    value = clean_attribute_value(value)
    if not attr_id or not value:
        return False
    attributes = create_payload.setdefault("attributes", [])
    existing = next((attr for attr in attributes if str(attr.get("id") or "").upper() == str(attr_id).upper()), None)
    if existing:
        existing.pop("value_id", None)
        existing.pop("values", None)
        existing.pop("value_struct", None)
        existing["value_name"] = value
    else:
        attributes.append({"id": attr_id, "value_name": value})
    return True


ATTRIBUTE_LABELS_PT = {
    "ALPHANUMERIC_MODEL": "Modelo alfanumérico",
    "DETAILED_MODEL": "Modelo detalhado",
    "MODEL": "Modelo",
    "LINE": "Linha",
    "BRAND": "Marca",
    "FAMILY_NAME": "Nome da família do produto",
    "UNITS_PER_PACKAGE": "Quantidade de unidades por embalagem",
    "CABLE_LENGTH": "Comprimento do cabo",
    "SELLER_PACKAGE_WEIGHT": "Peso da embalagem",
    "SELLER_PACKAGE_HEIGHT": "Altura da embalagem",
    "SELLER_PACKAGE_WIDTH": "Largura da embalagem",
    "SELLER_PACKAGE_LENGTH": "Comprimento da embalagem",
    "SELLER_PACKAGE_TYPE": "Tipo de embalagem",
    "PACKAGE_HEIGHT": "Altura da embalagem",
    "PACKAGE_WIDTH": "Largura da embalagem",
    "PACKAGE_LENGTH": "Comprimento da embalagem",
    "PACKAGE_WEIGHT": "Peso da embalagem",
    "PACKAGE_DATA_SOURCE": "Origem dos dados da embalagem",
    "GAUGES": "Calibres",
    "RECOMMENDED_INSTRUMENT": "Instrumento recomendado",
    "OUTPUT_CONNECTORS": "Conectores de saída",
    "GTIN": "Código universal do produto (GTIN/EAN)",
    "EAN": "Código EAN",
    "UPC": "Código UPC",
    "EMPTY_GTIN_REASON": "Motivo para não informar o código universal",
}


def clone_attribute_label(field_id, fallback=""):
    normalized = str(field_id or "").replace("attribute:", "").upper()
    return ATTRIBUTE_LABELS_PT.get(normalized) or fallback or normalized.replace("_", " ").title()


def clone_pending_field(field_id, label, kind="text", message="", item_id="", options=None, **metadata):
    field = {
        "id": field_id,
        "label": clone_attribute_label(field_id, label),
        "kind": kind,
        "message": message,
        "item_id": item_id,
        "options": options or [],
    }
    field.update({key: value for key, value in metadata.items() if value not in (None, "", [], {})})
    return field


def dedupe_pending_fields(fields):
    deduped = {}
    for field in fields:
        field_id = field.get("id")
        if not field_id:
            continue
        if field_id not in deduped:
            deduped[field_id] = field
    return list(deduped.values())


def normalize_family_name(value):
    text = clean_attribute_value(value)
    return text[:60] if len(text) > 60 else text


def source_attribute_value(source_item, ids):
    wanted = {item.upper() for item in ids}
    for attribute in source_item.get("attributes") or []:
        if str(attribute.get("id") or "").upper() not in wanted:
            continue
        value = attribute.get("value_name")
        if not value and attribute.get("values"):
            value = (attribute.get("values") or [{}])[0].get("name")
        if value:
            return value
    return ""


def normalized_attribute_label(value):
    return re.sub(
        r"[^a-z0-9]+",
        " ",
        unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower(),
    ).strip()


def cached_category_attributes(client, category_id):
    category_id = str(category_id or "").strip()
    if not category_id:
        return []
    ttl = max(300, int(os.getenv("CATEGORY_ATTRIBUTES_CACHE_SECONDS", "3600")))
    now = time.monotonic()
    with CATEGORY_ATTRIBUTES_LOCK:
        cached = CATEGORY_ATTRIBUTES_CACHE.get(category_id)
        if cached and now - cached["time"] < ttl:
            return cached["attributes"]
    attributes = client.category_attributes(category_id)
    attributes = attributes if isinstance(attributes, list) else []
    with CATEGORY_ATTRIBUTES_LOCK:
        CATEGORY_ATTRIBUTES_CACHE[category_id] = {"time": now, "attributes": attributes}
    return attributes


def clone_attribute_is_required(attribute):
    tags = attribute.get("tags") or {}
    if not isinstance(tags, dict):
        return False
    for key, value in tags.items():
        normalized = str(key or "").lower()
        if "required" not in normalized or normalized.startswith(("not_", "non_", "optional_")):
            continue
        if value not in (False, None, "", 0, "0", "false", "False"):
            return True
    return False


def clone_attribute_units(attribute):
    units = []
    default_unit = attribute.get("default_unit")
    if isinstance(default_unit, dict):
        default_unit = default_unit.get("name") or default_unit.get("id")
    if clean_attribute_value(default_unit):
        units.append(clean_attribute_value(default_unit))
    for unit in attribute.get("allowed_units") or []:
        if isinstance(unit, dict):
            unit = unit.get("name") or unit.get("id")
        unit = clean_attribute_value(unit)
        if unit and unit not in units:
            units.append(unit)
    return units


def category_attribute_definition(category_attributes, attr_id="", label=""):
    wanted_id = str(attr_id or "").replace("attribute:", "").upper()
    wanted_label = normalized_attribute_label(label)
    for attribute in category_attributes or []:
        current_id = str(attribute.get("id") or "").upper()
        current_label = normalized_attribute_label(attribute.get("name") or "")
        if (wanted_id and current_id == wanted_id) or (wanted_label and current_label == wanted_label):
            return attribute
    return {}


def category_attribute_ids(category_attributes):
    return {
        str(attribute.get("id") or "").upper()
        for attribute in category_attributes or []
        if attribute.get("id")
    }


def canonical_clone_attribute_id(attr_id):
    normalized = str(attr_id or "").replace("attribute:", "").upper()
    if normalized in CLONE_INTERNAL_READ_ONLY_ATTRIBUTES:
        return ""
    return CLONE_ATTRIBUTE_ALIASES.get(normalized, normalized)


def clone_attribute_user_editable(attribute, attr_id=""):
    normalized = str(attr_id or attribute.get("id") or "").upper()
    if normalized in CLONE_INTERNAL_READ_ONLY_ATTRIBUTES:
        return False
    if normalized in CLONE_ATTRIBUTE_ALIASES:
        return False
    tags = attribute.get("tags") or {}
    return not (isinstance(tags, dict) and tags.get("read_only") is True)


def source_attribute_label(source_item, attr_id):
    wanted = str(attr_id or "").upper()
    aliases = {wanted}
    aliases.update(key for key, value in CLONE_ATTRIBUTE_ALIASES.items() if value == wanted)
    for container in [source_item, *(source_item.get("variations") or [])]:
        sections = [container.get("attributes") or []]
        if container is not source_item:
            sections.append(container.get("attribute_combinations") or [])
        for section in sections:
            for attribute in section:
                if str(attribute.get("id") or "").upper() in aliases and attribute.get("name"):
                    return str(attribute.get("name"))
    return ""


def sanitize_clone_answers(answers, category_attributes):
    if not category_attributes:
        return dict(answers or {})
    allowed = category_attribute_ids(category_attributes)
    clean = {}
    for key, value in (answers or {}).items():
        if not key.startswith("attribute:"):
            clean[key] = value
            continue
        attr_id = canonical_clone_attribute_id(key.split(":", 1)[1])
        if attr_id and (attr_id in allowed or attr_id in CLONE_ALWAYS_ALLOWED_ATTRIBUTES):
            clean[f"attribute:{attr_id}"] = value
    return clean


def sanitize_clone_payload_attributes(create_payload, category_attributes):
    if not category_attributes:
        return []
    allowed = category_attribute_ids(category_attributes)
    removed = []

    def sanitize_section(section):
        kept = []
        seen = set()
        for attribute in section or []:
            original_id = str(attribute.get("id") or "").upper()
            attr_id = canonical_clone_attribute_id(original_id)
            clean = attribute
            if original_id in GTIN_IDENTIFIER_ATTRS and original_id != "GTIN" and "GTIN" in allowed and original_id not in allowed:
                attr_id = "GTIN"
                clean = {**attribute, "id": "GTIN"}
            if original_id in CLONE_ATTRIBUTE_ALIASES or not attr_id:
                removed.append(original_id)
            elif attr_id in allowed or attr_id in CLONE_ALWAYS_ALLOWED_ATTRIBUTES:
                signature = (attr_id, clone_attribute_display_value(clean), clean.get("value_id"))
                if signature not in seen:
                    seen.add(signature)
                    kept.append(clean)
            else:
                removed.append(original_id)
        return kept

    create_payload["attributes"] = sanitize_section(create_payload.get("attributes") or [])
    for variation in create_payload.get("variations") or []:
        variation["attributes"] = sanitize_section(variation.get("attributes") or [])
        variation["attribute_combinations"] = sanitize_section(variation.get("attribute_combinations") or [])
    return list(dict.fromkeys(removed))


def source_attribute_id_from_label(source_item, label):
    wanted = normalized_attribute_label(label)
    for attribute in source_item.get("attributes") or []:
        if normalized_attribute_label(attribute.get("name") or "") == wanted:
            return str(attribute.get("id") or "").upper()
    return ""


def pending_clone_attribute(attr_id, source_item, category_attributes, item_id="", fallback_label=""):
    original_id = str(attr_id or "").replace("attribute:", "").upper()
    resolved_id = canonical_clone_attribute_id(original_id)
    if not resolved_id:
        return {}
    definition = category_attribute_definition(category_attributes, resolved_id, fallback_label)
    if definition and not clone_attribute_user_editable(definition, resolved_id):
        return {}
    source_label = source_attribute_label(source_item, resolved_id)
    label = clone_attribute_label(resolved_id, source_label or definition.get("name") or fallback_label)
    options = []
    for value in definition.get("values") or []:
        name = clean_attribute_value(value.get("name"))
        if name and name not in options:
            options.append(name)
    if resolved_id == "SELLER_PACKAGE_TYPE" and not options:
        options = ["Sem embalagem adicional", "Com embalagem adicional"]
    value_type = str(definition.get("value_type") or ("number_unit" if resolved_id.startswith("SELLER_PACKAGE_") and resolved_id != "SELLER_PACKAGE_TYPE" else "string")).lower()
    units = clone_attribute_units(definition)
    if not units and resolved_id in {"SELLER_PACKAGE_HEIGHT", "SELLER_PACKAGE_LENGTH", "SELLER_PACKAGE_WIDTH"}:
        units = ["cm"]
    elif not units and resolved_id == "SELLER_PACKAGE_WEIGHT":
        units = ["g"]
    kind = "select" if options else "number" if value_type in {"number", "number_unit"} else "text"
    max_length = definition.get("value_max_length")
    message = "Selecione ou informe o valor obrigatório exigido pelo Mercado Livre."
    if units:
        message = f"Informe o valor e use uma das unidades aceitas: {', '.join(units)}."
    elif max_length:
        message = f"Informe o valor obrigatório com no máximo {max_length} caracteres."
    default_value = ""
    if resolved_id == "GTIN":
        default_value = ", ".join(source_clone_identifiers(source_item))
    return clone_pending_field(
        f"attribute:{resolved_id}",
        label,
        kind,
        message,
        item_id,
        options,
        units=units,
        max_length=max_length,
        value_type=value_type,
        default_value=default_value,
    )


def required_clone_attributes_from_error(exc, source_item, category_attributes):
    text = meli_error_text(exc)
    found = []

    for match in re.finditer(
        r"(?:attributes?|atributos?)\s*\[([^\]]+)\]\s*(?:are|is|são|sao)\s+(?:all\s+)?required",
        text,
        flags=re.I,
    ):
        for raw_id in match.group(1).split(","):
            attr_id = raw_id.strip().strip("'\"").upper()
            if re.fullmatch(r"[A-Z0-9_]+", attr_id):
                found.append((attr_id, ""))

    human_patterns = (
        r'(?:campo|atributo)\s+["\']([^"\']+)["\']\s+(?:é|e)\s+obrigat[oó]rio',
        r'(?:field|attribute)\s+["\']([^"\']+)["\']\s+is\s+required',
    )
    for pattern in human_patterns:
        for match in re.finditer(pattern, text, flags=re.I):
            label = match.group(1).strip()
            definition = category_attribute_definition(category_attributes, label=label)
            attr_id = str(definition.get("id") or "").upper() or source_attribute_id_from_label(source_item, label)
            if not attr_id:
                attr_id = attribute_id_from_human_name(label)
            found.append((attr_id, label))

    for match in re.finditer(
        r"(?:attribute|atributo)\s+([A-Z][A-Z0-9_]{2,})\s+(?:is\s+required|(?:é|e)\s+obrigat[oó]rio)",
        text,
        flags=re.I,
    ):
        found.append((match.group(1).upper(), ""))

    for cause in meli_error_causes(exc):
        code_message = f"{cause.get('code') or ''} {cause.get('message') or ''}".lower()
        if "required" not in code_message and "obrig" not in normalized_attribute_label(code_message):
            continue
        for reference in cause.get("references") or []:
            tokens = re.findall(r"[A-Z][A-Z0-9_]{2,}", str(reference or ""))
            for attr_id in tokens:
                if attr_id not in {"BODY", "ITEM", "ATTRIBUTES"}:
                    found.append((attr_id, ""))

    deduped = []
    seen = set()
    for attr_id, label in found:
        attr_id = str(attr_id or "").upper()
        if not attr_id or attr_id in seen:
            continue
        seen.add(attr_id)
        pending = pending_clone_attribute(attr_id, source_item, category_attributes, fallback_label=label)
        if pending:
            deduped.append(pending)
    return deduped


def dropped_clone_attributes_from_error(exc):
    text = meli_error_text(exc)
    attr_ids = []
    patterns = (
        r"Attribute:\s*([A-Z0-9_]+)\s+was dropped",
        r"Attribute\s+\[?([A-Z0-9_]+)\]?\s+(?:was dropped|does not exist|does not exists)",
        r"Atributo\s+\[?([A-Z0-9_]+)\]?\s+(?:não existe|nao existe|foi removido)",
    )
    for pattern in patterns:
        attr_ids.extend(match.group(1).upper() for match in re.finditer(pattern, text, flags=re.I))
    return list(dict.fromkeys(attr_ids))


def fill_missing_clone_fields(create_payload, source_item, fields):
    for field in fields:
        if create_payload.get(field):
            continue
        if field == "family_name":
            create_payload[field] = normalize_family_name(
                source_item.get("family_name")
                or source_attribute_value(source_item, ["FAMILY_NAME", "MODEL", "LINE", "BRAND"])
                or (source_item.get("title") or "")[:60]
            )
        elif field in source_item and source_item.get(field) not in (None, "", [], {}):
            create_payload[field] = source_item.get(field)
    return create_payload


def clone_payload_from_answers(create_payload, answers):
    answers = answers or {}
    for key, value in answers.items():
        if not clean_attribute_value(value):
            continue
        if key.startswith("attribute:"):
            attr_id = canonical_clone_attribute_id(key.split(":", 1)[1])
            if attr_id == "GTIN":
                apply_clone_gtin_override(create_payload, value)
            elif attr_id == "EMPTY_GTIN_REASON":
                remove_clone_attributes(create_payload, GTIN_IDENTIFIER_ATTRS)
                for variation in create_payload.get("variations") or []:
                    variation["attributes"] = remove_clone_product_identifiers(variation.get("attributes") or [])
                add_or_update_clone_attribute(create_payload, attr_id, value)
            else:
                add_or_update_clone_attribute(create_payload, attr_id, value)
        elif key == "family_name":
            create_payload["family_name"] = normalize_family_name(value)
        elif key in {"title", "price", "available_quantity", "listing_type_id", "condition"}:
            create_payload[key] = value
    return create_payload


def clone_retry_adjustments_from_error(exc, create_payload, source_item, pending_fields, item_id="", category_attributes=None):
    changed = False
    adjustments = []
    error_text = meli_error_text(exc)
    lowered_error = error_text.lower()
    immutable_terms = [
        match.group(1).upper()
        for match in re.finditer(r"(?:not allowed to modify sale term|não é permitido alterar o termo de venda)\s+([A-Z0-9_]+)", error_text, flags=re.I)
    ]
    removed_terms = remove_clone_sale_terms(create_payload, immutable_terms)
    if removed_terms:
        changed = True
        adjustments.append({"tipo": "termos_de_venda_imutaveis_removidos", "campos": removed_terms})
    if "official_store_id" in lowered_error and (
        "not allowed" in lowered_error or "invalid_official_store_id" in lowered_error or "invalid official" in lowered_error
    ):
        if strip_official_store_clone_fields(create_payload):
            changed = True
            adjustments.append({"tipo": "loja_oficial_incompativel_removida", "campos": ["official_store_id"]})
    dropped_attrs = dropped_clone_attributes_from_error(exc)
    removed_dropped_attrs = remove_clone_attributes(create_payload, dropped_attrs)
    if removed_dropped_attrs:
        changed = True
        adjustments.append({"tipo": "atributos_inexistentes_removidos", "campos": removed_dropped_attrs})
    for field in required_clone_attributes_from_error(exc, source_item, category_attributes or []):
        attr_id = str(field.get("id") or "").replace("attribute:", "")
        if restore_clone_attribute_from_source(create_payload, source_item, category_attributes or [], attr_id):
            changed = True
            adjustments.append({"tipo": "atributo_recuperado_do_anuncio_original", "campos": [canonical_clone_attribute_id(attr_id)]})
            continue
        field["item_id"] = item_id
        pending_fields.append(field)
    normalized_error = normalized_attribute_label(error_text)
    if "required" in normalized_error or "obrigatorio" in normalized_error:
        for attribute in category_attributes or []:
            attr_id = attribute.get("id")
            if (
                not attr_id
                or not clone_attribute_is_required(attribute)
                or not clone_attribute_user_editable(attribute, attr_id)
                or clone_required_attribute_satisfied(create_payload, attr_id)
            ):
                continue
            if restore_clone_attribute_from_source(create_payload, source_item, category_attributes or [], attr_id):
                changed = True
                adjustments.append({"tipo": "atributo_recuperado_do_anuncio_original", "campos": [canonical_clone_attribute_id(attr_id)]})
                continue
            pending_fields.append(
                pending_clone_attribute(
                    attr_id,
                    source_item,
                    category_attributes or [],
                    item_id,
                    attribute.get("name") or "",
                )
            )
    missing_fields = required_fields_from_error(exc)
    if missing_fields:
        body_field_names = {
            "family_name", "title", "category_id", "price", "currency_id", "available_quantity",
            "buying_mode", "listing_type_id", "condition", "pictures", "domain_id", "catalog_product_id",
        }
        body_fields = [field for field in missing_fields if str(field).lower() in body_field_names]
        before = json.dumps(create_payload, sort_keys=True, ensure_ascii=False)
        create_payload = fill_missing_clone_fields(create_payload, source_item, body_fields)
        after = json.dumps(create_payload, sort_keys=True, ensure_ascii=False)
        if before != after:
            changed = True
            adjustments.append({"tipo": "campos_obrigatorios_preenchidos", "campos": body_fields})
        for field in body_fields:
            if not create_payload.get(field):
                pending_fields.append(clone_pending_field(field, field.replace("_", " ").title(), "text", "Campo obrigatório exigido pelo Mercado Livre.", item_id))
        for field in missing_fields:
            if str(field).lower() in body_field_names or str(field).lower() == "official_store_id":
                continue
            canonical_id = canonical_clone_attribute_id(field)
            if not canonical_id:
                continue
            if restore_clone_attribute_from_source(create_payload, source_item, category_attributes or [], canonical_id):
                changed = True
                adjustments.append({"tipo": "atributo_recuperado_do_anuncio_original", "campos": [canonical_id]})
                continue
            pending = pending_clone_attribute(canonical_id, source_item, category_attributes or [], item_id)
            if pending:
                pending_fields.append(pending)

    invalid_fields = invalid_fields_from_error(exc)
    removed_fields = []
    for field in invalid_fields:
        if field in create_payload:
            create_payload.pop(field, None)
            removed_fields.append(field)
            changed = True
    if removed_fields:
        adjustments.append({"tipo": "campos_invalidos_removidos", "campos": removed_fields})

    causes = meli_error_causes(exc)
    attrs_to_remove = []
    for cause in causes:
        code = str(cause.get("code") or "")
        message = str(cause.get("message") or "")
        cause_type = str(cause.get("type") or "")
        if code in {"item.attributes.ignored", "invalid.item.attribute.values", "item.attribute.invalid_product_identifier"}:
            attrs_to_remove.extend(attribute_ids_from_error_text(message))
        if code == "item.attribute.invalid_product_identifier":
            source_identifiers = source_clone_identifiers(source_item)
            attrs_to_remove.extend(GTIN_IDENTIFIER_ATTRS)
            gtin_definition = category_attribute_definition(category_attributes or [], "GTIN")
            if gtin_definition and clone_attribute_is_required(gtin_definition):
                pending = pending_clone_attribute("GTIN", source_item, category_attributes or [], item_id)
                if pending:
                    pending["message"] = "Informe um GTIN/EAN/UPC válido. Para variações, separe um código por vírgula para cada variação."
                    if source_identifiers:
                        pending["default_value"] = ", ".join(source_identifiers)
                    pending_fields.append(pending)
                if not source_identifiers:
                    empty_reason = pending_clone_attribute("EMPTY_GTIN_REASON", source_item, category_attributes or [], item_id)
                    if empty_reason:
                        empty_reason["message"] = "Use este campo somente quando o produto realmente não possuir código universal."
                        pending_fields.append(empty_reason)
        if code == "item.attribute.invalid":
            attr_ids = attribute_ids_from_error_text(message)
            attr_ids.extend(attribute_id_from_human_name(name) for name in human_attribute_names_from_error_text(message))
            attrs_to_remove.extend(attr_ids)
            for attr_id in attr_ids:
                pending = pending_clone_attribute(attr_id, source_item, category_attributes or [], item_id)
                if pending:
                    pending["message"] = "O Mercado Livre exigiu um valor válido para este atributo."
                    pending_fields.append(pending)
        if "está incorreto" in message or "esta incorreto" in message or "value null was omitted" in message or "Struct or text must be present" in message:
            human_names = human_attribute_names_from_error_text(message)
            attr_ids = [attribute_id_from_human_name(name) for name in human_names]
            attr_ids.extend(attribute_ids_from_error_text(message))
            removed_attrs = remove_clone_attributes(create_payload, attr_ids)
            if removed_attrs:
                changed = True
                adjustments.append({"tipo": "atributos_com_valor_invalido_removidos", "campos": removed_attrs})
            for name in human_names:
                attr_id = attribute_id_from_human_name(name)
                pending = pending_clone_attribute(attr_id, source_item, category_attributes or [], item_id, name)
                if pending:
                    pending["message"] = "Informe exatamente no formato aceito pelo Mercado Livre."
                    pending_fields.append(pending)
        if code == "item.family_name.length_invalid" or "Family Name length is over" in message:
            if create_payload.get("family_name"):
                create_payload["family_name"] = normalize_family_name(create_payload.get("family_name"))
                changed = True
                adjustments.append({"tipo": "family_name_ajustado", "campos": ["family_name"]})
        if "shipping" in code or "mode me1" in message.lower():
            if create_payload.pop("shipping", None) is not None:
                changed = True
                adjustments.append({"tipo": "frete_removido_para_padrao_da_conta", "campos": ["shipping"]})
        if cause_type == "warning" and code == "item.attributes.ignored":
            attrs_to_remove.extend(attribute_ids_from_error_text(message))

    if "Struct or text must be present" in error_text or "value null was omitted" in error_text or "está incorreto" in error_text or "esta incorreto" in error_text:
        human_names = human_attribute_names_from_error_text(error_text)
        attr_ids = [attribute_id_from_human_name(name) for name in human_names]
        attr_ids.extend(attribute_ids_from_error_text(error_text))
        removed_attrs = remove_clone_attributes(create_payload, attr_ids)
        if removed_attrs:
            changed = True
            adjustments.append({"tipo": "atributos_com_valor_invalido_removidos", "campos": removed_attrs})
        for name in human_names:
            attr_id = attribute_id_from_human_name(name)
            pending = pending_clone_attribute(attr_id, source_item, category_attributes or [], item_id, name)
            if pending:
                pending["message"] = "Informe exatamente no formato aceito pelo Mercado Livre."
                pending_fields.append(pending)

    if "item.attribute.invalid" in error_text or "Value name of attribute" in error_text:
        for attr_id in attribute_ids_from_error_text(error_text):
            pending = pending_clone_attribute(attr_id, source_item, category_attributes or [], item_id)
            if pending:
                pending["message"] = "Informe um valor válido para este campo obrigatório."
                pending_fields.append(pending)

    removed_attrs = remove_clone_attributes(create_payload, attrs_to_remove)
    if removed_attrs:
        changed = True
        adjustments.append({"tipo": "atributos_removidos", "campos": removed_attrs})

    if create_payload.get("family_name") and len(str(create_payload["family_name"])) > 60:
        create_payload["family_name"] = normalize_family_name(create_payload["family_name"])
        changed = True
        adjustments.append({"tipo": "family_name_ajustado", "campos": ["family_name"]})

    return create_payload, changed, adjustments


def friendly_clone_error(exc):
    if meli_rate_limited_error(exc):
        return "O Mercado Livre limitou temporariamente as criações. A aplicação tentou novamente com espera progressiva; aguarde alguns instantes e repita se necessário."
    error_text = meli_error_text(exc).lower()
    if "official_store_id" in error_text:
        return "O Mercado Livre recusou o vínculo de loja oficial da origem. A aplicação tentou removê-lo automaticamente neste mesmo processamento."
    causes = meli_error_causes(exc)
    if not causes:
        return str(exc)
    messages = []
    for cause in causes:
        code = str(cause.get("code") or "")
        message = str(cause.get("message") or "")
        if code == "item.family_name.length_invalid":
            messages.append("Nome de família do catálogo passou de 60 caracteres; ajuste o campo solicitado e tente novamente.")
        elif code == "item.attribute.invalid_product_identifier":
            messages.append("Código universal do produto não pode ser reutilizado nessa categoria. O clone remove esse código; se a categoria exigir, informe um novo.")
        elif code == "invalid.item.attribute.values":
            attrs = ", ".join(attribute_ids_from_error_text(message)) or "atributo"
            messages.append(f"Atributo com valor inválido: {attrs}. Informe um valor válido ou deixe o clone remover quando não for obrigatório.")
        elif code == "item.attribute.invalid":
            attrs = ", ".join(attribute_ids_from_error_text(message)) or "atributo"
            messages.append(f"Atributo obrigatório sem valor válido: {attrs}.")
        elif code == "item.attributes.ignored":
            continue
        elif code:
            messages.append(message or code)
    return " ".join(dict.fromkeys(messages)) or str(exc)


def create_item_with_clone_retries(target_client, create_payload, source_item, answers=None, item_id="", category_attributes=None, cross_account=False, destination_store_id=None):
    payload = json.loads(json.dumps(create_payload, ensure_ascii=False))
    sanitized_attributes = sanitize_clone_payload_attributes(payload, category_attributes or [])
    payload = clone_payload_from_answers(payload, sanitize_clone_answers(answers or {}, category_attributes or []))
    adjustments = []
    if sanitized_attributes:
        adjustments.append({"tipo": "atributos_fora_da_categoria_removidos", "campos": sanitized_attributes})
    last_error = None
    pending_fields = []
    validation_attempts = 0
    rate_limit_attempts = 0
    max_rate_limit_attempts = max(0, int(os.getenv("MELI_CLONE_RATE_LIMIT_RETRIES", "3")))
    while validation_attempts < 5:
        try:
            if cross_account:
                removed_store = prepare_cross_account_official_store_payload(payload, destination_store_id)
                if removed_store and not any(row.get("tipo") == "vinculo_loja_oficial_origem_substituido" for row in adjustments):
                    adjustments.append({
                        "tipo": "vinculo_loja_oficial_origem_substituido",
                        "campos": ["official_store_id"],
                        "destino": destination_store_id,
                    })
            created = target_client.create_item(payload)
            if adjustments and isinstance(created, dict):
                created["_clone_adjustments"] = adjustments
            return created
        except Exception as exc:
            last_error = exc
            if meli_rate_limited_error(exc) and rate_limit_attempts < max_rate_limit_attempts:
                delay = clone_rate_limit_delay(rate_limit_attempts)
                adjustments.append({"tipo": "limite_temporario_aguardado", "tentativa": rate_limit_attempts + 1, "espera_segundos": round(delay, 2)})
                rate_limit_attempts += 1
                time.sleep(delay)
                continue
            validation_attempts += 1
            payload, changed, new_adjustments = clone_retry_adjustments_from_error(
                exc,
                payload,
                source_item,
                pending_fields,
                item_id,
                category_attributes or [],
            )
            adjustments.extend(new_adjustments)
            if pending_fields:
                break
            if not changed:
                break
    if pending_fields:
        error = RuntimeError("Revise campos obrigatórios antes de copiar este anúncio.")
        error.pending_fields = dedupe_pending_fields(pending_fields)
        error.original_error = str(last_error)
        raise error
    raise last_error


def create_official_clone(payload, job, source_item_id, edits):
    source_account = official_account_by_name(payload, job.get("source_account_id") or job.get("source"))
    target_account = official_account_by_name(payload, job.get("target_account_id") or job.get("target"))
    if not source_account or not source_account.get("official"):
        raise RuntimeError("Conta origem oficial não encontrada para clonar via Mercado Livre.")
    if not target_account or not target_account.get("official"):
        raise RuntimeError("Conta destino oficial não encontrada para clonar via Mercado Livre.")
    source_client = account_client(source_account)
    target_client = account_client(target_account)
    bundle = clone_source_bundle(payload, source_account.get("id"), source_item_id, include_description=True)
    source_item = bundle["source_item"]
    source_account_id = str(source_account.get("id") or source_account.get("seller_id") or "")
    target_account_id = str(target_account.get("id") or target_account.get("seller_id") or "")
    cross_account = bool(source_account_id and target_account_id and source_account_id != target_account_id)
    create_payload = build_clone_item_payload(source_item, edits)
    try:
        category_attributes = cached_category_attributes(source_client, source_item.get("category_id"))
    except Exception:
        category_attributes = []
    catalog_product = bundle.get("catalog_product") or {}
    if catalog_product:
        source_item["_clone_catalog_product"] = catalog_product
    hydrate_clone_package_attributes(create_payload, source_item)
    hydrate_required_clone_attributes(create_payload, source_item, category_attributes, catalog_product)
    create_payload = apply_target_account_clone_rules(create_payload, source_item, source_account, target_account)
    destination_store_id = target_official_store_id(target_client, target_account, source_item) if cross_account else None
    answers = (job.get("field_answers") or {}).get(source_item_id) or {}
    created = create_item_with_clone_retries(
        target_client,
        create_payload,
        source_item,
        answers,
        source_item_id,
        category_attributes,
        cross_account,
        destination_store_id,
    )
    verified_item = {}
    if created.get("id"):
        try:
            verified_item = target_client.item(created["id"])
        except Exception as exc:
            created["_verification_warning"] = f"Anúncio criado, mas ainda não apareceu na leitura imediata da API: {exc}"
    description_text = (edits.get("description") or "").strip()
    if not description_text:
        description_text = bundle.get("description") or ""
    if description_text and created.get("id"):
        try:
            target_client.create_item_description(created["id"], description_text)
        except Exception:
            pass
    return created, source_item, verified_item


def clone_payload_has_attribute(create_payload, attr_id):
    attr_id = str(attr_id or "").upper()
    def section_has(section):
        for attribute in section or []:
            if str(attribute.get("id") or "").upper() != attr_id:
                continue
            if clean_attribute_value(attribute.get("value_name")) or attribute.get("value_id") or attribute.get("values"):
                return True
        return False

    if section_has(create_payload.get("attributes") or []):
        return True
    variations = create_payload.get("variations") or []
    if variations and all(
        section_has([*(variation.get("attributes") or []), *(variation.get("attribute_combinations") or [])])
        for variation in variations
    ):
        return True
    return False


def clone_preflight_pending_fields(payload, source_name, item_ids, edits):
    source_account = official_account_by_name(payload, source_name)
    if not source_account or not source_account.get("official"):
        return []
    client = account_client(source_account)
    pending = []
    category_cache = {}
    for item_id in item_ids:
        try:
            bundle = clone_source_bundle(payload, source_account.get("id"), item_id)
            source_item = bundle["source_item"]
            create_payload = build_clone_item_payload(source_item, edits)
            category_id = create_payload.get("category_id")
            if not category_id:
                continue
            if category_id not in category_cache:
                category_cache[category_id] = cached_category_attributes(client, category_id)
            catalog_product = bundle.get("catalog_product") or {}
            hydrate_clone_package_attributes(create_payload, source_item)
            hydrate_required_clone_attributes(
                create_payload,
                source_item,
                category_cache.get(category_id) or [],
                catalog_product,
            )
            item_fields = []
            for attribute in category_cache.get(category_id) or []:
                attr_id = attribute.get("id")
                if not clone_attribute_is_required(attribute) or not attr_id or not clone_attribute_user_editable(attribute, attr_id):
                    continue
                if clone_required_attribute_satisfied(create_payload, attr_id):
                    continue
                item_fields.append(
                    pending_clone_attribute(
                        attr_id,
                        source_item,
                        category_cache.get(category_id) or [],
                        item_id,
                        attribute.get("name") or "",
                    )
                )
            if item_fields:
                pending.append(
                    {
                        "item_id": item_id,
                        "error": "O Mercado Livre exige informações antes de criar este anúncio.",
                        "pending_fields": dedupe_pending_fields(item_fields),
                    }
                )
        except Exception:
            continue
    return pending


def clone_source_snapshot(payload, account_identifier, item_id):
    bundle = clone_source_bundle(payload, account_identifier, item_id, include_description=True)
    source_item = bundle["source_item"]
    catalog_product = bundle.get("catalog_product") or {}
    identifiers = source_clone_identifiers(source_item, catalog_product)
    description_text = bundle.get("description") or ""
    sku = item_sku(source_item)
    return {
        "item_id": source_item.get("id") or item_id,
        "title": source_item.get("title") or "",
        "sku": "" if sku in (None, "", "-") else sku,
        "price": source_item.get("price") or 0,
        "stock": item_available_quantity(source_item),
        "listing_type_id": source_item.get("listing_type_id") or "",
        "catalog_listing": is_catalog_listing(source_item),
        "description": description_text,
        "gtin": ", ".join(identifiers),
        "identifier_count": len(identifiers),
        "variation_count": len(source_item.get("variations") or []),
        "hydrated_from_user_product": bool(source_item.get("_user_product_hydrated")),
    }


def statistics_date_window(date_from, date_to):
    try:
        start = date.fromisoformat(str(date_from or "")[:10])
        end = date.fromisoformat(str(date_to or "")[:10])
    except ValueError as exc:
        raise RuntimeError("Informe um período válido para consultar as vendas.") from exc
    if end < start:
        raise RuntimeError("A data final precisa ser igual ou posterior à data inicial.")
    maximum_days = max(1, int(os.getenv("MELI_STATISTICS_MAX_DAYS", "366")))
    if (end - start).days + 1 > maximum_days:
        raise RuntimeError(f"O período máximo por consulta é de {maximum_days} dias.")
    start_at = datetime.combine(start, datetime_time.min).replace(tzinfo=APP_TZ)
    end_at = datetime.combine(end, datetime_time.max).replace(tzinfo=APP_TZ)
    return start, end, start_at.isoformat(timespec="milliseconds"), end_at.isoformat(timespec="milliseconds")


def statistics_order_windows(start, end):
    days_per_window = max(1, min(31, int(os.getenv("MELI_STATISTICS_WINDOW_DAYS", "31"))))
    current = start
    while current <= end:
        window_end = min(end, current + timedelta(days=days_per_window - 1))
        start_at = datetime.combine(current, datetime_time.min).replace(tzinfo=APP_TZ)
        end_at = datetime.combine(window_end, datetime_time.max).replace(tzinfo=APP_TZ)
        yield start_at.isoformat(timespec="milliseconds"), end_at.isoformat(timespec="milliseconds")
        current = window_end + timedelta(days=1)


def fetch_statistics_orders(client, seller_id, start, end):
    maximum = max(1, int(os.getenv("MELI_STATISTICS_ORDERS_LIMIT", "50000")))
    orders = []
    seen = set()
    truncated = False
    for window_from, window_to in statistics_order_windows(start, end):
        offset = 0
        while len(orders) < maximum:
            data = client.seller_orders(
                seller_id,
                limit=50,
                offset=offset,
                date_from=window_from,
                date_to=window_to,
            )
            batch = data.get("results") or []
            for order in batch:
                order_id = str(order.get("id") or "")
                signature = order_id or json.dumps(order, sort_keys=True, ensure_ascii=False)
                if signature in seen:
                    continue
                seen.add(signature)
                orders.append(order)
                if len(orders) >= maximum:
                    truncated = True
                    break
            total = int((data.get("paging") or {}).get("total") or len(batch))
            offset += len(batch)
            if not batch or offset >= total or len(orders) >= maximum:
                break
        if len(orders) >= maximum:
            truncated = True
            break
    return orders, truncated


def shipment_id_from_order(order):
    shipping = order.get("shipping")
    if isinstance(shipping, dict):
        return shipping.get("id") or shipping.get("shipment_id")
    if shipping not in (None, "", 0, "0"):
        return shipping
    return order.get("shipping_id")


def flex_hint_from_payload(value):
    if not isinstance(value, dict):
        return None
    shipping = value.get("shipping") if isinstance(value.get("shipping"), dict) else {}
    candidates = [
        value.get("logistic_type"),
        value.get("shipping_logistic_type"),
        first_present(value, ["logistic.type", "logistic_type"], ""),
        shipping.get("logistic_type"),
        first_present(
            value,
            [
                "shipping_option.logistic_type",
                "shipping_option.logistic.type",
                "shipping_option.shipping_method.logistic_type",
                "shipping_option.shipping_method.logistic.type",
            ],
            "",
        ),
    ]
    tags = {str(tag).strip().lower() for tag in value.get("tags") or [] if str(tag).strip()}
    normalized_candidates = {str(candidate or "").strip().lower() for candidate in candidates if str(candidate or "").strip()}
    if "self_service" in normalized_candidates or tags.intersection({"self_service", "self_service_in"}):
        return True
    if normalized_candidates or tags.intersection({"self_service_out", "self_service_available"}):
        return False
    return None


def order_is_flex(client, order, catalog_by_id, shipment_cache, lookup_state):
    hint = flex_hint_from_payload(order)
    if hint is not None:
        return hint
    shipment_id = shipment_id_from_order(order)
    maximum_lookups = max(0, int(lookup_state.get("maximum", os.getenv("MELI_STATISTICS_SHIPMENT_LOOKUPS", "500"))))
    if shipment_id not in (None, "", 0, "0"):
        cache_key = str(shipment_id)
        if cache_key in shipment_cache:
            return shipment_cache[cache_key]
        cache_ttl = max(300, int(os.getenv("MELI_STATISTICS_SHIPMENT_CACHE_SECONDS", "86400")))
        now = time.monotonic()
        with SHIPMENT_MODE_CACHE_LOCK:
            cached = SHIPMENT_MODE_CACHE.get(cache_key)
        if cached and now - cached.get("time", 0) < cache_ttl:
            shipment_cache[cache_key] = cached.get("flex")
            return cached.get("flex")
        elif lookup_state["count"] < maximum_lookups:
            lookup_state["count"] += 1
            try:
                shipment = client.shipment(shipment_id)
            except Exception:
                shipment = {}
            hint = flex_hint_from_payload(shipment)
            shipment_cache[cache_key] = hint
            with SHIPMENT_MODE_CACHE_LOCK:
                SHIPMENT_MODE_CACHE[cache_key] = {"time": now, "flex": hint}
            return hint
        else:
            shipment_cache[cache_key] = None
    # Never infer a historical order from the listing's current Flex state. A
    # listing can enter or leave Flex after the sale, which previously inflated
    # broad-period Flex statistics.
    return None


def order_item_sku(order_item, catalog_item):
    item = order_item.get("item") or {}
    candidates = (
        order_item.get("seller_sku"),
        item.get("seller_sku"),
        item.get("seller_custom_field"),
        catalog_item.get("sku"),
    )
    return next((clean_attribute_value(value) for value in candidates if clean_attribute_value(value)), "")


def aggregate_sku_statistics(account_orders, catalog, maximum_shipment_lookups=None):
    catalog_by_id = {str(item.get("id")): item for item in catalog or [] if item.get("id")}
    selected_accounts = {
        str(account.get("nickname") or "")
        for account, _client, _orders, _truncated in account_orders
        if account.get("nickname")
    }
    current_stock_by_sku = {}
    for catalog_item in catalog or []:
        if selected_accounts and str(catalog_item.get("account") or "") not in selected_accounts:
            continue
        sku_key = clean_attribute_value(catalog_item.get("sku")).upper()
        if not sku_key or sku_key == "-":
            continue
        try:
            current_stock_by_sku[sku_key] = current_stock_by_sku.get(sku_key, 0) + max(0, int(catalog_item.get("stock") or 0))
        except (TypeError, ValueError):
            continue
    aggregates = {}
    ignored_statuses = {"cancelled", "canceled", "invalid"}
    warnings = []
    truncated = False
    for account, client, orders, account_truncated in account_orders:
        truncated = truncated or account_truncated
        shipment_cache = {}
        lookup_state = {
            "count": 0,
            "maximum": max(
                0,
                int(
                    maximum_shipment_lookups
                    if maximum_shipment_lookups is not None
                    else os.getenv("MELI_STATISTICS_SHIPMENT_LOOKUPS", "500")
                ),
            ),
        }
        for order in orders:
            if str(order.get("status") or "").lower() in ignored_statuses:
                continue
            order_id = str(order.get("id") or "-")
            flex = order_is_flex(client, order, catalog_by_id, shipment_cache, lookup_state)
            for order_item in order.get("order_items") or []:
                item = order_item.get("item") or {}
                item_id = str(item.get("id") or "")
                catalog_item = catalog_by_id.get(item_id) or {}
                try:
                    quantity = max(0, int(float(order_item.get("quantity") or 0)))
                except (TypeError, ValueError):
                    quantity = 0
                if quantity <= 0:
                    continue
                sku = order_item_sku(order_item, catalog_item)
                key = sku.upper() if sku else f"SEM-SKU:{item_id or order_id}"
                row = aggregates.setdefault(
                    key,
                    {
                        "sku": sku or "Sem SKU",
                        "product": item.get("title") or catalog_item.get("title") or "Produto sem título",
                        "thumbnail": item.get("thumbnail") or item.get("secure_thumbnail") or catalog_item.get("thumbnail") or "",
                        "units": 0,
                        "flex_units": 0,
                        "non_flex_units": 0,
                        "unknown_units": 0,
                        "revenue": 0.0,
                        "flex_revenue": 0.0,
                        "non_flex_revenue": 0.0,
                        "unknown_revenue": 0.0,
                        "accounts": set(),
                        "item_ids": set(),
                        "order_ids": set(),
                        "flex_order_ids": set(),
                        "non_flex_order_ids": set(),
                        "unknown_order_ids": set(),
                    },
                )
                unit_price = order_item.get("unit_price") or order_item.get("full_unit_price") or 0
                try:
                    line_total = float(unit_price or 0) * quantity
                except (TypeError, ValueError):
                    line_total = 0
                row["units"] += quantity
                row["revenue"] += line_total
                row["accounts"].add(account.get("nickname") or str(account.get("seller_id") or "Conta"))
                if item_id:
                    row["item_ids"].add(item_id)
                row["order_ids"].add(order_id)
                if flex is True:
                    row["flex_units"] += quantity
                    row["flex_revenue"] += line_total
                    row["flex_order_ids"].add(order_id)
                elif flex is False:
                    row["non_flex_units"] += quantity
                    row["non_flex_revenue"] += line_total
                    row["non_flex_order_ids"].add(order_id)
                else:
                    row["unknown_units"] += quantity
                    row["unknown_revenue"] += line_total
                    row["unknown_order_ids"].add(order_id)
    rows = []
    for row in aggregates.values():
        rows.append(
            {
                **{key: value for key, value in row.items() if not isinstance(value, set)},
                "revenue": round(row["revenue"], 2),
                "flex_revenue": round(row["flex_revenue"], 2),
                "non_flex_revenue": round(row["non_flex_revenue"], 2),
                "unknown_revenue": round(row["unknown_revenue"], 2),
                "current_stock": current_stock_by_sku.get(str(row.get("sku") or "").upper(), 0),
                "accounts": sorted(row["accounts"]),
                "item_ids": sorted(row["item_ids"]),
                "order_ids": sorted(row["order_ids"]),
                "flex_order_ids": sorted(row["flex_order_ids"]),
                "non_flex_order_ids": sorted(row["non_flex_order_ids"]),
                "unknown_order_ids": sorted(row["unknown_order_ids"]),
            }
        )
    rows.sort(key=lambda row: (-row["units"], -row["revenue"], row["sku"]))
    return {"rows": rows, "truncated": truncated, "warnings": warnings}


def query_sku_statistics(payload, request):
    start, end, _, _ = statistics_date_window(request.get("date_from"), request.get("date_to"))
    account_filter = str(request.get("account") or "all")
    requested_flex = str(request.get("flex") or "all")
    default_lookups = int(os.getenv("MELI_STATISTICS_SHIPMENT_LOOKUPS", "500"))
    focused_lookups = int(os.getenv("MELI_STATISTICS_FLEX_LOOKUPS", "5000"))
    maximum_lookups = focused_lookups if requested_flex in {"flex", "non_flex"} else default_lookups
    accounts = [
        account
        for account in payload.get("accounts") or []
        if account.get("official") and account.get("access_token") and account.get("status") == "connected"
    ]
    if account_filter != "all":
        accounts = [
            account
            for account in accounts
            if account.get("id") == account_filter
            or str(account.get("seller_id")) == account_filter
            or account.get("nickname") == account_filter
        ]
    if not accounts:
        raise RuntimeError("Nenhuma conta oficial conectada corresponde ao filtro selecionado.")

    cache_key = (
        "flex-tristate-v2",
        tuple(sorted(str(account.get("seller_id") or account.get("id")) for account in accounts)),
        start.isoformat(),
        end.isoformat(),
        maximum_lookups,
    )
    cache_ttl = max(30, int(os.getenv("MELI_STATISTICS_CACHE_SECONDS", "300")))
    now = time.monotonic()
    with STATISTICS_CACHE_LOCK:
        cached = STATISTICS_CACHE.get(cache_key)
    if cached and now - cached["time"] < cache_ttl:
        base = json.loads(json.dumps(cached["result"], ensure_ascii=False))
        from_cache = True
    else:
        account_orders = []
        warnings = []
        for account in accounts:
            try:
                client = account_client(account)
                orders, truncated = fetch_statistics_orders(client, account.get("seller_id"), start, end)
                account_orders.append((account, client, orders, truncated))
            except Exception as exc:
                warnings.append(f"{account.get('nickname')}: {policy_error_message(exc, 'a leitura das vendas do período')}")
        if not account_orders and warnings:
            raise RuntimeError(" ".join(warnings))
        base = aggregate_sku_statistics(account_orders, payload.get("catalog") or [], maximum_lookups)
        unknown_units = sum(row.get("unknown_units") or 0 for row in base.get("rows") or [])
        if unknown_units:
            base.setdefault("warnings", []).append(
                f"{unknown_units} unidade(s) ficaram com a modalidade de envio não confirmada pela remessa histórica e não foram atribuídas ao Flex."
            )
        base["warnings"] = [*base.get("warnings", []), *warnings]
        with STATISTICS_CACHE_LOCK:
            STATISTICS_CACHE[cache_key] = {"time": now, "result": base}
        from_cache = False

    sku_filter = normalized_attribute_label(request.get("sku") or "")
    flex_filter = str(request.get("flex") or "all")
    rows = []
    selected_order_ids = set()
    for original in base.get("rows") or []:
        if sku_filter and sku_filter not in normalized_attribute_label(original.get("sku") or ""):
            continue
        row = dict(original)
        row["total_units"] = row.get("units") or 0
        if flex_filter == "flex":
            row["units"] = row.get("flex_units") or 0
            row["orders"] = len(row.get("flex_order_ids") or [])
            row["revenue"] = row.get("flex_revenue") or 0
            selected_order_ids.update(row.get("flex_order_ids") or [])
        elif flex_filter == "non_flex":
            row["units"] = row.get("non_flex_units") or 0
            row["orders"] = len(row.get("non_flex_order_ids") or [])
            row["revenue"] = row.get("non_flex_revenue") or 0
            selected_order_ids.update(row.get("non_flex_order_ids") or [])
        else:
            row["orders"] = len(row.get("order_ids") or [])
            selected_order_ids.update(row.get("order_ids") or [])
        if row["units"] <= 0:
            continue
        row.pop("order_ids", None)
        row.pop("flex_order_ids", None)
        row.pop("non_flex_order_ids", None)
        row.pop("unknown_order_ids", None)
        rows.append(row)
    rows.sort(key=lambda row: (-row["units"], -row["revenue"], row["sku"]))
    return {
        "ok": True,
        "date_from": start.isoformat(),
        "date_to": end.isoformat(),
        "account": account_filter,
        "flex": flex_filter,
        "rows": rows,
        "summary": {
            "units": sum(row.get("units") or 0 for row in rows),
            "skus": len(rows),
            "orders": len(selected_order_ids),
            "flex_units": (
                sum(row.get("units") or 0 for row in rows)
                if flex_filter == "flex"
                else 0
                if flex_filter == "non_flex"
                else sum(row.get("flex_units") or 0 for row in rows)
            ),
            "unknown_units": sum(row.get("unknown_units") or 0 for row in rows),
            "revenue": round(sum(row.get("revenue") or 0 for row in rows), 2),
        },
        "warnings": base.get("warnings") or [],
        "truncated": bool(base.get("truncated")),
        "cached": from_cache,
        "generated_at": now_label(),
    }


def statistics_job_signature(request):
    keys = ("account", "sku", "flex", "date_from", "date_to")
    normalized = {key: str((request or {}).get(key) or "") for key in keys}
    return json.dumps(normalized, sort_keys=True, ensure_ascii=False)


def cleanup_statistics_jobs(now=None):
    now = now or time.time()
    maximum_age = max(900, int(os.getenv("MELI_STATISTICS_JOB_TTL_SECONDS", "21600")))
    with STATISTICS_JOBS_LOCK:
        expired = [
            job_id
            for job_id, job in STATISTICS_JOBS.items()
            if now - float(job.get("created_epoch") or now) > maximum_age
        ]
        for job_id in expired:
            STATISTICS_JOBS.pop(job_id, None)


def start_statistics_job(request):
    cleanup_statistics_jobs()
    safe_request = {
        key: str((request or {}).get(key) or "")
        for key in ("account", "sku", "flex", "date_from", "date_to")
    }
    signature = statistics_job_signature(safe_request)
    with STATISTICS_JOBS_LOCK:
        existing = next(
            (
                job
                for job in STATISTICS_JOBS.values()
                if job.get("signature") == signature and job.get("status") in {"queued", "processing"}
            ),
            None,
        )
        if existing:
            return {key: value for key, value in existing.items() if key not in {"result", "signature", "created_epoch"}}
        job_id = uuid.uuid4().hex
        job = {
            "id": job_id,
            "signature": signature,
            "status": "queued",
            "message": "Consulta adicionada à fila.",
            "created_at": now_label(),
            "created_epoch": time.time(),
        }
        STATISTICS_JOBS[job_id] = job

    def worker():
        with STATISTICS_JOBS_LOCK:
            current = STATISTICS_JOBS.get(job_id)
            if current:
                current.update({"status": "processing", "message": "Consultando vendas oficiais no Mercado Livre."})
        try:
            result = query_sku_statistics(read_payload(), safe_request)
            with STATISTICS_JOBS_LOCK:
                current = STATISTICS_JOBS.get(job_id)
                if current:
                    current.update(
                        {
                            "status": "completed",
                            "message": "Consulta concluída.",
                            "result": result,
                            "finished_at": now_label(),
                        }
                    )
        except Exception as exc:
            with STATISTICS_JOBS_LOCK:
                current = STATISTICS_JOBS.get(job_id)
                if current:
                    current.update(
                        {
                            "status": "error",
                            "message": str(exc),
                            "finished_at": now_label(),
                        }
                    )

    threading.Thread(target=worker, daemon=True).start()
    return {key: value for key, value in job.items() if key not in {"result", "signature", "created_epoch"}}


def statistics_job_result(job_id, include_result=True):
    cleanup_statistics_jobs()
    with STATISTICS_JOBS_LOCK:
        job = STATISTICS_JOBS.get(str(job_id or ""))
        if not job:
            return None
        public = {key: value for key, value in job.items() if key not in {"signature", "created_epoch"}}
        if not include_result:
            public.pop("result", None)
        return json.loads(json.dumps(public, ensure_ascii=False))


def cleanup_report_jobs(now=None):
    now = now or time.time()
    maximum_age = max(900, int(os.getenv("MELI_REPORT_JOB_TTL_SECONDS", "21600")))
    with REPORT_JOBS_LOCK:
        expired = [
            job_id
            for job_id, job in REPORT_JOBS.items()
            if now - float(job.get("created_epoch") or now) > maximum_age
        ]
        for job_id in expired:
            REPORT_JOBS.pop(job_id, None)


def report_job_result(job_id, include_body=False):
    cleanup_report_jobs()
    with REPORT_JOBS_LOCK:
        job = REPORT_JOBS.get(str(job_id or ""))
        if not job:
            return None
        public = {key: value for key, value in job.items() if key not in {"body", "created_epoch"}}
        if include_body:
            public["body"] = job.get("body")
        return public


def start_report_job(request):
    report_type = str((request or {}).get("report_type") or "").lower()
    output_format = str((request or {}).get("format") or "xlsx").lower()
    if report_type not in {"statistics", "catalog", "ads", "equalization"}:
        raise RuntimeError("Selecione Estatísticas, Catálogo, Anúncios ou Equalização para exportar.")
    if output_format not in {"xlsx", "pdf"}:
        raise RuntimeError("Formato de relatório inválido.")
    statistics_job_id = str((request or {}).get("statistics_job_id") or "")
    if report_type == "statistics" and statistics_job_id:
        statistics_job = statistics_job_result(statistics_job_id)
        if not statistics_job or statistics_job.get("status") != "completed" or not statistics_job.get("result"):
            raise RuntimeError("A consulta de estatísticas ainda não foi concluída. Aguarde e tente exportar novamente.")

    cleanup_report_jobs()
    job_id = uuid.uuid4().hex
    job = {
        "id": job_id,
        "status": "queued",
        "message": "Relatório adicionado à fila.",
        "report_type": report_type,
        "format": output_format,
        "created_at": now_label(),
        "created_epoch": time.time(),
    }
    with REPORT_JOBS_LOCK:
        REPORT_JOBS[job_id] = job

    filters = json.loads(json.dumps((request or {}).get("filters") or {}, ensure_ascii=False))

    def worker():
        try:
            with REPORT_JOBS_LOCK:
                REPORT_JOBS[job_id].update({"status": "processing", "message": "Preparando os dados do relatório."})
            statistics_result = None
            if report_type == "statistics" and statistics_job_id:
                statistics_result = statistics_job_result(statistics_job_id).get("result")
            title, columns, rows, metadata = report_dataset(
                read_payload(include_catalog=report_type in {"catalog", "ads", "equalization"}),
                report_type,
                filters,
                statistics_result,
            )
            maximum_rows = max(10000, int(os.getenv("MELI_REPORT_MAX_ROWS", "100000")))
            if len(rows) > maximum_rows:
                raise RuntimeError(
                    f"O relatório possui {len(rows)} linhas e excede o limite configurado de {maximum_rows}. Divida-o usando os filtros."
                )
            with REPORT_JOBS_LOCK:
                REPORT_JOBS[job_id].update(
                    {"message": f"Gerando {output_format.upper()} com {len(rows)} linha(s).", "row_count": len(rows)}
                )
            stamp = datetime.now(APP_TZ).strftime("%Y%m%d-%H%M")
            if output_format == "xlsx":
                body = build_report_xlsx(title, columns, rows, metadata)
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                body = build_report_pdf(title, columns, rows, metadata)
                content_type = "application/pdf"
            filename = f"competidor-{report_type}-{stamp}.{output_format}"
            with REPORT_JOBS_LOCK:
                REPORT_JOBS[job_id].update(
                    {
                        "status": "completed",
                        "message": "Relatório concluído e pronto para baixar.",
                        "body": body,
                        "content_type": content_type,
                        "filename": filename,
                        "size": len(body),
                        "finished_at": now_label(),
                    }
                )
        except Exception as exc:
            with REPORT_JOBS_LOCK:
                REPORT_JOBS[job_id].update(
                    {"status": "error", "message": str(exc), "finished_at": now_label()}
                )

    threading.Thread(target=worker, daemon=True).start()
    return report_job_result(job_id)


def report_filtered_catalog(payload, report_type, filters):
    account_filter = str(filters.get("account") or "all")
    search = normalized_attribute_label(filters.get("search") or filters.get("product") or "")
    brand_search = normalized_attribute_label(filters.get("brand") or "")
    sku_search = normalized_attribute_label(filters.get("sku") or "")
    code_search = normalized_attribute_label(filters.get("code") or "")
    status_filter = str(filters.get("status") or "all")
    ml_status_filter = str(filters.get("ml_status") or "all")
    listing_type_filter = str(filters.get("listing_type") or "all")
    stock_filter = str(filters.get("stock") or "all")
    catalog_filter = str(filters.get("catalog") or "all")
    flex_filter = str(filters.get("flex") or "all")
    rows = []
    for item in payload.get("catalog") or []:
        if report_type == "catalog" and not is_catalog_listing(item):
            continue
        if account_filter != "all" and account_filter not in {
            str(item.get("account_id") or ""),
            str(item.get("account") or ""),
        }:
            continue
        if status_filter != "all":
            if status_filter == "internal" and not item.get("internal_competition"):
                continue
            if status_filter != "internal" and item.get("status") != status_filter:
                continue
        if ml_status_filter != "all" and item.get("meli_status") != ml_status_filter:
            continue
        if listing_type_filter != "all" and item.get("listing_type_id") != listing_type_filter:
            continue
        if stock_filter == "zero" and int(float(item.get("stock") or 0)) != 0:
            continue
        if stock_filter == "available" and int(float(item.get("stock") or 0)) <= 0:
            continue
        if catalog_filter == "catalog" and not is_catalog_listing(item):
            continue
        if catalog_filter == "traditional" and is_catalog_listing(item):
            continue
        is_flex = item.get("shipping_logistic_type") == "self_service"
        if flex_filter == "active" and not is_flex:
            continue
        if flex_filter == "inactive" and is_flex:
            continue
        haystack = normalized_attribute_label(f"{item.get('title', '')} {item.get('sku', '')} {item.get('id', '')}")
        if search and search not in haystack:
            continue
        if brand_search and brand_search not in normalized_attribute_label(f"{item.get('brand', '')} {item.get('title', '')}"):
            continue
        if sku_search and sku_search not in normalized_attribute_label(item.get("sku") or ""):
            continue
        if code_search and code_search not in normalized_attribute_label(item.get("id") or ""):
            continue
        rows.append(item)
    return rows


def report_dataset(payload, report_type, filters, statistics_result=None):
    if report_type == "statistics":
        result = statistics_result or query_sku_statistics(payload, filters)
        columns = [
            ("rank", "Posição", "integer"),
            ("sku", "SKU", "text"),
            ("product", "Produto", "text"),
            ("accounts_label", "Contas", "text"),
            ("units", "Unidades", "integer"),
            ("orders", "Pedidos", "integer"),
            ("flex_units", "Flex confirmado", "integer"),
            ("unknown_units", "Modalidade pendente", "integer"),
            ("revenue", "Valor dos itens", "currency"),
        ]
        rows = []
        for index, row in enumerate(result.get("rows") or [], 1):
            rows.append({**row, "rank": index, "accounts_label": ", ".join(row.get("accounts") or [])})
        title = "Estatísticas de vendas por SKU"
        metadata = {
            "Período": f"{result.get('date_from')} a {result.get('date_to')}",
            "Conta": filters.get("account") or "Todas",
            "Modalidade": filters.get("flex") or "Todas",
            "SKU": filters.get("sku") or "Todos",
        }
        return title, columns, rows, metadata

    if report_type == "equalization":
        accounts = [str(account.get("nickname") or "") for account in payload.get("accounts") or [] if account.get("official")]
        ml_status_filter = str(filters.get("ml_status") or "all").lower()
        matrix = {}
        for item in payload.get("catalog") or []:
            if ml_status_filter != "all" and str(item.get("meli_status") or "").lower() != ml_status_filter:
                continue
            sku = str(item.get("sku") or "").strip().upper()
            account_name = str(item.get("account") or "")
            if not sku or sku == "-" or account_name not in accounts:
                continue
            key = (account_name, sku)
            row = matrix.setdefault(key, {"account": account_name, "sku": sku, "product": item.get("title") or "", "classic": False, "premium": False})
            row["classic"] = row["classic"] or item.get("listing_type_id") == "gold_special"
            row["premium"] = row["premium"] or item.get("listing_type_id") == "gold_pro"
        report_mode = str(filters.get("report_mode") or "listing_type_gap")
        account_filter = str(filters.get("account") or "all")
        search = normalized_attribute_label(filters.get("search") or "")
        output = []
        if report_mode == "listing_type_gap":
            for row in matrix.values():
                if row["classic"] == row["premium"]:
                    continue
                if account_filter != "all" and row["account"] != account_filter:
                    continue
                output.append({
                    **row,
                    "present": "Clássico" if row["classic"] else "Premium",
                    "missing": "Premium" if row["classic"] else "Clássico",
                })
            columns = [
                ("sku", "SKU", "text"), ("product", "Produto", "text"),
                ("account", "Conta", "text"), ("present", "Possui", "text"),
                ("missing", "Falta criar", "text"),
            ]
            title = "Equalização de anúncios Clássicos e Premium"
        else:
            by_sku = {}
            for row in matrix.values():
                summary = by_sku.setdefault(row["sku"], {"sku": row["sku"], "product": row["product"], "present_accounts": []})
                if row["account"] not in summary["present_accounts"]:
                    summary["present_accounts"].append(row["account"])
            for row in by_sku.values():
                missing = [account for account in accounts if account not in row["present_accounts"]]
                if not missing or (account_filter != "all" and account_filter not in row["present_accounts"]):
                    continue
                output.append({
                    **row,
                    "present_label": ", ".join(row["present_accounts"]),
                    "missing_label": ", ".join(missing),
                })
            columns = [
                ("sku", "SKU", "text"), ("product", "Produto", "text"),
                ("present_label", "Presente em", "text"), ("missing_label", "Ausente em", "text"),
            ]
            title = "Equalização de SKUs entre contas"
        if search:
            output = [row for row in output if search in normalized_attribute_label(" ".join(str(value) for value in row.values()))]
        output.sort(key=lambda row: str(row.get("sku") or ""))
        metadata = {
            "Conta de referência": account_filter if account_filter != "all" else "Todas",
            "Status do anúncio": {"active": "Ativos", "paused": "Pausados"}.get(ml_status_filter, "Todos"),
            "Busca": filters.get("search") or "Todos",
            "Contas comparadas": len(accounts),
            "Gerado em": now_label(),
        }
        return title, columns, output, metadata

    reclassify_internal_competition(payload)
    rows = report_filtered_catalog(payload, report_type, filters)
    common = [
        ("account", "Conta", "text"),
        ("id", "ID do anúncio", "text"),
        ("sku", "SKU", "text"),
        ("title", "Produto", "text"),
        ("listing_type_label", "Tipo", "text"),
        ("catalog_mode", "Modalidade", "text"),
        ("meli_status_label", "Status ML", "text"),
        ("stock", "Estoque", "integer"),
        ("price", "Preço", "currency"),
        ("shipping_cost", "Frete estimado ML", "currency"),
        ("shipping_cost_status_label", "Situação do frete", "text"),
        ("shipping_cost_updated_at", "Frete consultado em", "text"),
    ]
    if report_type == "catalog":
        columns = [
            *common,
            ("catalog_product_id", "ID do catálogo", "text"),
            ("status_label", "Situação", "text"),
            ("winner_name", "Vencedor", "text"),
            ("winner_price", "Preço vencedor", "currency"),
            ("price_to_win", "Preço para ganhar", "currency"),
            ("internal_label", "Entre contas conectadas", "text"),
            ("competition_checked_at", "Verificado em", "text"),
        ]
        title = "Disputa de catálogo"
    else:
        columns = [
            *common,
            ("gtin", "EAN / UPC / GTIN", "text"),
            ("flex_label", "Mercado Envios Flex", "text"),
            ("package_weight", "Peso", "text"),
            ("package_height", "Altura", "text"),
            ("package_width", "Largura", "text"),
            ("package_length", "Comprimento", "text"),
        ]
        title = "Anúncios Mercado Livre"
    normalized_rows = []
    for item in rows:
        normalized_rows.append(
            {
                **item,
                "listing_type_label": "Premium" if item.get("listing_type_id") == "gold_pro" else "Clássico" if item.get("listing_type_id") == "gold_special" else item.get("listing_type_id") or "-",
                "catalog_mode": "Catálogo" if is_catalog_listing(item) else "Tradicional",
                "meli_status_label": {"active": "Ativo", "paused": "Pausado", "under_review": "Aguardando revisão"}.get(item.get("meli_status"), item.get("meli_status") or "-"),
                "status_label": "Entre suas contas" if item.get("internal_competition") else {"winning": "Ganhando", "losing": "Perdendo", "sharing": "Compartilhando", "paused": "Pausado"}.get(item.get("status"), item.get("status") or "-"),
                "internal_label": "Sim" if item.get("internal_competition") else "Não",
                "flex_label": "Ativo" if item.get("shipping_logistic_type") == "self_service" else "Inativo",
                "shipping_cost_status_label": {
                    "ok": "Cotação oficial",
                    "not_available": "Não informado pela API",
                    "error": "Erro na consulta",
                }.get(item.get("shipping_cost_status"), "Aguardando cotação"),
            }
        )
    metadata = {
        "Conta": filters.get("account") or "Todas",
        "Busca": filters.get("search") or filters.get("product") or "Todos",
        "Status": filters.get("status") or filters.get("ml_status") or "Todos",
        "Gerado em": now_label(),
    }
    return title, columns, normalized_rows, metadata


def build_report_xlsx(title, columns, rows, metadata):
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet()
    sheet.title = "RELATORIO"
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="FFD600")
    header = []
    widths = [len(label) for _, label, _ in columns]
    for _, label, _ in columns:
        cell = WriteOnlyCell(sheet, value=label)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(vertical="center")
        header.append(cell)
    sheet.append(header)
    for row in rows:
        cells = []
        for column_index, (key, _, kind) in enumerate(columns):
            value = row.get(key, "") if row.get(key, "") is not None else ""
            cell = WriteOnlyCell(sheet, value=value)
            if kind == "currency" and cell.value not in (None, ""):
                cell.number_format = 'R$ #,##0.00'
            elif kind == "integer" and cell.value not in (None, ""):
                cell.number_format = '#,##0'
            widths[column_index] = max(widths[column_index], len(str(value or "")))
            cells.append(cell)
        sheet.append(cells)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    for column_index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 12), 46)
    info = workbook.create_sheet("FILTROS")
    title_cell = WriteOnlyCell(info, value=title)
    title_cell.font = Font(size=16, bold=True)
    info.append([title_cell])
    info.append(["Filtro", "Valor"])
    for key, value in metadata.items():
        info.append([key, value])
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 65
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_report_pdf(title, columns, rows, metadata):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A3, A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle

    stream = BytesIO()
    page_size = landscape(A3 if len(columns) > 14 else A4)
    document = SimpleDocTemplate(
        stream,
        pagesize=page_size,
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=9 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("report-cell", parent=styles["BodyText"], fontName="Helvetica", fontSize=6.2, leading=7.3, alignment=TA_LEFT)
    header_style = ParagraphStyle("report-header", parent=cell_style, fontName="Helvetica-Bold", textColor=colors.HexColor("#111827"))
    story = [Paragraph(html.escape(title), styles["Title"])]
    story.append(Paragraph(" · ".join(f"<b>{html.escape(str(key))}:</b> {html.escape(str(value))}" for key, value in metadata.items()), styles["BodyText"]))
    story.append(Spacer(1, 5 * mm))
    available_width = page_size[0] - 18 * mm
    weights = []
    for key, _, kind in columns:
        weights.append(2.4 if key in {"title", "product"} else 1.7 if key in {"winner_name", "accounts_label"} else 1.0)
    total_weight = sum(weights)
    widths = [available_width * weight / total_weight for weight in weights]
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD600")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
    )
    chunk_size = max(100, int(os.getenv("MELI_REPORT_PDF_CHUNK_ROWS", "1000")))
    header = [Paragraph(html.escape(label), header_style) for _, label, _ in columns]
    for chunk_start in range(0, len(rows), chunk_size):
        table_rows = [header]
        for row in rows[chunk_start:chunk_start + chunk_size]:
            values = []
            for key, _, kind in columns:
                value = row.get(key, "")
                if kind == "currency" and value not in (None, ""):
                    value = brl(value)
                values.append(Paragraph(html.escape(str(value if value not in (None, "") else "-")), cell_style))
            table_rows.append(values)
        table = LongTable(table_rows, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(table_style)
        story.append(table)
        if chunk_start + chunk_size < len(rows):
            story.append(PageBreak())
    document.build(story)
    return stream.getvalue()


BULK_SHEET_HEADERS = [
    "AÇÃO",
    "CONTA",
    "ID_ANÚNCIO",
    "SKU",
    "TÍTULO",
    "PREÇO",
    "ESTOQUE",
    "STATUS",
    "EAN_UPC_GTIN",
    "ME_FLEX",
    "PESO_KG",
    "ALTURA_CM",
    "LARGURA_CM",
    "COMPRIMENTO_CM",
]


def spreadsheet_number(value, field_name):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError as exc:
        raise RuntimeError(f"{field_name}: informe um número válido.") from exc


def measure_sheet_value(value):
    if clean_attribute_value(value) in ("", "-"):
        return ""
    parsed = parse_decimal_number(value)
    return parsed if parsed else ""


def build_bulk_spreadsheet(payload, filters):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Protection
    from openpyxl.worksheet.datavalidation import DataValidation

    rows = report_filtered_catalog(payload, "ads", filters)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ANUNCIOS"
    sheet.append(BULK_SHEET_HEADERS)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="FFD600")
        cell.font = Font(bold=True, color="111827")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.protection = Protection(locked=True)
    for item in rows:
        sheet.append(
            [
                "",
                item.get("account") or "",
                item.get("id") or "",
                item.get("sku") or "",
                item.get("title") or "",
                item.get("price") or 0,
                item.get("stock") or 0,
                {"active": "ATIVO", "paused": "PAUSADO", "under_review": "EM_REVISÃO"}.get(item.get("meli_status"), item.get("meli_status") or ""),
                item.get("gtin") or "",
                "ATIVO" if item.get("shipping_logistic_type") == "self_service" else "INATIVO",
                measure_sheet_value(item.get("package_weight")),
                measure_sheet_value(item.get("package_height")),
                measure_sheet_value(item.get("package_width")),
                measure_sheet_value(item.get("package_length")),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 22
    sheet.column_dimensions["E"].width = 58
    for letter in "FGHIJKLMN":
        sheet.column_dimensions[letter].width = 18
    for row_index in range(2, sheet.max_row + 1):
        sheet.cell(row_index, 3).number_format = "@"
        sheet.cell(row_index, 4).number_format = "@"
        sheet.cell(row_index, 9).number_format = "@"
        sheet.cell(row_index, 6).number_format = 'R$ #,##0.00'
        sheet.cell(row_index, 7).number_format = '#,##0'
    action_validation = DataValidation(type="list", formula1='"ATUALIZAR"', allow_blank=True)
    status_validation = DataValidation(type="list", formula1='"ATIVO,PAUSADO"', allow_blank=True)
    flex_validation = DataValidation(type="list", formula1='"ATIVO,INATIVO"', allow_blank=True)
    sheet.add_data_validation(action_validation)
    sheet.add_data_validation(status_validation)
    sheet.add_data_validation(flex_validation)
    if sheet.max_row >= 2:
        action_validation.add(f"A2:A{sheet.max_row}")
        status_validation.add(f"H2:H{sheet.max_row}")
        flex_validation.add(f"J2:J{sheet.max_row}")
    instructions = workbook.create_sheet("INSTRUCOES")
    instructions.append(["Edição em massa CompeTIDOR / Mercado Livre"])
    instructions["A1"].font = Font(size=16, bold=True)
    instructions.append(["1", "Edite somente os valores necessários e escreva ATUALIZAR na coluna A da linha."])
    instructions.append(["2", "Não altere nomes, ordem ou quantidade das colunas."])
    instructions.append(["3", "Preço usa reais; medidas usam kg e cm. Ponto e vírgula decimal são aceitos."])
    instructions.append(["4", "Para anúncio com variações, separe os códigos EAN/UPC/GTIN por vírgula, na ordem das variações."])
    instructions.append(["5", "Importe a planilha no CompeTIDOR, confira a prévia e só então aplique as alterações."])
    instructions.column_dimensions["A"].width = 8
    instructions.column_dimensions["B"].width = 110
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def parse_bulk_spreadsheet(payload, encoded_file, actor):
    from openpyxl import load_workbook

    try:
        raw = base64.b64decode(encoded_file, validate=True)
    except Exception as exc:
        raise RuntimeError("O arquivo enviado não é uma planilha válida.") from exc
    if len(raw) > 20 * 1024 * 1024:
        raise RuntimeError("A planilha excede o limite de 20 MB.")
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook["ANUNCIOS"]
    except Exception as exc:
        raise RuntimeError("Use a planilha XLSX gerada pelo CompeTIDOR e preserve a aba ANUNCIOS.") from exc
    headers = [clean_attribute_value(cell.value) for cell in sheet[1]]
    if headers != BULK_SHEET_HEADERS:
        raise RuntimeError("Os cabeçalhos ou a ordem das colunas foram alterados. Baixe uma nova planilha e mantenha a primeira linha intacta.")
    catalog_by_id = {str(item.get("id")): item for item in payload.get("catalog") or [] if item.get("id")}
    accounts = payload.get("accounts") or []
    changes = []
    errors = []
    maximum_rows = max(1, int(os.getenv("MELI_SPREADSHEET_MAX_ROWS", "10000")))
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if row_index > maximum_rows + 1:
            errors.append({"row": row_index, "error": f"A planilha ultrapassou o limite de {maximum_rows} anúncios."})
            break
        row = dict(zip(BULK_SHEET_HEADERS, values))
        if normalized_attribute_label(row.get("AÇÃO")) != "atualizar":
            continue
        item_id = clean_attribute_value(row.get("ID_ANÚNCIO"))
        item = catalog_by_id.get(item_id)
        if not item:
            errors.append({"row": row_index, "item_id": item_id, "error": "Anúncio não encontrado entre as contas conectadas."})
            continue
        account = next(
            (
                account
                for account in accounts
                if account.get("id") == item.get("account_id")
                or normalized_account_name(account.get("nickname")) == normalized_account_name(row.get("CONTA"))
            ),
            None,
        )
        if not account or not account.get("official") or not account.get("access_token"):
            errors.append({"row": row_index, "item_id": item_id, "error": "Conta oficial do anúncio não está conectada."})
            continue
        try:
            change = {
                "item_id": item_id,
                "account_id": account.get("id"),
                "account": account.get("nickname"),
                "row": row_index,
            }
            price = spreadsheet_number(row.get("PREÇO"), "Preço")
            stock = spreadsheet_number(row.get("ESTOQUE"), "Estoque")
            if price is not None and abs(price - float(item.get("price") or 0)) > 0.0001:
                change["price"] = price
            if stock is not None and int(stock) != int(float(item.get("stock") or 0)):
                change["available_quantity"] = int(stock)
            title = clean_attribute_value(row.get("TÍTULO"))
            if title and title != clean_attribute_value(item.get("title")):
                change["title"] = title[:60]
            status = normalized_attribute_label(row.get("STATUS"))
            status_value = {"ativo": "active", "pausado": "paused"}.get(status)
            if status_value and status_value != item.get("meli_status"):
                change["status_action"] = "activate" if status_value == "active" else "pause"
            gtin = clean_attribute_value(row.get("EAN_UPC_GTIN"))
            if gtin and gtin != clean_attribute_value(item.get("gtin")):
                normalize_clone_identifier_codes(gtin)
                change["gtin"] = gtin
            flex = normalized_attribute_label(row.get("ME_FLEX"))
            current_flex = item.get("shipping_logistic_type") == "self_service"
            if flex in {"ativo", "inativo"} and (flex == "ativo") != current_flex:
                change["flex_action"] = "activate" if flex == "ativo" else "remove"
            for sheet_key, request_key in (
                ("PESO_KG", "package_weight"),
                ("ALTURA_CM", "package_height"),
                ("LARGURA_CM", "package_width"),
                ("COMPRIMENTO_CM", "package_length"),
            ):
                value = spreadsheet_number(row.get(sheet_key), sheet_key)
                if value is None:
                    continue
                unit = "kg" if request_key == "package_weight" else "cm"
                current = parse_decimal_number(item.get(request_key))
                if abs(value - current) > 0.0001:
                    change[request_key] = f"{value:g} {unit}"
            editable = {key: value for key, value in change.items() if key not in {"item_id", "account_id", "account", "row"}}
            if editable:
                change["changes"] = editable
                changes.append(change)
        except Exception as exc:
            errors.append({"row": row_index, "item_id": item_id, "error": str(exc)})
    token = secrets.token_urlsafe(24)
    with SPREADSHEET_JOBS_LOCK:
        SPREADSHEET_JOBS[token] = {
            "created": time.monotonic(),
            "user_id": (actor or {}).get("id"),
            "changes": changes,
            "errors": errors,
        }
        expired = [key for key, value in SPREADSHEET_JOBS.items() if time.monotonic() - value.get("created", 0) > 1800]
        for key in expired:
            SPREADSHEET_JOBS.pop(key, None)
    return {"token": token, "changes": changes, "errors": errors, "ready": len(changes), "invalid": len(errors)}


def apply_spreadsheet_change(payload, row, actor):
    item_id = row.get("item_id")
    account = official_account_by_name(payload, row.get("account_id") or row.get("account"))
    item = next((entry for entry in payload.get("catalog") or [] if entry.get("id") == item_id), None)
    if not account or not item:
        raise RuntimeError("Conta ou anúncio não encontrado no momento da aplicação.")
    changes = row.get("changes") or {}
    client = account_client(account)
    update = {}
    for key in ("price", "available_quantity", "title"):
        if key in changes:
            update[key] = changes[key]
    if changes.get("status_action") == "pause":
        update["status"] = "paused"
    elif changes.get("status_action") == "activate":
        update["status"] = "active"
    dimension_map = {
        "package_weight": "SELLER_PACKAGE_WEIGHT",
        "package_height": "SELLER_PACKAGE_HEIGHT",
        "package_width": "SELLER_PACKAGE_WIDTH",
        "package_length": "SELLER_PACKAGE_LENGTH",
    }
    attributes = []
    expected_package = {}
    for request_key, attr_id in dimension_map.items():
        if changes.get(request_key):
            api_value = seller_package_api_value(request_key, changes[request_key])
            expected_package[request_key] = api_value
            attributes.append({"id": attr_id, "value_name": api_value})
    expected_gtin = []
    if changes.get("gtin"):
        fragment, expected_gtin = item_gtin_update_fragment(client, item_id, changes["gtin"])
        attributes.extend(fragment.get("attributes") or [])
        if fragment.get("variations"):
            update["variations"] = fragment["variations"]
    if attributes:
        update["attributes"] = attributes
    if changes.get("flex_action"):
        update["shipping"] = {
            "mode": item.get("shipping_mode") or "me2",
            "logistic_type": "self_service" if changes["flex_action"] == "activate" else "drop_off",
        }
    if not update:
        return {"item_id": item_id, "status": "ignored"}
    client.update_item(item_id, update)
    verified = verify_package_update(client, item_id, expected_package) if expected_package else {}
    if expected_gtin:
        verified = verify_gtin_update(client, item_id, expected_gtin)
    local_changes = {}
    for key, local_key in (("price", "price"), ("available_quantity", "stock"), ("title", "title")):
        if key in update:
            local_changes[local_key] = {"from": item.get(local_key), "to": update[key]}
            item[local_key] = update[key]
    package_values = package_values_from_item(verified) if verified else {}
    for request_key in dimension_map:
        if request_key in changes:
            value = package_values.get(request_key) or changes[request_key]
            local_changes[request_key] = {"from": item.get(request_key), "to": value}
            item[request_key] = value
    if expected_gtin:
        value = ", ".join(expected_gtin)
        local_changes["gtin"] = {"from": item.get("gtin"), "to": value}
        item["gtin"] = value
    if update.get("status"):
        local_changes["status"] = {"from": item.get("meli_status"), "to": update["status"]}
        item["meli_status"] = update["status"]
        item["status"] = "paused" if update["status"] == "paused" else ("sharing" if is_catalog_listing(item) else "winning")
    if update.get("shipping"):
        local_changes["shipping_logistic_type"] = {"from": item.get("shipping_logistic_type"), "to": update["shipping"]["logistic_type"]}
        item["shipping_logistic_type"] = update["shipping"]["logistic_type"]
    item["updated_at"] = now_label()
    append_item_log(payload, item, actor, "Atualização por planilha", local_changes)
    return {"item_id": item_id, "status": "updated", "changes": list(local_changes)}


def execute_clone_job(payload, job, incoming_answers=None):
    incoming_answers = incoming_answers or {}
    if isinstance(incoming_answers, dict) and incoming_answers:
        saved_answers = job.setdefault("field_answers", {})
        for item_id, answers in incoming_answers.items():
            if isinstance(answers, dict):
                saved_answers.setdefault(item_id, {}).update(
                    {key: value for key, value in answers.items() if value not in (None, "")}
                )
    catalog = payload.setdefault("catalog", [])
    existing_ids = {item.get("id") for item in catalog if item.get("id")}
    copied = []
    errors = []
    created_details = []
    edits = job.get("edits") or {}
    for item_id in job.get("item_ids", []):
        source_item = next((item for item in catalog if item.get("id") == item_id), None)
        if not source_item:
            errors.append({"item_id": item_id, "error": "O anúncio de origem não foi encontrado na base sincronizada."})
            continue
        try:
            created, official_source, verified_item = create_official_clone(payload, job, item_id, edits)
        except Exception as exc:
            pending_fields = getattr(exc, "pending_fields", [])
            if pending_fields:
                errors.append(
                    {
                        "item_id": item_id,
                        "error": str(exc),
                        "pending_fields": pending_fields,
                        "original_error": getattr(exc, "original_error", ""),
                    }
                )
            else:
                errors.append({"item_id": item_id, "error": friendly_clone_error(exc)})
            continue
        new_id = created.get("id") or f"COPY-{uuid.uuid4().hex[:6].upper()}"
        if new_id in existing_ids:
            continue
        existing_ids.add(new_id)
        target_account = official_account_by_name(payload, job.get("target_account_id") or job.get("target")) or {}
        official_sku = item_sku(official_source)
        source_sku = official_sku if official_sku and official_sku != "-" else source_item.get("sku") or source_item.get("id") or "SEM-SKU"
        official_created_item = {**official_source, **created, **verified_item}
        new_item = synced_catalog_item(target_account, official_created_item)
        display_sku = item_sku(official_created_item) or f"{source_sku}{edits.get('sku_suffix') or ''}"
        new_item.update(
            {
                "id": new_id,
                "account": job.get("target"),
                "account_id": target_account.get("id"),
                "sku": display_sku,
                "price": official_created_item.get("price") or float(edits.get("price") or source_item.get("price") or 0),
                "stock": item_available_quantity(official_created_item) or int(float(edits.get("stock") or item_available_quantity(official_source) or source_item.get("stock") or 0)),
                "status": "sharing",
                "share": 0,
                "clone_source_item_id": item_id,
                "description_override": edits.get("description", ""),
                "permalink": official_created_item.get("permalink") or created.get("permalink") or "",
                "meli_status": official_created_item.get("status") or created.get("status") or new_item.get("meli_status"),
                "verification_warning": created.get("_verification_warning", ""),
                "action": f"Anúncio criado oficialmente no Mercado Livre a partir de {source_item.get('id')}. Novo código: {new_id}.",
            }
        )
        catalog.append(new_item)
        copied.append(new_item)
        created_details.append(
            {
                "source_item_id": item_id,
                "item_id": new_id,
                "title": new_item.get("title"),
                "sku": new_item.get("sku"),
                "status": new_item.get("meli_status") or "-",
                "permalink": new_item.get("permalink") or "",
                "verification_warning": new_item.get("verification_warning") or "",
            }
        )
    has_pending = any(row.get("pending_fields") for row in errors)
    job["validation_version"] = 3
    job["status"] = "copied" if copied and not errors else "review_required" if has_pending else "partial_error" if copied else "error"
    job["copied_items"] = [item.get("id") for item in copied]
    job["created_details"] = created_details
    job["errors"] = errors
    created_codes = ", ".join(item.get("item_id") for item in created_details if item.get("item_id"))
    job["note"] = f"{len(copied)} anúncio(s) criado(s) oficialmente no Mercado Livre para {job.get('target')}."
    if created_codes:
        job["note"] += f" Códigos criados: {created_codes}."
    if errors:
        job["note"] += f" {len(errors)} anúncio(s) precisam de atenção; veja os detalhes abaixo."
    return copied


def prepare_clone_preview(request):
    payload = read_payload(include_catalog=False)
    item_ids = request.get("item_ids") or []
    edits = request.get("edits") or {}
    if isinstance(item_ids, str):
        item_ids = [item.strip() for item in item_ids.split(",") if item.strip()]
    item_ids = [str(item_id).strip() for item_id in item_ids if str(item_id).strip()]
    source = str(request.get("source") or "").strip()
    target_values = request.get("targets") or [request.get("target") or ""]
    if isinstance(target_values, str):
        target_values = [target_values]
    target_values = list(dict.fromkeys(str(value or "").strip() for value in target_values if str(value or "").strip()))
    variants = request.get("variants") or [{}]
    if not isinstance(variants, list):
        variants = [{}]
    if not source or not target_values:
        raise RuntimeError("Informe conta origem e ao menos uma conta destino.")
    source_account = official_account_by_name(payload, source)
    if not source_account or not source_account.get("official") or not source_account.get("access_token"):
        raise RuntimeError("A conta origem não está conectada oficialmente. Reautentique a conta e tente novamente.")
    target_accounts = []
    for target in target_values:
        target_account = official_account_by_name(payload, target)
        if not target_account or not target_account.get("official") or not target_account.get("access_token"):
            raise RuntimeError(f"A conta destino {target} não está conectada oficialmente. Reautentique-a e tente novamente.")
        target_accounts.append(target_account)
    if not item_ids:
        raise RuntimeError("Selecione ao menos um anúncio específico.")
    if len(item_ids) > 1:
        raise RuntimeError("A cópia permite apenas um anúncio por vez.")
    catalog = read_json(CATALOG_DATA_FILE, [])
    payload["catalog"] = catalog
    payload["_catalog_loaded"] = False
    source_items = [
        item
        for item in catalog
        if (
            item.get("account_id") == source_account.get("id")
            or item.get("account") == source_account.get("nickname")
        )
        and item.get("id") in item_ids
    ]
    found_ids = {item.get("id") for item in source_items}
    missing_ids = [item_id for item_id in item_ids if item_id not in found_ids]
    if missing_ids:
        raise RuntimeError(f"Anúncios não encontrados na conta origem: {', '.join(missing_ids)}.")
    combinations = len(target_accounts) * len(variants)
    if combinations > max(1, int(os.getenv("MELI_CLONE_MAX_COMBINATIONS", "20"))):
        raise RuntimeError("Selecione no máximo 20 combinações de conta e tipo por lote.")
    jobs = payload.get("clone_jobs")
    if not isinstance(jobs, list):
        jobs = []
        payload["clone_jobs"] = jobs
    batch_id = f"batch-{uuid.uuid4().hex[:8]}"
    created_jobs = []
    validation_cache = {}
    for target_account in target_accounts:
        for variant in variants:
            variant = variant if isinstance(variant, dict) else {}
            variant_edits = dict(edits)
            listing_type = str(variant.get("listing_type_id") or "").strip()
            if listing_type:
                variant_edits["listing_type_id"] = listing_type
            if variant.get("price") not in (None, ""):
                variant_edits["price"] = variant.get("price")
            validation_key = json.dumps(variant_edits, sort_keys=True, ensure_ascii=False)
            if validation_key not in validation_cache:
                validation_cache[validation_key] = clone_preflight_pending_fields(
                    payload,
                    source_account.get("id"),
                    item_ids,
                    variant_edits,
                )
            preflight_errors = json.loads(json.dumps(validation_cache[validation_key], ensure_ascii=False))
            type_label = "Premium" if listing_type == "gold_pro" else "Clássico" if listing_type == "gold_special" else "Mesmo tipo do anúncio"
            created_jobs.append(
                {
                    "id": f"clone-{uuid.uuid4().hex[:8]}",
                    "batch_id": batch_id,
                    "validation_version": 3,
                    "source": source_account.get("nickname"),
                    "target": target_account.get("nickname"),
                    "source_account_id": source_account.get("id"),
                    "target_account_id": target_account.get("id"),
                    "item_ids": item_ids,
                    "items": len(item_ids),
                    "status": "review_required" if preflight_errors else "preview_ready",
                    "edits": variant_edits,
                    "variant_label": type_label,
                    "errors": preflight_errors,
                    "note": f"Preview preparado para {target_account.get('nickname')} em {type_label}. Campos opcionais em branco mantêm as informações originais.",
                }
            )
    for job in reversed(created_jobs):
        jobs.insert(0, job)
    write_payload(payload)
    legacy_request = not request.get("targets") and not request.get("variants")
    return created_jobs[0] if legacy_request else {"batch_id": batch_id, "jobs": created_jobs}


def execute_clone_request(request):
    payload = read_payload()
    jobs = payload.get("clone_jobs") if isinstance(payload.get("clone_jobs"), list) else []
    job = next((item for item in jobs if item.get("id") == request.get("job_id")), None)
    if not job:
        raise RuntimeError("Preview não encontrado.")
    copied = execute_clone_job(payload, job, request.get("field_answers") or {})
    write_payload(payload)
    return {"ok": True, "job": job, "copied": copied}


def execute_clone_batch_request(request):
    payload = read_payload()
    jobs = payload.get("clone_jobs") if isinstance(payload.get("clone_jobs"), list) else []
    requested_ids = request.get("job_ids") or []
    batch_id = str(request.get("batch_id") or "")
    selected = [
        job
        for job in jobs
        if (requested_ids and job.get("id") in requested_ids)
        or (batch_id and job.get("batch_id") == batch_id)
    ]
    if not selected:
        raise RuntimeError("Nenhum preview do lote foi encontrado.")
    if len(selected) > max(1, int(os.getenv("MELI_CLONE_MAX_COMBINATIONS", "20"))):
        raise RuntimeError("O lote excede o limite de combinações permitido.")
    copied = []
    for job in selected:
        if job.get("status") == "copied":
            continue
        copied.extend(execute_clone_job(payload, job, (request.get("field_answers") or {}).get(job.get("id"), {})))
    write_payload(payload)
    return {"ok": True, "jobs": selected, "copied": copied}


class App(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        return

    def send_json(self, payload, status=200, headers=None):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_security_headers()
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_bytes(self, body, content_type, filename, status=200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_security_headers()
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_security_headers(self):
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Permissions-Policy", "camera=(), microphone=(), geolocation=(), payment=()")
        self.send_header(
            "Content-Security-Policy",
            "default-src 'self'; "
            "script-src 'self'; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data: https: http:; "
            "connect-src 'self'; "
            "frame-ancestors 'none'; "
            "base-uri 'self'; "
            "form-action 'self'",
        )

    def validate_same_origin(self):
        origin = self.headers.get("Origin")
        if not origin:
            return True
        origin_host = urlparse(origin).netloc
        request_host = self.headers.get("Host", "")
        if origin_host == request_host:
            return True
        self.send_json({"error": "Requisição bloqueada por origem inválida."}, status=403)
        return False

    def cookie_value(self, name):
        cookies = self.headers.get("Cookie", "")
        for chunk in cookies.split(";"):
            if "=" not in chunk:
                continue
            key, value = chunk.strip().split("=", 1)
            if key == name:
                return value
        return ""

    def current_user(self, payload=None):
        token = self.cookie_value("competidor_session")
        if not token:
            return None
        sessions = read_json("sessions.json", [])
        now = int(time.time())
        session = next((item for item in sessions if item.get("token") == token and item.get("expires_at", 0) > now), None)
        if not session:
            return None
        payload = payload or read_payload()
        users = ensure_users(payload)
        return next((public_user(user) for user in users if user.get("id") == session.get("user_id")), None)

    def create_session(self, user):
        token = secrets.token_urlsafe(32)
        now = int(time.time())
        with DATA_LOCK:
            sessions = [item for item in read_json("sessions.json", []) if item.get("expires_at", 0) > now]
            sessions.append({"token": token, "user_id": user["id"], "created_at": now, "expires_at": now + SESSION_SECONDS})
            write_json("sessions.json", sessions[-50:])
        return token

    def clear_session(self):
        token = self.cookie_value("competidor_session")
        if token:
            with DATA_LOCK:
                sessions = [item for item in read_json("sessions.json", []) if item.get("token") != token]
                write_json("sessions.json", sessions)

    def require_auth(self, payload=None):
        user = self.current_user(payload)
        if user:
            return user
        self.send_json({"error": "Login necessário"}, status=401)
        return None

    def redirect_uri(self):
        return meli_config()["redirect_uri"]

    def suggested_redirect_uri(self):
        host = self.headers.get("Host", "localhost:8765")
        hostname = host.split(":")[0]
        return f"https://{hostname}:{https_port()}/api/oauth/callback"

    def serve_file(self, path):
        if path == "/":
            path = "/index.html"
        target = (PUBLIC / path.lstrip("/")).resolve()
        if not str(target).startswith(str(PUBLIC.resolve())) or not target.exists():
            if not path.startswith("/api/"):
                target = PUBLIC / "index.html"
            else:
                self.send_error(404)
                return
        content_types = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "application/javascript; charset=utf-8",
            ".svg": "image/svg+xml",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".webp": "image/webp",
        }
        body = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_types.get(target.suffix, "application/octet-stream"))
        self.send_security_headers()
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        payload = read_payload(include_catalog=parsed.path == "/api/catalog")

        if parsed.path == "/api/auth/me":
            user = self.current_user(payload)
            self.send_json({"authenticated": bool(user), "user": user, "setup_required": len(ensure_users(payload)) == 0})
            return

        if parsed.path == "/api/auth/logout":
            self.clear_session()
            self.send_response(302)
            self.send_header("Set-Cookie", "competidor_session=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")
            self.send_header("Location", "/#/dashboard")
            self.end_headers()
            return

        if parsed.path.startswith("/api/") and parsed.path not in {"/api/oauth/callback"}:
            if not self.require_auth(payload):
                return

        if parsed.path == "/api/meta":
            tenant = payload.get("tenant") or default_tenant()
            actor = self.current_user(payload)
            master_view = is_master(actor)
            self.send_json(
                {
                    "name": "CompeTIDOR",
                    "tenant": {
                        **tenant,
                        "official_accounts": len([account for account in payload.get("accounts", []) if account.get("official")]),
                        "connected_accounts": len([account for account in payload.get("accounts", []) if account.get("status") == "connected"]),
                    },
                    "meli": {
                        "client_configured": app_configured(),
                        "auth_url": MELI_AUTH_URL if master_view else "",
                        "token_url": MELI_TOKEN_URL if master_view else "",
                        "api_url": MELI_API_URL if master_view else "",
                        "redirect_uri": self.redirect_uri() if master_view else "",
                        "suggested_redirect_uri": self.suggested_redirect_uri() if master_view else "",
                        "oauth_issues": oauth_issues() if master_view else [],
                    },
                    "generated_at": int(time.time()),
                }
            )
            return
        if parsed.path == "/api/dashboard":
            notifications_migrated = migrate_legacy_notifications(payload)
            alerts_enriched = enrich_product_alerts(payload)
            if notifications_migrated or alerts_enriched:
                write_payload(payload)
            self.send_json(public_payload(payload, self.current_user(payload), include_catalog=False))
            return
        if parsed.path == "/api/catalog":
            reclassify_internal_competition(payload)
            catalog = payload.get("catalog", [])
            self.send_json(
                {
                    "catalog": catalog,
                    "item_logs": payload.get("item_logs", [])[:500],
                    "catalog_counts": {
                        "total": len(catalog),
                        "winning": len([item for item in catalog if item.get("status") == "winning"]),
                        "losing": len([item for item in catalog if item.get("status") == "losing"]),
                        "sharing": len([item for item in catalog if item.get("status") == "sharing"]),
                        "paused": len([item for item in catalog if item.get("status") == "paused" or item.get("meli_status") == "paused"]),
                    },
                }
            )
            return
        if parsed.path.startswith("/api/async/jobs/"):
            job_id = parsed.path.rsplit("/", 1)[-1]
            job = async_operation_result(job_id)
            if not job:
                self.send_json({"error": "Processamento não encontrado ou expirado."}, status=404)
            else:
                self.send_json({"ok": True, **job})
            return
        if parsed.path.startswith("/api/statistics/jobs/"):
            job_id = parsed.path.rsplit("/", 1)[-1]
            job = statistics_job_result(job_id)
            if not job:
                self.send_json({"error": "Consulta de estatísticas não encontrada ou expirada."}, status=404)
            else:
                self.send_json({"ok": True, **job})
            return
        if parsed.path.startswith("/api/reports/jobs/"):
            job_id = parsed.path.rsplit("/", 1)[-1]
            query = parse_qs(parsed.query)
            download = query.get("download", ["0"])[0] == "1"
            job = report_job_result(job_id, include_body=download)
            if not job:
                self.send_json({"error": "Relatório não encontrado ou expirado."}, status=404)
            elif download:
                if job.get("status") != "completed" or not job.get("body"):
                    self.send_json({"error": job.get("message") or "O relatório ainda não está pronto."}, status=409)
                else:
                    self.send_bytes(job["body"], job["content_type"], job["filename"])
            else:
                self.send_json({"ok": True, **job})
            return
        if parsed.path == "/api/accounts":
            self.send_json([public_account(account) for account in payload["accounts"]])
            return
        if parsed.path == "/api/meli/config":
            user = self.current_user(payload)
            if not is_master(user):
                self.send_json({"error": "Apenas usuário master pode ver ou editar Client ID, Client Secret e Redirect URI."}, status=403)
                return
            config = meli_config()
            self.send_json(
                {
                    "client_id": config["client_id"],
                    "client_secret_set": bool(config["client_secret"]),
                    "redirect_uri": config["redirect_uri"],
                    "suggested_redirect_uri": self.suggested_redirect_uri(),
                    "issues": oauth_issues(),
                }
            )
            return
        if parsed.path == "/api/items":
            query = parse_qs(parsed.query)
            account_name = query.get("account", [""])[0]
            items = payload["catalog"]
            if account_name:
                items = [item for item in items if item["account"] == account_name]
            self.send_json(items)
            return
        if parsed.path == "/api/oauth/login":
            query = parse_qs(parsed.query)
            switch_account = query.get("switch_account", ["0"])[0] == "1"
            issues = oauth_issues()
            if issues:
                self.send_json(
                    {
                        "error": "Corrija a configuração OAuth antes de conectar.",
                        "issues": issues,
                        "hint": "O MELI_REDIRECT_URI precisa ser exatamente igual ao Redirect URI cadastrado na aplicação do Mercado Livre.",
                    },
                    status=400,
                )
                return
            state = str(uuid.uuid4())
            states = read_json("oauth_states.json", [])
            states.append({"state": state, "created_at": int(time.time())})
            write_json("oauth_states.json", states[-20:])
            self.send_json(
                {
                    "authorization_url": MercadoLivreClient().auth_url(state, self.redirect_uri(), switch_account),
                    "state": state,
                    "switch_account": switch_account,
                }
            )
            return
        if parsed.path == "/api/oauth/start":
            query = parse_qs(parsed.query)
            switch_account = query.get("switch_account", ["0"])[0] == "1"
            issues = oauth_issues()
            if issues:
                self.send_response(302)
                self.send_header("Location", "/#/contas?oauth=config")
                self.end_headers()
                return
            state = str(uuid.uuid4())
            states = read_json("oauth_states.json", [])
            states.append({"state": state, "created_at": int(time.time())})
            write_json("oauth_states.json", states[-20:])
            self.send_response(302)
            self.send_header("Location", MercadoLivreClient().auth_url(state, self.redirect_uri(), switch_account))
            self.end_headers()
            return
        if parsed.path == "/api/oauth/callback":
            query = parse_qs(parsed.query)
            code = query.get("code", [""])[0]
            state = query.get("state", [""])[0]
            states = read_json("oauth_states.json", [])
            valid_state = any(item.get("state") == state for item in states)
            if not code or not valid_state:
                self.send_response(302)
                self.send_header("Location", "/#/contas?oauth=erro")
                self.end_headers()
                return
            try:
                token = MercadoLivreClient().exchange_code(code, self.redirect_uri())
                client = MercadoLivreClient(token.get("access_token"))
                me = client.me()
                existing_account = next(
                    (item for item in payload.get("accounts", []) if str(item.get("seller_id")) == str(me.get("id"))),
                    None,
                )
                account = {
                    "id": f"meli-{me.get('id')}",
                    "tenant_id": (payload.get("tenant") or default_tenant())["id"],
                    "nickname": me.get("nickname") or me.get("first_name") or "Conta Mercado Livre",
                    "seller_id": str(me.get("id")),
                    "site_id": me.get("site_id") or "MLB",
                    "status": "connected",
                    "official": True,
                    "color": "#2563eb",
                    "last_sync": now_label(),
                    "access_token": token.get("access_token"),
                    "refresh_token": token.get("refresh_token"),
                    "expires_in": token.get("expires_in"),
                    "token_created_at": int(time.time()),
                    "permalink": me.get("permalink", ""),
                }
                add_or_update_account(payload, account)
                oauth_status = "updated" if existing_account else "connected"
            except Exception as exc:
                oauth_status = "error"
                add_or_update_account(
                    payload,
                    {
                        "id": f"meli-{state[:8]}",
                        "nickname": "Conta Mercado Livre com erro OAuth",
                        "seller_id": "Não sincronizado",
                        "site_id": "MLB",
                        "status": "oauth_error",
                        "official": False,
                        "color": "#dc2626",
                        "last_sync": now_label(),
                        "error": str(exc),
                    },
                )
            write_payload(payload)
            self.send_response(302)
            self.send_header("Location", f"/#/contas?oauth={oauth_status}")
            self.end_headers()
            return
        if parsed.path == "/api/notifications/meli":
            self.send_json({"ok": True})
            return
        if not parsed.path.startswith("/api/") and "." not in Path(parsed.path).name:
            self.serve_file("/")
            return
        self.serve_file(parsed.path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/notifications/meli":
            try:
                length = int(self.headers.get("Content-Length", 0))
                body = self.rfile.read(length).decode("utf-8") if length else "{}"
                event = json.loads(body or "{}")
                configured_app_id = str(meli_config().get("client_id") or "")
                event_app_id = str(event.get("application_id") or "")
                if configured_app_id and event_app_id and configured_app_id != event_app_id:
                    self.send_json({"ok": False, "error": "Aplicação de origem inválida."}, status=403)
                    return
                MELI_NOTIFICATION_QUEUE.put_nowait(event)
                self.send_json({"ok": True}, status=200)
            except queue.Full:
                self.send_json({"ok": False, "error": "Fila temporariamente cheia."}, status=503)
            except Exception:
                self.send_json({"ok": True}, status=200)
            return
        if not self.validate_same_origin():
            return
        length = int(self.headers.get("Content-Length", 0))
        maximum_body = max(1024 * 1024, int(os.getenv("MAX_REQUEST_BODY_BYTES", str(30 * 1024 * 1024))))
        if length > maximum_body:
            self.send_json({"error": "O arquivo ou a requisição excede o limite permitido pelo servidor."}, status=413)
            return
        body = self.rfile.read(length).decode("utf-8") if length else "{}"
        request = json.loads(body or "{}")
        lightweight_paths = {
            "/api/auth/setup-master",
            "/api/auth/login",
            "/api/auth/logout",
            "/api/users",
            "/api/users/update",
            "/api/meli/config",
            "/api/notifications/config",
            "/api/notifications/test",
            "/api/meli/item/clone-source",
            "/api/meli/item/description",
            "/api/clone/preview",
            "/api/clone/execute",
            "/api/clone/execute-batch",
        }
        payload = read_payload(include_catalog=parsed.path not in lightweight_paths)

        if parsed.path == "/api/auth/setup-master":
            users = ensure_users(payload)
            if users:
                self.send_json({"error": "O usuário master inicial já foi criado."}, status=409)
                return
            name = (request.get("name") or "").strip()
            email = (request.get("email") or "").strip().lower()
            password = request.get("password") or ""
            if not name or not email or not password:
                self.send_json({"error": "Informe nome, e-mail e senha para criar o master."}, status=400)
                return
            if len(password) < 8:
                self.send_json({"error": "A senha do master precisa ter pelo menos 8 caracteres."}, status=400)
                return
            salt, password_hash = hash_password(password)
            user = {
                "id": f"user-{uuid.uuid4().hex[:8]}",
                "tenant_id": (payload.get("tenant") or default_tenant())["id"],
                "name": name,
                "email": email,
                "role": "master",
                "status": "ativo",
                "password_salt": salt,
                "password_hash": password_hash,
                "created_at": now_label(),
            }
            payload["users"] = [user]
            payload.setdefault("user_notifications", {})[user["id"]] = blank_notifications()
            write_payload(payload)
            token = self.create_session(user)
            self.send_json(
                {"ok": True, "user": public_user(user)},
                status=201,
                headers={"Set-Cookie": f"competidor_session={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_SECONDS}"},
            )
            return

        if parsed.path == "/api/auth/login":
            users = ensure_users(payload)
            write_payload(payload)
            if not users:
                self.send_json({"error": "Crie o usuário master no primeiro acesso.", "setup_required": True}, status=409)
                return
            email = (request.get("email") or "").strip().lower()
            password = request.get("password") or ""
            user = next((item for item in users if (item.get("email") or "").lower() == email), None)
            if not user or not verify_password(password, user) or user.get("status") not in ("ativo", "active"):
                self.send_json({"error": "E-mail ou senha inválidos."}, status=401)
                return
            token = self.create_session(user)
            self.send_json(
                {"ok": True, "user": public_user(user)},
                headers={"Set-Cookie": f"competidor_session={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_SECONDS}"},
            )
            return

        if parsed.path == "/api/auth/logout":
            self.clear_session()
            self.send_json(
                {"ok": True},
                headers={"Set-Cookie": "competidor_session=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0"},
            )
            return

        if parsed.path.startswith("/api/") and not parsed.path.startswith("/api/auth/"):
            if not self.require_auth(payload):
                return

        if parsed.path == "/api/alerts/read":
            alert_id = request.get("id")
            for alert in payload["alerts"]:
                if alert["id"] == alert_id:
                    alert["read"] = True
            write_payload(payload)
            self.send_json({"ok": True, "alerts": payload["alerts"]})
            return

        if parsed.path == "/api/users":
            actor = self.current_user(payload)
            if not can_manage_users(actor):
                self.send_json({"error": "Seu usuário não tem permissão para criar usuários."}, status=403)
                return
            name = (request.get("name") or "").strip()
            email = (request.get("email") or "").strip().lower()
            role = (request.get("role") or "operator").strip()
            password = request.get("password") or ""
            if not name or not email or not password:
                self.send_json({"error": "Informe nome, e-mail e senha para criar o usuário."}, status=400)
                return
            if len(password) < 6:
                self.send_json({"error": "A senha precisa ter pelo menos 6 caracteres."}, status=400)
                return
            if role == "master" and not is_master(actor):
                self.send_json({"error": "Apenas o usuário master pode criar outro master."}, status=403)
                return
            users = payload.get("users") or default_users(payload.get("tenant") or default_tenant())
            if any((user.get("email") or "").lower() == email for user in users):
                self.send_json({"error": "Já existe um usuário com este e-mail."}, status=400)
                return
            salt, password_hash = hash_password(password)
            user = {
                "id": f"user-{uuid.uuid4().hex[:8]}",
                "tenant_id": (payload.get("tenant") or default_tenant())["id"],
                "name": name,
                "email": email,
                "role": role if role in {"master", "admin", "manager", "operator", "viewer"} else "operator",
                "status": "ativo",
                "password_salt": salt,
                "password_hash": password_hash,
                "created_at": now_label(),
            }
            users.append(user)
            payload["users"] = users
            write_payload(payload)
            self.send_json({"ok": True, "user": public_user(user), "users": visible_users_for(actor, users)}, status=201)
            return

        if parsed.path == "/api/users/update":
            actor = self.current_user(payload)
            if not can_manage_users(actor):
                self.send_json({"error": "Seu usuário não tem permissão para editar usuários."}, status=403)
                return
            user_id = request.get("id")
            users = ensure_users(payload)
            user = next((item for item in users if item.get("id") == user_id), None)
            if not user:
                self.send_json({"error": "Usuário não encontrado."}, status=404)
                return
            if user.get("role") == "master" and not is_master(actor):
                self.send_json({"error": "Administradores não podem ver nem editar usuários master."}, status=403)
                return
            name = (request.get("name") or "").strip()
            email = (request.get("email") or "").strip().lower()
            role = (request.get("role") or user.get("role") or "operator").strip()
            status = (request.get("status") or user.get("status") or "ativo").strip()
            password = request.get("password") or ""
            if not name or not email:
                self.send_json({"error": "Informe nome e e-mail para editar o usuário."}, status=400)
                return
            if any(item.get("id") != user_id and (item.get("email") or "").lower() == email for item in users):
                self.send_json({"error": "Já existe outro usuário com este e-mail."}, status=400)
                return
            if password and len(password) < 6:
                self.send_json({"error": "A nova senha precisa ter pelo menos 6 caracteres."}, status=400)
                return
            if role == "master" and not is_master(actor):
                self.send_json({"error": "Apenas o usuário master pode conceder perfil master."}, status=403)
                return
            current_masters = [item for item in users if item.get("role") == "master" and item.get("status") in {"ativo", "active"}]
            if user.get("role") == "master" and (role != "master" or status in {"inativo", "inactive"}) and len(current_masters) <= 1:
                self.send_json({"error": "Mantenha pelo menos um usuário master ativo."}, status=400)
                return
            user["name"] = name
            user["email"] = email
            user["role"] = role if role in {"master", "admin", "manager", "operator", "viewer"} else "operator"
            user["status"] = status if status in {"ativo", "inativo", "active", "inactive"} else "ativo"
            if password:
                salt, password_hash = hash_password(password)
                user["password_salt"] = salt
                user["password_hash"] = password_hash
            user["updated_at"] = now_label()
            payload["users"] = users
            write_payload(payload)
            self.send_json({"ok": True, "user": public_user(user), "users": visible_users_for(actor, users)})
            return

        if parsed.path == "/api/meli/config":
            actor = self.current_user(payload)
            if not is_master(actor):
                self.send_json({"error": "Apenas usuário master pode alterar Client ID, Client Secret e Redirect URI."}, status=403)
                return
            settings = app_settings()
            current = settings.get("meli", {})
            client_secret = request.get("client_secret", "")
            settings["meli"] = {
                "client_id": request.get("client_id", "").strip(),
                "client_secret": client_secret.strip() if client_secret and client_secret != "********" else current.get("client_secret", ""),
                "redirect_uri": request.get("redirect_uri", "").strip(),
            }
            write_json("settings.json", settings)
            self.send_json(
                {
                    "ok": True,
                    "config": {
                        "client_id": settings["meli"]["client_id"],
                        "client_secret_set": bool(settings["meli"]["client_secret"]),
                        "redirect_uri": settings["meli"]["redirect_uri"],
                        "issues": oauth_issues(),
                    },
                }
            )
            return

        if parsed.path == "/api/notifications/config":
            actor = self.current_user(payload)
            notifications = user_notifications(payload, actor, create=True)
            if "telegram" in request:
                telegram_request = dict(request["telegram"])
                telegram_request.pop("status", None)
                if not telegram_request.get("bot_token") or telegram_request.get("bot_token") == "********":
                    telegram_request.pop("bot_token", None)
                notifications["telegram"] = {**notifications.get("telegram", {}), **telegram_request}
                telegram = notifications["telegram"]
                if telegram.get("enabled") and telegram.get("bot_token") and telegram.get("chat_id"):
                    telegram["status"] = "Telegram pronto para envio via Bot API"
                elif telegram.get("enabled"):
                    telegram["status"] = "Telegram ativado, mas falta token ou Chat ID"
                else:
                    telegram["status"] = "Telegram desativado"
            payload.setdefault("user_notifications", {})[actor["id"]] = notifications
            write_payload(payload)
            safe = json.loads(json.dumps(notifications))
            if safe.get("telegram", {}).get("bot_token"):
                safe["telegram"]["bot_token"] = "********"
            self.send_json({"ok": True, "notifications": safe})
            return

        if parsed.path == "/api/notifications/test":
            text = request.get("message") or "Teste de alerta do CompeTIDOR"
            channels = ["telegram"]
            results = {}
            actor = self.current_user(payload)
            notifier = Notifier(user_notifications(payload, actor, create=False))
            for channel in channels:
                try:
                    if channel == "telegram":
                        results[channel] = notifier.send_telegram(text)
                except Exception as exc:
                    results[channel] = {"ok": False, "error": str(exc)}
            self.send_json({"ok": True, "results": results})
            return

        if parsed.path == "/api/notifications/telegram/updates":
            try:
                actor = self.current_user(payload)
                notifier = Notifier(user_notifications(payload, actor, create=False))
                updates = notifier.telegram_updates()
                chats = []
                seen = set()
                for update in updates.get("result", []) or []:
                    message = update.get("message") or update.get("channel_post") or update.get("my_chat_member", {})
                    chat = message.get("chat") if isinstance(message, dict) else None
                    if not chat:
                        continue
                    chat_id = str(chat.get("id"))
                    if not chat_id or chat_id in seen:
                        continue
                    seen.add(chat_id)
                    chats.append(
                        {
                            "id": chat_id,
                            "type": chat.get("type", ""),
                            "title": chat.get("title") or " ".join([chat.get("first_name", ""), chat.get("last_name", "")]).strip() or chat.get("username", ""),
                            "username": chat.get("username", ""),
                        }
                    )
                self.send_json({"ok": True, "chats": chats, "raw_count": len(updates.get("result", []) or [])})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/scan/items":
            name = (request.get("name") or "").strip()
            url = (request.get("url") or "").strip()
            minimum_price = request.get("minimum_price") or 0
            item_id = extract_meli_item_id(url)
            if not name or not url or not item_id:
                self.send_json({"error": "Informe nome do produto e link ou código MLB válido para criar o scan."}, status=400)
                return
            scan = {
                "id": f"scan-{uuid.uuid4().hex[:8]}",
                "name": name,
                "url": url,
                "item_id": item_id,
                "minimum_price": float(minimum_price or 0),
                "history": [],
                "created_at": now_label(),
                "status": "active",
            }
            payload.setdefault("scan_items", []).insert(0, scan)
            write_payload(payload)
            self.send_json({"ok": True, "scan": scan, "scan_items": payload["scan_items"]}, status=201)
            return

        if parsed.path == "/api/scan/run":
            try:
                scan, entry = scan_meli_item(payload, request.get("id", ""))
                write_payload(payload)
                self.send_json({"ok": True, "scan": scan, "entry": entry, "scan_items": payload.get("scan_items", [])})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/scan/update":
            scan_id = request.get("id", "")
            scans = payload.setdefault("scan_items", [])
            scan = next((item for item in scans if item.get("id") == scan_id), None)
            if not scan:
                self.send_json({"error": "Produto de scan não encontrado."}, status=404)
                return
            try:
                minimum_price = float(request.get("minimum_price") or 0)
            except (TypeError, ValueError):
                self.send_json({"error": "Informe um preço mínimo válido."}, status=400)
                return
            scan["minimum_price"] = max(0, minimum_price)
            scan["last_alert_signature"] = ""
            scan["updated_at"] = now_label()
            write_payload(payload)
            self.send_json({"ok": True, "scan": scan, "scan_items": scans})
            return

        if parsed.path == "/api/competitors/scan":
            try:
                competitor = scan_competitor_profile(payload, request.get("seller_id", ""), request.get("limit", 50))
                write_payload(payload)
                self.send_json({"ok": True, "competitor": competitor, "competitors": payload.get("competitors", [])})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/statistics/query":
            try:
                job = start_statistics_job(request)
                self.send_json({"ok": True, **job}, status=202)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/reports/export":
            try:
                job = start_report_job(request)
                self.send_json({"ok": True, **job}, status=202)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/spreadsheet/template":
            try:
                body = build_bulk_spreadsheet(payload, request.get("filters") or {})
                stamp = datetime.now(APP_TZ).strftime("%Y%m%d-%H%M")
                self.send_bytes(body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"competidor-edicao-anuncios-{stamp}.xlsx")
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/spreadsheet/import":
            try:
                result = parse_bulk_spreadsheet(payload, request.get("file") or "", self.current_user(payload))
                self.send_json({"ok": True, **result})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/spreadsheet/apply":
            try:
                actor = self.current_user(payload)
                token = str(request.get("token") or "")
                with SPREADSHEET_JOBS_LOCK:
                    job = SPREADSHEET_JOBS.get(token)
                if not job or job.get("user_id") != (actor or {}).get("id") or time.monotonic() - job.get("created", 0) > 1800:
                    raise RuntimeError("A prévia da planilha expirou. Importe o arquivo novamente.")
                maximum = max(1, int(os.getenv("MELI_SPREADSHEET_MAX_CHANGES", "500")))
                if len(job.get("changes") or []) > maximum:
                    raise RuntimeError(f"Aplique no máximo {maximum} alterações por planilha para respeitar os limites da API.")
                results = []
                for row in job.get("changes") or []:
                    try:
                        results.append(apply_spreadsheet_change(payload, row, actor))
                    except Exception as exc:
                        results.append({"item_id": row.get("item_id"), "row": row.get("row"), "status": "error", "error": str(exc)})
                write_payload(payload)
                with SPREADSHEET_JOBS_LOCK:
                    SPREADSHEET_JOBS.pop(token, None)
                self.send_json(
                    {
                        "ok": not any(row.get("status") == "error" for row in results),
                        "updated": sum(row.get("status") == "updated" for row in results),
                        "failed": sum(row.get("status") == "error" for row in results),
                        "results": results,
                    }
                )
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/sync":
            try:
                raw_limit = request.get("limit", "all")
                limit = None if raw_limit in ("all", "", None) else max(1, int(raw_limit))
                if limit is None or limit > int(os.getenv("MELI_SYNC_INLINE_LIMIT", "200")):
                    result = enqueue_official_sync(request.get("account_id", ""), limit, "manual")
                    self.send_json({"ok": True, **result}, status=202)
                else:
                    result = sync_official_account(payload, request.get("account_id", ""), limit)
                    write_payload(payload)
                    self.send_json({"ok": True, **result})
            except Exception as exc:
                message = str(exc)
                if "PA_UNAUTHORIZED_RESULT_FROM_POLICIES" in message or "HTTP 403" in message:
                    message = (
                        "A conta OAuth está conectada, mas a aplicação ainda não tem permissão/política "
                        "para ler anúncios. Habilite as permissões de anúncios/vendas no painel do Mercado Livre "
                        "e refaça o login desta conta."
                    )
                self.send_json({"error": message}, status=400)
            return

        if parsed.path == "/api/meli/unlink":
            try:
                account = unlink_official_account(payload, request.get("account_id", ""))
                write_payload(payload, replace_collections={"accounts"})
                self.send_json({"ok": True, "account": account})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/shipping-costs/refresh":
            try:
                results = refresh_shipping_costs_for_items(payload, request.get("item_ids") or [])
                write_payload(payload)
                self.send_json({"ok": True, "items": results})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/identifiers/refresh":
            try:
                results = refresh_identifiers_for_items(payload, request.get("item_ids") or [])
                write_payload(payload)
                self.send_json({"ok": True, "items": results})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/update":
            try:
                user = self.current_user(payload)
                item_id = request.get("item_id", "")
                account_id = request.get("account_id", "")
                account = next((item for item in payload.get("accounts", []) if item.get("id") == account_id or item.get("nickname") == request.get("account")), None)
                if not account or not account.get("official"):
                    raise RuntimeError("Conta oficial não encontrada para atualizar o anúncio.")
                client = account_client(account)
                update = {}
                for key in ["price", "available_quantity", "title"]:
                    if key in request and request[key] not in ("", None):
                        update[key] = request[key]
                if request.get("status_action") == "pause":
                    update["status"] = "paused"
                if request.get("status_action") == "activate":
                    update["status"] = "active"
                dimension_attrs = []
                expected_package_values = {}
                dimension_map = {
                    "package_weight": "SELLER_PACKAGE_WEIGHT",
                    "package_height": "SELLER_PACKAGE_HEIGHT",
                    "package_width": "SELLER_PACKAGE_WIDTH",
                    "package_length": "SELLER_PACKAGE_LENGTH",
                }
                for request_key, attr_id in dimension_map.items():
                    if clean_attribute_value(request.get(request_key)):
                        api_value = seller_package_api_value(request_key, request.get(request_key))
                        expected_package_values[request_key] = api_value
                        dimension_attrs.append({"id": attr_id, "value_name": api_value})
                if dimension_attrs:
                    update["attributes"] = dimension_attrs
                expected_gtin_codes = []
                if clean_attribute_value(request.get("gtin")):
                    gtin_fragment, expected_gtin_codes = item_gtin_update_fragment(client, item_id, request.get("gtin"))
                    if gtin_fragment.get("attributes"):
                        update["attributes"] = [
                            *(update.get("attributes") or []),
                            *gtin_fragment["attributes"],
                        ]
                    if gtin_fragment.get("variations"):
                        update["variations"] = gtin_fragment["variations"]
                if not update:
                    raise RuntimeError("Nenhum campo informado para atualizar.")
                official = client.update_item(item_id, update)
                verified_item = verify_package_update(client, item_id, expected_package_values) if expected_package_values else {}
                if expected_gtin_codes:
                    verified_gtin_item = verify_gtin_update(client, item_id, expected_gtin_codes)
                    if not verified_item:
                        verified_item = verified_gtin_item
                verified_package_values = package_values_from_item(verified_item) if verified_item else {}
                stock_transition = None
                for item in payload.get("catalog", []):
                    if item.get("id") == item_id:
                        changes = {}
                        if "price" in update:
                            changes["price"] = {"from": item.get("price"), "to": update["price"]}
                            item["price"] = update["price"]
                        if "available_quantity" in update:
                            changes["stock"] = {"from": item.get("stock"), "to": update["available_quantity"]}
                            stock_transition = (int(item.get("stock") or 0), int(update["available_quantity"] or 0), item)
                            item["stock"] = update["available_quantity"]
                        if "title" in update:
                            changes["title"] = {"from": item.get("title"), "to": update["title"]}
                            item["title"] = update["title"]
                        for request_key in dimension_map:
                            if clean_attribute_value(request.get(request_key)):
                                normalized_value = verified_package_values.get(request_key) or normalize_package_value(
                                    expected_package_values[request_key], [dimension_map[request_key]]
                                )
                                changes[request_key] = {"from": item.get(request_key), "to": normalized_value}
                                item[request_key] = normalized_value
                        if expected_gtin_codes:
                            new_gtin = ", ".join(expected_gtin_codes)
                            changes["gtin"] = {"from": item.get("gtin") or "", "to": new_gtin}
                            item["gtin"] = new_gtin
                        if update.get("status") == "paused":
                            changes["status"] = {"from": item.get("meli_status"), "to": "paused"}
                            item["status"] = "paused"
                            item["meli_status"] = "paused"
                        if update.get("status") == "active":
                            changes["status"] = {"from": item.get("meli_status"), "to": "active"}
                            item["status"] = "sharing" if is_catalog_listing(item) else "winning"
                            item["meli_status"] = "active"
                        item["updated_at"] = now_label()
                        append_item_log(payload, item, user, "Atualização manual", changes)
                if stock_transition and stock_transition[0] > 0 and stock_transition[1] == 0:
                    old_stock, new_stock, changed_item = stock_transition
                    alert_id = f"stock-{account.get('id')}-{item_id}-{uuid.uuid4().hex[:8]}"
                    alert = stock_alert(alert_id, account, changed_item)
                    payload.setdefault("alerts", []).insert(0, alert)
                    notify_alert(payload, alert)
                write_payload(payload)
                self.send_json({"ok": True, "official": official, "catalog": public_payload(payload, user)["catalog"]})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/clone-source":
            try:
                request_copy = json.loads(json.dumps(request, ensure_ascii=False))
                operation = start_async_operation(
                    "clone_source",
                    lambda: clone_source_snapshot(
                        read_payload(include_catalog=False),
                        request_copy.get("account_id") or request_copy.get("account") or "",
                        request_copy.get("item_id") or "",
                    ),
                    "Carregando EAN/GTIN, descrição e atributos oficiais.",
                )
                self.send_json({"ok": True, **operation}, status=202)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/description":
            try:
                request_copy = json.loads(json.dumps(request, ensure_ascii=False))
                actor = self.current_user(payload)
                action_label = "Salvando descrição oficial." if request_copy.get("action") == "update" else "Carregando descrição oficial."
                operation = start_async_operation(
                    "item_description",
                    lambda: item_description_operation(request_copy, actor),
                    action_label,
                )
                self.send_json({"ok": True, **operation}, status=202)
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/win_catalog":
            try:
                user = self.current_user(payload)
                item_id = request.get("item_id", "")
                item = next((row for row in payload.get("catalog", []) if row.get("id") == item_id), None)
                if not item:
                    raise RuntimeError("Anúncio não encontrado.")
                if not item.get("price_to_win"):
                    raise RuntimeError("O Mercado Livre não retornou preço sugerido para ganhar este catálogo.")
                account = next((row for row in payload.get("accounts", []) if row.get("id") == item.get("account_id")), None)
                if not account:
                    raise RuntimeError("Conta oficial não encontrada.")
                official = account_client(account).update_item(item_id, {"price": item["price_to_win"]})
                old_price = item.get("price")
                item["price"] = item["price_to_win"]
                item["updated_at"] = now_label()
                append_item_log(payload, item, user, "Ganhar catálogo", {"price": {"from": old_price, "to": item["price_to_win"]}})
                write_payload(payload)
                self.send_json({"ok": True, "official": official, "item": item})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/remove_flex":
            try:
                user = self.current_user(payload)
                item_id = request.get("item_id", "")
                account_id = request.get("account_id", "")
                item = next((row for row in payload.get("catalog", []) if row.get("id") == item_id), None)
                if not item:
                    raise RuntimeError("Anúncio não encontrado.")
                account = next(
                    (row for row in payload.get("accounts", []) if row.get("id") == account_id or row.get("id") == item.get("account_id")),
                    None,
                )
                if not account or not account.get("official"):
                    raise RuntimeError("Conta oficial não encontrada para remover o Mercado Envios Flex.")
                update = {"shipping": {"mode": item.get("shipping_mode") or "me2", "logistic_type": "drop_off"}}
                official = account_client(account).update_item(item_id, update)
                old_logistic_type = item.get("shipping_logistic_type") or ""
                item["shipping_mode"] = update["shipping"]["mode"]
                item["shipping_logistic_type"] = update["shipping"]["logistic_type"]
                item["updated_at"] = now_label()
                append_item_log(
                    payload,
                    item,
                    user,
                    "Remoção Mercado Envios Flex",
                    {"shipping_logistic_type": {"from": old_logistic_type, "to": "drop_off"}},
                )
                write_payload(payload)
                self.send_json({"ok": True, "official": official, "item": item})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/meli/item/activate_flex":
            try:
                user = self.current_user(payload)
                item_id = request.get("item_id", "")
                account_id = request.get("account_id", "")
                item = next((row for row in payload.get("catalog", []) if row.get("id") == item_id), None)
                if not item:
                    raise RuntimeError("Anúncio não encontrado.")
                account = next(
                    (row for row in payload.get("accounts", []) if row.get("id") == account_id or row.get("id") == item.get("account_id")),
                    None,
                )
                if not account or not account.get("official"):
                    raise RuntimeError("Conta oficial não encontrada para ativar o Mercado Envios Flex.")
                update = {"shipping": {"mode": item.get("shipping_mode") or "me2", "logistic_type": "self_service"}}
                official = account_client(account).update_item(item_id, update)
                old_logistic_type = item.get("shipping_logistic_type") or ""
                item["shipping_mode"] = update["shipping"]["mode"]
                item["shipping_logistic_type"] = update["shipping"]["logistic_type"]
                item["updated_at"] = now_label()
                append_item_log(
                    payload,
                    item,
                    user,
                    "Ativação Mercado Envios Flex",
                    {"shipping_logistic_type": {"from": old_logistic_type, "to": "self_service"}},
                )
                write_payload(payload)
                self.send_json({"ok": True, "official": official, "item": item})
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=400)
            return

        if parsed.path == "/api/clone/preview":
            try:
                request_copy = json.loads(json.dumps(request, ensure_ascii=False))
                operation = start_async_operation(
                    "clone_preview",
                    lambda: prepare_clone_preview(request_copy),
                    "Preparando e validando a cópia em segundo plano.",
                )
                self.send_json({"ok": True, **operation}, status=202)
            except Exception as exc:
                self.send_json({"error": f"Não foi possível gerar o preview: {exc}"}, status=400)
            return

        if parsed.path == "/api/clone/execute-batch":
            try:
                request_copy = json.loads(json.dumps(request, ensure_ascii=False))
                operation = start_async_operation(
                    "clone_execute_batch",
                    lambda: execute_clone_batch_request(request_copy),
                    "Cópia adicionada à fila do Mercado Livre.",
                )
                self.send_json({"ok": True, **operation}, status=202)
            except Exception as exc:
                self.send_json({"error": f"Não foi possível executar o lote: {exc}"}, status=400)
            return

        if parsed.path == "/api/clone/execute":
            try:
                request_copy = json.loads(json.dumps(request, ensure_ascii=False))
                operation = start_async_operation(
                    "clone_execute",
                    lambda: execute_clone_request(request_copy),
                    "Cópia adicionada à fila do Mercado Livre.",
                )
                self.send_json({"ok": True, **operation}, status=202)
            except Exception as exc:
                self.send_json({"error": f"Não foi possível copiar o anúncio: {exc}"}, status=400)
            return

        self.send_json({"error": "Endpoint não encontrado"}, status=404)


if __name__ == "__main__":
    port = int(os.getenv("PORT", "8765"))
    http_server = ThreadingHTTPServer(("127.0.0.1", port), App)

    cert_file, key_file = ensure_dev_certificate()
    tls_context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
    tls_context.load_cert_chain(certfile=str(cert_file), keyfile=str(key_file))
    https_server = ThreadingHTTPServer(("127.0.0.1", https_port()), App)
    https_server.socket = tls_context.wrap_socket(https_server.socket, server_side=True)

    threading.Thread(target=https_server.serve_forever, daemon=True).start()
    threading.Thread(target=meli_notification_loop, daemon=True).start()
    threading.Thread(target=resume_pending_official_syncs, daemon=True).start()
    threading.Thread(target=auto_scan_loop, daemon=True).start()
    threading.Thread(target=auto_official_sync_loop, daemon=True).start()
    threading.Thread(target=auto_catalog_competition_loop, daemon=True).start()
    print(f"CompeTIDOR rodando em http://127.0.0.1:{port}")
    print(f"Callback HTTPS em https://127.0.0.1:{https_port()}/api/oauth/callback")
    http_server.serve_forever()




